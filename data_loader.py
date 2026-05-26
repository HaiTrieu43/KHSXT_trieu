"""
data_loader.py - Module đọc TẤT CẢ file Excel đầu vào cho hệ thống KHSX

Đọc dữ liệu từ:
  1. Sale Forecast (tuần)
  2. Kế hoạch SILO (xe bồn)
  3. Kế hoạch Ba Cang (đại lý)
  4. FFSTOCK (tồn kho thành phẩm)
  5. Tồn bồn (cám trong bồn)
  6. Empty Bag (bao bì rỗng)
  7. Công suất (Plan.xlsm)
  8. Feedcode (KHSX → line máy)
  9. Kháng sinh (KHSX → mã KS/HC)
 10. KHSX hôm qua (thực hiện ngày trước)

Tất cả đều dùng openpyxl với data_only=True, read_only=True.
"""

import os
import re
import glob
import traceback
from typing import Optional
from openpyxl import load_workbook

from models import ProductSpec, ForecastItem


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def _safe_float(val, default=0.0) -> float:
    """Chuyển giá trị sang float an toàn. Trả về default nếu không hợp lệ."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=0) -> int:
    """Chuyển giá trị sang int an toàn. Trả về default nếu không hợp lệ."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _safe_str(val, default='') -> str:
    """Chuyển giá trị sang string, strip khoảng trắng và uppercase."""
    if val is None:
        return default
    s = str(val).strip()
    return s.upper() if s else default


def _normalize_product_code(val) -> str:
    """Chuẩn hóa mã sản phẩm: uppercase, bỏ khoảng trắng, bỏ annotation ngoặc.

    Quy tắc xử lý ngoặc trong mã từ Sale Forecast:
      - (SILO) / (silo) ở cuối → bỏ hoàn toàn (chỉ là ghi chú silo)
      - (5*5) / (NxM) ở cuối  → bỏ hoàn toàn (ghi chú quy cách đóng bao)
      - (TH) / (th) ở giữa   → giữ lại TH, bỏ ngoặc (TH = Thái Hòa, phần tên)
    Ví dụ:
      '551(5*5)'      → '551'
      '553(silo)'     → '553'
      '551F(SILO)'    → '551F'
      '005(TH)S'      → '005THS'
      '005 (th)Q'     → '005THQ'
    """
    if val is None:
        return ''
    s = str(val).strip().upper()
    # Loại bỏ khoảng trắng
    s = s.replace(' ', '')
    # Bước 1: Chuyển (TH) thành TH (phần tên thực sự, không phải annotation)
    s = s.replace('(TH)', 'TH')
    # Bước 2: Bỏ tất cả phần ngoặc còn lại: (SILO), (5*5), v.v.
    #   Regex: loại bỏ chuỗi dạng (...) bất kỳ còn lại
    s = re.sub(r'\([^)]*\)', '', s)
    # Bước 3: Xóa khoảng trắng thừa sau khi xử lý
    s = s.strip()
    return s


def normalize_ks_code(val) -> str:
    """Chuẩn hóa mã kháng sinh sang format KS/ HC (mã1)/(mã2)"""
    if val is None:
        return ''
    s = str(val).strip()
    if not s or s.upper() in {'#N/A', 'SILO', 'NONE', 'NULL'}:
        return ''
        
    # Loại bỏ tiền tố KS/HC (nếu có) để chuẩn hóa phần sau
    s_clean = re.sub(r'^(?:KS\s*/?\s*HC\s*\(?|KS\s*/?\s*HC\s*|KS\s*HC\s*\(?|\(?)(.*)', r'\1', s, flags=re.IGNORECASE)
    
    # Nếu s_clean bắt đầu bằng '(', giữ nguyên, ngược lại thêm '('
    if not s_clean.startswith('('):
        s_clean = '(' + s_clean
        
    # Sửa dấu gạch chéo kép "//" thành "/"
    s_clean = s_clean.replace('//', '/')
    
    # Chuẩn hóa khoảng trắng quanh các ký tự phân tách
    s_clean = s_clean.replace(' )', ')').replace('( ', '(')
    
    return f"KS/ HC {s_clean}"


def _find_latest_file(directory: str, pattern: str) -> Optional[str]:
    """
    Tìm file mới nhất trong thư mục theo glob pattern.
    Bỏ qua các file tạm thời bắt đầu bằng '~$'.
    Ưu tiên trích xuất và so sánh ngày tháng từ tên file (dạng DD-MM-YYYY).
    Nếu không có hoặc bằng nhau, sắp xếp theo thời gian chỉnh sửa file (mtime).
    """
    import re
    from datetime import datetime

    if not os.path.isdir(directory):
        print(f"  ⚠️  Thư mục không tồn tại: {directory}")
        return None

    # Bước 1: Tìm ở cấp gốc (nhanh)
    search_path = os.path.join(directory, pattern)
    files = glob.glob(search_path)

    # Bước 2: Nếu không tìm thấy, tìm đệ quy trong thư mục con
    if not files:
        search_path_recursive = os.path.join(directory, '**', pattern)
        files = glob.glob(search_path_recursive, recursive=True)

    # Loại bỏ các file tạm thời (Excel temporary files bắt đầu bằng ~$)
    files = [f for f in files if not os.path.basename(f).startswith('~$')]

    if not files:
        print(f"  ⚠️  Không tìm thấy file: {pattern} trong {directory}")
        return None

    def _extract_date_from_filename(filepath: str) -> Optional[datetime]:
        basename = os.path.basename(filepath)
        # Tìm kiếm chuỗi có dạng DD-MM-YYYY (2 chữ số - 2 chữ số - 4 chữ số)
        match = re.search(r'(\d{2})[-_\s]+(\d{2})[-_\s]+(\d{4})', basename)
        if match:
            day, month, year = map(int, match.groups())
            try:
                return datetime(year, month, day)
            except ValueError:
                pass
        return None

    def sort_key(filepath: str):
        parsed_date = _extract_date_from_filename(filepath)
        has_date = 1 if parsed_date is not None else 0
        date_val = parsed_date if parsed_date is not None else datetime.min
        mtime = os.path.getmtime(filepath)
        return (has_date, date_val, mtime)

    # Lấy file mới nhất dựa trên sort_key (ưu tiên ngày trong tên file, sau đó là mtime)
    latest = max(files, key=sort_key)
    return latest



def get_file_info(directory, pattern, exact_path=None):
    """Lấy thông tin của file mới nhất theo pattern hoặc đường dẫn chính xác.
    
    Trả về dict với keys: exists, filename, last_modified, size, path
    """
    import datetime
    
    if exact_path:
        path = exact_path
    else:
        path = _find_latest_file(directory, pattern)
        
    if not path or not os.path.isfile(path):
        return {
            'exists': False,
            'filename': 'Không tìm thấy',
            'last_modified': '-',
            'size': '-',
            'path': ''
        }
        
    stat = os.stat(path)
    mtime = datetime.datetime.fromtimestamp(stat.st_mtime)
    size_kb = round(stat.st_size / 1024, 1)
    
    return {
        'exists': True,
        'filename': os.path.basename(path),
        'last_modified': mtime.strftime('%d-%m-%Y %H:%M:%S'),
        'size': f"{size_kb} KB",
        'path': path
    }


def _open_workbook(file_path: str, data_only=True, read_only=True):
    """
    Mở workbook Excel an toàn với xử lý encoding.
    Trả về workbook hoặc None nếu lỗi.
    """
    if not os.path.isfile(file_path):
        print(f"  ❌ File không tồn tại: {file_path}")
        return None
    try:
        wb = load_workbook(file_path, data_only=data_only, read_only=read_only)
        return wb
    except Exception as e:
        print(f"  ❌ Lỗi mở file {os.path.basename(file_path)}: {e}")
        return None


def _get_sheet(wb, sheet_name=None):
    """
    Lấy sheet theo tên. Nếu không chỉ định, trả về sheet cuối.
    Trả về (sheet, tên_sheet) hoặc (None, None).
    """
    if sheet_name:
        # Tìm sheet chứa tên (case-insensitive)
        for name in wb.sheetnames:
            if sheet_name.upper() in name.upper():
                return wb[name], name
        # Tìm chính xác
        if sheet_name in wb.sheetnames:
            return wb[sheet_name], sheet_name
        print(f"  ⚠️  Không tìm thấy sheet: {sheet_name}")
        return None, None
    else:
        # Trả về sheet cuối cùng
        name = wb.sheetnames[-1]
        return wb[name], name


# Các từ khóa dòng tổng cần bỏ qua trong Forecast
_SKIP_KEYWORDS = {
    'PRE-STARTER', 'PRE STARTER', 'PRESTARTER',
    'FATTENNING', 'FATTENING',
    'BREEDER',
    'TOTAL', 'GRAND TOTAL',
    'BROILER',
    'LOCAL',
    'LAYER',
    'DUCK',
    'SHRIMP',
    'FISH',
    'PET',
    'HOG',
    'QUAIL',
    'COW',
    'GOAT',
    'LABORATORY',
}


# ============================================================
# 1. LOAD FORECAST
# ============================================================

def load_forecast(file_path: str, week_sheet: str = None) -> list:
    """
    Đọc Sale Forecast Excel → list[ForecastItem]

    File có 21 sheets (W1-W21). Nếu week_sheet=None, dùng sheet CUỐI CÙNG.
    Dữ liệu bắt đầu từ dòng 10, tối đa 250 dòng, 31 cột (A-AE).

    Args:
        file_path: Đường dẫn file Forecast
        week_sheet: Tên sheet (VD: 'W21'). None = sheet cuối.

    Returns:
        list[ForecastItem] - Danh sách nhu cầu forecast
    """
    print(f"📊 Đọc FORECAST: {os.path.basename(file_path)}")
    results = []

    wb = _open_workbook(file_path)
    if wb is None:
        return results

    ws, sheet_name = _get_sheet(wb, week_sheet)
    if ws is None:
        wb.close()
        return results

    print(f"  📋 Sheet: {sheet_name}")

    row_count = 0
    for row in ws.iter_rows(min_row=10, max_row=350, max_col=31):
        row_count += 1
        if row_count > 340:
            break

        # Lấy giá trị các cột (0-indexed từ tuple)
        formular_code = _safe_str(row[0].value)  # A - FORMULAR CODE

        # Bỏ qua dòng trống
        if not formular_code:
            continue

        # Bỏ qua dòng tổng/phân loại
        if any(kw in formular_code for kw in _SKIP_KEYWORDS):
            continue

        # DIE size - có thể là số (2.5, 2.8, 4.0) hoặc chữ (M, MC, BC, SC, C)
        die_raw = row[1].value  # B
        try:
            die_size = float(die_raw) if die_raw is not None else 0.0
        except (ValueError, TypeError):
            die_size = 0.0  # Chuỗi M, MC, BC, SC, C → die_size = 0

        packing_size = _safe_str(row[2].value)  # C - PACKING SIZE

        # Feed codes theo thương hiệu (cột D-I) — chuẩn hóa bỏ annotation ngoặc
        fc_higro = _normalize_product_code(row[3].value)   # D
        fc_cp = _normalize_product_code(row[4].value)      # E
        fc_star = _normalize_product_code(row[5].value)    # F
        fc_nuvo = _normalize_product_code(row[6].value)    # G
        fc_nasa = _normalize_product_code(row[7].value)    # H
        fc_farm = _normalize_product_code(row[8].value)    # I

        # Product code = mã đầu tiên không rỗng theo thứ tự D, I, E, F, G, H
        product_code = (fc_higro or fc_farm or fc_cp or fc_star
                        or fc_nuvo or fc_nasa)
        product_code = _normalize_product_code(product_code)

        if not product_code:
            continue

        # Tấn - kênh Dealer (J-O)
        dealer_higro = _safe_float(row[9].value)     # J
        dealer_cp = _safe_float(row[10].value)       # K
        dealer_star = _safe_float(row[11].value)     # L
        dealer_nuvo = _safe_float(row[12].value)     # M
        dealer_nasa = _safe_float(row[13].value)     # N
        dealer_total = _safe_float(row[14].value)    # O

        # Tấn - kênh Farm (P-R)
        farm_swine = _safe_float(row[15].value)      # P
        farm_integrate = _safe_float(row[16].value)  # Q
        farm_total = _safe_float(row[17].value)      # R

        # Tổng (S-U)
        grand_total_tons = _safe_float(row[18].value)  # S
        silo_tons = _safe_float(row[19].value)         # T
        total_with_silo = _safe_float(row[20].value)   # U

        # Bỏ qua dòng không có sản lượng
        if total_with_silo == 0 and grand_total_tons == 0 and silo_tons == 0:
            continue

        # Bao bì (X-AE, index 23-30)
        bag_higro = _safe_int(row[23].value)          # X
        bag_cp = _safe_int(row[24].value)             # Y
        bag_star = _safe_int(row[25].value)           # Z
        bag_nuvo = _safe_int(row[26].value)           # AA
        bag_nasa = _safe_int(row[27].value)           # AB
        bag_dealer_total = _safe_int(row[28].value)   # AC
        bag_farm = _safe_int(row[29].value)           # AD
        bag_grand_total = _safe_int(row[30].value)    # AE

        item = ForecastItem(
            product_code=product_code,
            packing_size=packing_size,
            die_size=die_size,
            # Dealer
            dealer_higro=dealer_higro,
            dealer_cp=dealer_cp,
            dealer_star=dealer_star,
            dealer_nuvo=dealer_nuvo,
            dealer_nasa=dealer_nasa,
            dealer_total=dealer_total,
            # Farm
            farm_swine=farm_swine,
            farm_integrate=farm_integrate,
            farm_total=farm_total,
            # Tổng
            grand_total_tons=grand_total_tons,
            silo_tons=silo_tons,
            total_with_silo=total_with_silo,
            # Bao
            bag_higro=bag_higro,
            bag_cp=bag_cp,
            bag_star=bag_star,
            bag_nuvo=bag_nuvo,
            bag_nasa=bag_nasa,
            bag_dealer_total=bag_dealer_total,
            bag_farm=bag_farm,
            bag_grand_total=bag_grand_total,
            # Feed codes
            feed_code_higro=fc_higro,
            feed_code_cp=fc_cp,
            feed_code_star=fc_star,
            feed_code_nuvo=fc_nuvo,
            feed_code_nasa=fc_nasa,
            feed_code_farm=fc_farm,
        )
        results.append(item)

    wb.close()
    print(f"  ✅ Đọc được {len(results)} sản phẩm từ Forecast")
    return results


# ============================================================
# 2. LOAD SILO PLAN
# ============================================================

def load_silo_plan(file_path: str, week_sheet: str = None) -> dict:
    """
    Đọc Kế hoạch SILO → dict {day_index(1-6) → {product_code → tons}}

    Cột A = tên sản phẩm, các cột sau = sản lượng từng ngày (Thứ 2-7).
    Đơn vị file là KG → chuyển sang TẤN.
    Mỗi sản phẩm có 2 dòng (PLAN/ACTUAL) - chỉ lấy dòng PLAN (dòng lẻ).

    Args:
        file_path: Đường dẫn file SILO
        week_sheet: Tên sheet (None = sheet cuối)

    Returns:
        dict {1: {'552SF': 200.0, ...}, 2: {...}, ..., 6: {...}}
    """
    print(f"🚛 Đọc SILO PLAN: {os.path.basename(file_path)}")
    result = {d: {} for d in range(1, 7)}

    wb = _open_workbook(file_path)
    if wb is None:
        return result

    ws, sheet_name = _get_sheet(wb, week_sheet)
    if ws is None:
        wb.close()
        return result

    print(f"  📋 Sheet: {sheet_name}")

    # Từ khóa cần bỏ qua trong tên sản phẩm
    skip_kw = {'ACTUAL', 'TOTAL', 'SỐ LƯỢNG', 'PLAN', 'FARM', 'DEALER',
               'ĐẠI LÝ', 'TỔNG', 'GHI CHÚ', 'REMARK', 'NOTE'}

    for row in ws.iter_rows(min_row=3, max_row=100, max_col=10):
        val_str = _safe_str(row[0].value)
        if not val_str:
            continue

        # Bỏ qua dòng tổng, dòng ACTUAL, dòng header
        if any(kw in val_str for kw in skip_kw):
            continue

        product_raw = _normalize_product_code(row[0].value)

        # Đọc sản lượng 6 ngày (cột C-H hoặc C-I tuỳ cấu trúc)
        # Thường cột B = pellet size, cột C trở đi = ngày
        # Tìm cột bắt đầu chứa số liệu ngày
        day_start_col = 2  # Cột C (index 2) = ngày đầu tiên (sau pellet size)
        for day_idx in range(1, 7):
            col_idx = day_start_col + day_idx - 1
            if col_idx >= len(row):
                break

            val = _safe_float(row[col_idx].value)
            if val <= 0:
                continue

            # Đơn vị file là KG → chuyển tấn
            tons = val / 1000.0

            # Cộng dồn nếu cùng sản phẩm
            if product_raw in result[day_idx]:
                result[day_idx][product_raw] += tons
            else:
                result[day_idx][product_raw] = tons

    wb.close()

    total_products = len(set(
        p for day_data in result.values() for p in day_data
    ))
    print(f"  ✅ Đọc SILO: {total_products} sản phẩm, 6 ngày")
    return result


# ============================================================
# 3. LOAD BA CANG
# ============================================================

def load_bacang(file_path: str, week_sheet: str = None) -> dict:
    """
    Đọc Kế hoạch Ba Cang → dict {day_index(1-6) → {product → tons}}

    Ba Cang = đại lý lớn Võ Bá Cang, đặt hàng theo tuần.
    Nếu giá trị > 100 → đơn vị BAO (nhân 25/1000 → tấn).
    Nếu giá trị ≤ 100 → đã là tấn.

    Args:
        file_path: Đường dẫn file Ba Cang
        week_sheet: Tên sheet (None = sheet đầu tiên / mới nhất)

    Returns:
        dict {1: {'552S': 3.0, ...}, 2: {...}, ..., 6: {...}}
    """
    print(f"📦 Đọc BA CANG: {os.path.basename(file_path)}")
    result = {d: {} for d in range(1, 7)}

    wb = _open_workbook(file_path)
    if wb is None:
        return result

    ws, sheet_name = _get_sheet(wb, week_sheet)
    if ws is None:
        wb.close()
        return result

    print(f"  📋 Sheet: {sheet_name}")

    skip_kw = {'TOTAL', 'TỔNG', 'TON', 'TẤN', 'SILO', 'ACTUAL',
               'DỰ KIẾN', 'LỊCH', 'THỨ', 'NGÀY'}

    for row in ws.iter_rows(min_row=3, max_row=50, max_col=9):
        val_str = _safe_str(row[0].value)
        if not val_str:
            continue

        if any(kw in val_str for kw in skip_kw):
            continue

        product_raw = _normalize_product_code(row[0].value)

        # Cột B-G (index 1-6) = Thứ 2 → Thứ 7
        for day_idx in range(1, 7):
            if day_idx >= len(row):
                break

            val = _safe_float(row[day_idx].value)
            if val <= 0:
                continue

            # Nếu > 100 → đơn vị BAO (bao 25kg)
            if val > 100:
                tons = val * 25.0 / 1000.0
            else:
                tons = val

            if product_raw in result[day_idx]:
                result[day_idx][product_raw] += tons
            else:
                result[day_idx][product_raw] = tons

    wb.close()

    total = sum(
        sum(day_data.values()) for day_data in result.values()
    )
    print(f"  ✅ Đọc BA CANG: tổng {total:.1f} tấn/tuần")
    return result


# ============================================================
# 4. LOAD FFSTOCK
# ============================================================

def load_ffstock(file_path: str) -> dict:
    """
    Đọc FFSTOCK (tồn kho thành phẩm) → dict {product → tons_in_stock}

    File FFSTOCK có nhiều sheet (PRO-NASA, BRAN, INTGRATE, REMIX).
    Tìm sản phẩm ở cột B (FEED CODE), tồn kho ở cột M (BALANCE bao)
    hoặc cột N (BALANCE kg).

    Args:
        file_path: Đường dẫn file FFSTOCK

    Returns:
        dict {'552SF': 135.0, '550S': 25.6, ...}  (tấn)
    """
    print(f"📦 Đọc FFSTOCK: {os.path.basename(file_path)}")
    result = {}

    wb = _open_workbook(file_path)
    if wb is None:
        return result

    # Các sheet chính chứa dữ liệu tồn kho
    target_sheets = ['PRO-NASA', 'BRAN', 'INTGRATE', 'REMIX']

    for sname in wb.sheetnames:
        # Chỉ đọc các sheet liên quan
        sname_upper = sname.upper()
        is_target = any(t in sname_upper for t in
                        ['PRO', 'NASA', 'BRAN', 'INTGRATE', 'INTEGRATE', 'REMIX'])
        if not is_target:
            continue

        ws = wb[sname]

        for row in ws.iter_rows(min_row=4, max_row=400, max_col=20):
            # Cột B (index 1) = FEED CODE
            if len(row) < 14:
                continue

            val_str = _safe_str(row[1].value)
            if not val_str:
                continue

            # Bỏ qua dòng tổng/header
            if any(kw in val_str for kw in {'TOTAL', 'STT', 'FEED', 'GRAND',
                                            'HI-GRO', 'BRAND', 'FARM'}):
                continue

            product = _normalize_product_code(row[1].value)

            # Cột N (index 13) = BALANCE (kg)
            balance_kg = _safe_float(row[13].value)

            # Nếu cột N trống, thử cột M (index 12) = BALANCE (bao)
            if balance_kg == 0:
                balance_bags = _safe_float(row[12].value)
                # Xác định pack size từ cột D (index 3)
                pack = _safe_str(row[3].value) if len(row) > 3 else '25'
                pack_kg = 50 if '50' in pack else 25
                balance_kg = balance_bags * pack_kg

            if balance_kg <= 0:
                continue

            # Chuyển kg → tấn
            tons = balance_kg / 1000.0

            # Cộng dồn nếu sản phẩm xuất hiện ở nhiều sheet/dòng
            if product in result:
                result[product] += tons
            else:
                result[product] = tons

    wb.close()
    print(f"  ✅ Đọc FFSTOCK: {len(result)} sản phẩm, "
          f"tổng {sum(result.values()):.1f} tấn")
    return result


def load_ffstock_details(file_path: str) -> dict:
    """
    Đọc chi tiết DOH, Sales Average và Tồn kho từ FFSTOCK.
    Sử dụng thuật toán dò tìm cột động (Dynamic Header Search) tối ưu hóa hiệu năng.
    
    Returns:
        dict: {
            '552SF': {
                'stock_tons': float,
                'sales_avg_kg': float,
                'doh': float
            },
            ...
        }
    """
    print(f"📦 Đọc chi tiết FFSTOCK (DOH): {os.path.basename(file_path)}")
    raw_data = {} # {product: {'balance_kg': 0.0, 'sales_avg_kg': 0.0}}

    wb = _open_workbook(file_path)
    if wb is None:
        return {}

    for sname in wb.sheetnames:
        sname_upper = sname.upper()
        is_target = any(t in sname_upper for t in ['PRO', 'NASA', 'BRAN', 'INTGRATE', 'INTEGRATE'])
        if not is_target:
            continue

        ws = wb[sname]
        
        # Đọc toàn bộ 400 dòng và 30 cột đầu tiên bằng iter_rows (values_only=True)
        # Đây là cách tối ưu hóa hiệu năng cực cao cho openpyxl read_only mode
        rows_data = list(ws.iter_rows(min_row=1, max_row=400, min_col=1, max_col=30, values_only=True))
        if not rows_data:
            continue
            
        # Lấy 15 dòng đầu làm headers để dò cột
        headers = rows_data[:15]
            
        feed_code_col = None
        balance_kg_col = None
        balance_bag_col = None
        sales_avg_col = None
        doh_col = None
        
        # Duyệt qua các cột (0-indexed)
        for col_idx in range(min(30, len(headers[0]))):
            col_vals = []
            for r_idx in range(min(15, len(headers))):
                val = headers[r_idx][col_idx]
                if val is not None:
                    col_vals.append(str(val).strip().upper())
            
            # 1. Mã cám
            if any("FEED CODE" in v or "MÃ CÁM" in v or "TEN FEED" in v for v in col_vals):
                feed_code_col = col_idx
            
            # 2. Tồn kho Kg
            is_balance = any("BALANCE" in v or "TỒN CUỐI" in v or "TON CUOI" in v for v in col_vals)
            if is_balance:
                has_kg = any("KG" in v or "KÝ" in v or "KILO" in v for v in col_vals)
                if has_kg:
                    balance_kg_col = col_idx
                else:
                    balance_bag_col = col_idx
            
            # 3. TB bán
            if any("SALES AVERAGE" in v or "TB BÁN" in v or "BÁN TB" in v or "TB BAN" in v or "AVERAGE" in v for v in col_vals):
                sales_avg_col = col_idx
                
            # 4. DOH
            if any("DAY ON HAND" in v or "DOH" in v or "SỐ NGÀY TỒN" in v or "SO NGAY TON" in v or "NGÀY TỒN" in v or "ON HAND" in v for v in col_vals):
                doh_col = col_idx

        # Fallback balance_kg_col nếu chỉ tìm thấy balance_bag_col
        if balance_kg_col is None and balance_bag_col is not None:
            balance_kg_col = balance_bag_col + 1

        # Nếu thiếu cột quan trọng, bỏ qua sheet này
        if feed_code_col is None or balance_kg_col is None or sales_avg_col is None or doh_col is None:
            print(f"  ⚠️ Bỏ qua sheet {sname} do không tìm đủ các cột tiêu đề bắt buộc.")
            continue

        # Tìm dòng bắt đầu dữ liệu thực tế (dòng đầu tiên có mã cám thực tế)
        data_start_row = 12 # Fallback mặc định
        for r in range(4, min(25, len(rows_data))):
            val = rows_data[r - 1][feed_code_col]
            if val is not None:
                val_str = str(val).strip().upper()
                if val_str and not any(kw in val_str for kw in ['TOTAL', 'STT', 'FEED', 'GRAND', 'HI-GRO', 'BRAND', 'FARM', 'CODE', 'TÊN']):
                    data_start_row = r
                    break

        # Đọc dữ liệu từ data_start_row đến hết dòng 400 (hoặc độ dài rows_data)
        for r in range(data_start_row, len(rows_data) + 1):
            row_cells = rows_data[r - 1]
            if feed_code_col >= len(row_cells):
                continue
                
            feed_val = row_cells[feed_code_col]
            if feed_val is None:
                continue
                
            val_str = str(feed_val).strip().upper()
            if not val_str or any(kw in val_str for kw in ['TOTAL', 'STT', 'FEED', 'GRAND', 'HI-GRO', 'BRAND', 'FARM', 'CODE', 'TÊN']):
                continue
                
            product = _normalize_product_code(feed_val)
            if not product:
                continue
                
            # Đọc các giá trị số an toàn từ hàng
            bal_val = row_cells[balance_kg_col] if balance_kg_col < len(row_cells) else None
            avg_val = row_cells[sales_avg_col] if sales_avg_col < len(row_cells) else None
            
            balance_kg = _safe_float(bal_val)
            sales_avg_kg = _safe_float(avg_val)
            
            # Lưu vào raw_data và cộng dồn
            if product not in raw_data:
                raw_data[product] = {'balance_kg': 0.0, 'sales_avg_kg': 0.0}
                
            raw_data[product]['balance_kg'] += balance_kg
            raw_data[product]['sales_avg_kg'] += sales_avg_kg

    wb.close()

    # Tính toán lại DOH cuối cùng và chuẩn hóa kết quả đầu ra
    result = {}
    for product, vals in raw_data.items():
        balance_kg = vals['balance_kg']
        sales_avg_kg = vals['sales_avg_kg']
        
        stock_tons = balance_kg / 1000.0
        
        if sales_avg_kg > 0:
            doh = balance_kg / sales_avg_kg
        else:
            doh = 99.0
            
        result[product] = {
            'stock_tons': stock_tons,
            'sales_avg_kg': sales_avg_kg,
            'doh': doh
        }

    # In ra một số mẫu DOH thấp để kiểm tra
    low_doh_items = {p: v for p, v in result.items() if v['doh'] < 3.0 and v['sales_avg_kg'] > 0}
    if low_doh_items:
        print(f"  ⚠️ Phát hiện {len(low_doh_items)} mã cám có DOH < 3.0 ngày trong FFSTOCK:")
        for p, v in sorted(low_doh_items.items(), key=lambda x: x[1]['doh'])[:10]:
            print(f"    - {p:10s}: DOH = {v['doh']:.2f} ngày | Tồn = {v['stock_tons']:.1f}T | TB bán = {v['sales_avg_kg']:.0f} kg/ngày")
    else:
        print("  ✅ Không phát hiện mã cám nào có DOH < 3.0 ngày trong FFSTOCK")

    print(f"  ✅ Đọc chi tiết FFSTOCK: {len(result)} sản phẩm")
    return result


# ============================================================
# 5. LOAD TỒN BỒN
# ============================================================

def load_tonbon(file_path: str, day: int = None) -> dict:
    """
    Đọc Tồn Bồn (cám trong bồn) → dict {product → total_tons}

    File có 32 sheets (1 mỗi ngày). Mỗi sheet ghi nhận bồn nào chứa cám gì.
    Nếu day=None, tìm sheet số lớn nhất có dữ liệu.

    Cấu trúc: 2 nửa bảng:
      - Cột A(0)=bồn, B(1)=mã cám, C(2)=kg
      - Cột E(4)=bồn, F(5)=mã cám, G(6)=kg

    Args:
        file_path: Đường dẫn file tồn bồn
        day: Ngày cần đọc (1-31). None = ngày mới nhất.

    Returns:
        dict {'312001': 9.652, '320001': 16.404, ...}  (tấn, mã FORMULAR)
    """
    print(f"🏗️ Đọc TỒN BỒN: {os.path.basename(file_path)}")
    result = {}

    wb = _open_workbook(file_path)
    if wb is None:
        return result

    # Tìm sheet ngày phù hợp
    if day is not None:
        sheet_name = str(day)
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet ngày '{day}' không tồn tại")
            wb.close()
            return result
    else:
        # Tìm sheet số lớn nhất có dữ liệu
        numeric_sheets = []
        for name in wb.sheetnames:
            try:
                num = int(name)
                numeric_sheets.append(num)
            except ValueError:
                continue

        if not numeric_sheets:
            print("  ⚠️  Không tìm thấy sheet ngày nào")
            wb.close()
            return result

        # Duyệt từ lớn nhất để tìm sheet có dữ liệu
        sheet_name = None
        for num in sorted(numeric_sheets, reverse=True):
            ws_test = wb[str(num)]
            # Kiểm tra có dữ liệu thực tế không (check các dòng 10-45 có chứa mã cám và kg > 0)
            has_data = False
            for row in ws_test.iter_rows(min_row=10, max_row=45, max_col=8):
                if len(row) > 2:
                    p_left = _normalize_product_code(row[1].value)
                    kg_left = _safe_float(row[2].value)
                    if p_left and kg_left > 0:
                        has_data = True
                        break
                if len(row) > 6:
                    p_right = _normalize_product_code(row[5].value)
                    kg_right = _safe_float(row[6].value)
                    if p_right and kg_right > 0:
                        has_data = True
                        break
            if has_data:
                sheet_name = str(num)
                break

        if sheet_name is None:
            print("  ⚠️  Không tìm thấy sheet có dữ liệu")
            wb.close()
            return result

    ws = wb[sheet_name]
    print(f"  📋 Sheet ngày: {sheet_name}")

    # Đọc dữ liệu bồn - cấu trúc 2 nửa bảng
    for row in ws.iter_rows(min_row=8, max_row=45, max_col=8):
        # Nửa trái: B(1)=mã cám, C(2)=kg
        if len(row) > 2:
            product_left = _normalize_product_code(row[1].value)
            kg_left = _safe_float(row[2].value)
            if product_left and kg_left > 0:
                tons = kg_left / 1000.0
                result[product_left] = result.get(product_left, 0) + tons

        # Nửa phải: F(5)=mã cám, G(6)=kg
        if len(row) > 6:
            product_right = _normalize_product_code(row[5].value)
            kg_right = _safe_float(row[6].value)
            if product_right and kg_right > 0:
                tons = kg_right / 1000.0
                result[product_right] = result.get(product_right, 0) + tons

    wb.close()
    print(f"  ✅ Đọc TỒN BỒN ngày {sheet_name}: {len(result)} mã cám, "
          f"tổng {sum(result.values()):.1f} tấn")
    return result


def load_tonbon_detail(file_path: str, day: int = None) -> dict:
    """
    Đọc Tồn Bồn chi tiết theo từng bồn chứa -> dict {silo_number: {'product': product_code, 'tons': tons}}
    Chỉ lấy các bồn chứa bán thành phẩm cho Pellet (86-98) và bồn cám thành phẩm đặc biệt của PL6/PL7 (129, 130).
    """
    print(f"🏗️ Đọc TỒN BỒN CHI TIẾT: {os.path.basename(file_path)}")
    result = {}

    wb = _open_workbook(file_path)
    if wb is None:
        return result

    # Tìm sheet ngày phù hợp
    if day is not None:
        sheet_name = str(day)
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet ngày '{day}' không tồn tại")
            wb.close()
            return result
    else:
        # Tìm sheet số lớn nhất có dữ liệu
        numeric_sheets = []
        for name in wb.sheetnames:
            try:
                num = int(name)
                numeric_sheets.append(num)
            except ValueError:
                continue

        if not numeric_sheets:
            print("  ⚠️  Không tìm thấy sheet ngày nào")
            wb.close()
            return result

        # Duyệt từ lớn nhất để tìm sheet có dữ liệu
        sheet_name = None
        for num in sorted(numeric_sheets, reverse=True):
            ws_test = wb[str(num)]
            # Kiểm tra có dữ liệu thực tế không (check các dòng 10-45 có chứa mã cám và kg > 0)
            has_data = False
            for row in ws_test.iter_rows(min_row=10, max_row=45, max_col=8):
                if len(row) > 2:
                    p_left = _normalize_product_code(row[1].value)
                    kg_left = _safe_float(row[2].value)
                    if p_left and kg_left > 0:
                        has_data = True
                        break
                if len(row) > 6:
                    p_right = _normalize_product_code(row[5].value)
                    kg_right = _safe_float(row[6].value)
                    if p_right and kg_right > 0:
                        has_data = True
                        break
            if has_data:
                sheet_name = str(num)
                break

        if sheet_name is None:
            print("  ⚠️  Không tìm thấy sheet có dữ liệu")
            wb.close()
            return result

    ws = wb[sheet_name]
    print(f"  📋 Sheet ngày: {sheet_name}")

    # Đọc dữ liệu bồn - cấu trúc 2 nửa bảng
    # Dòng 8 đến 45: cột A đến H
    for row in ws.iter_rows(min_row=8, max_row=45, max_col=8):
        # Nửa trái: A(0)=bồn, B(1)=mã cám, C(2)=kg
        if len(row) > 2:
            silo_val = _safe_int(row[0].value)
            product_left = _normalize_product_code(row[1].value)
            kg_left = _safe_float(row[2].value)
            if silo_val and product_left and kg_left > 0:
                # Lọc các silo cần quan tâm
                if (86 <= silo_val <= 98) or silo_val in (129, 130):
                    tons = kg_left / 1000.0
                    result[silo_val] = {'product': product_left, 'tons': tons}

        # Nửa phải: E(4)=bồn, F(5)=mã cám, G(6)=kg
        if len(row) > 6:
            silo_val = _safe_int(row[4].value)
            product_right = _normalize_product_code(row[5].value)
            kg_right = _safe_float(row[6].value)
            if silo_val and product_right and kg_right > 0:
                if (86 <= silo_val <= 98) or silo_val in (129, 130):
                    tons = kg_right / 1000.0
                    result[silo_val] = {'product': product_right, 'tons': tons}

    wb.close()
    print(f"  ✅ Đọc TỒN BỒN CHI TIẾT ngày {sheet_name}: {len(result)} bồn được ghi nhận")
    return result


# ============================================================
# 6. LOAD EMPTY BAG

# ============================================================

def load_empty_bag(file_path: str) -> dict:
    """
    Đọc Empty Bag Report → dict {product → {brand → bags_available}}

    File có nhiều sheet theo thương hiệu: HIGRO, CP&STAR, NASA & NUVO, FARM.
    Mỗi sheet: Cột B=mã bao, Cột M(12)=TOTAL tồn cuối (bao).

    Args:
        file_path: Đường dẫn file Empty Bag

    Returns:
        dict {
            '552': {'HIGRO': 5000, 'CP': 200, 'STAR': 300},
            '552SF': {'FARM': 29313},
            ...
        }
    """
    print(f"🎒 Đọc EMPTY BAG: {os.path.basename(file_path)}")
    result = {}

    wb = _open_workbook(file_path)
    if wb is None:
        return result

    # Mapping sheet → brand
    brand_mapping = {
        'HIGRO': 'HIGRO',
        'CP': 'CP',
        'STAR': 'STAR',
        'NASA': 'NASA',
        'NUVO': 'NUVO',
        'FARM': 'FARM',
    }

    for sname in wb.sheetnames:
        sname_upper = sname.upper()

        # Xác định brand từ tên sheet
        current_brand = None
        for key, brand in brand_mapping.items():
            if key in sname_upper:
                current_brand = brand
                break

        if current_brand is None:
            continue

        # Xử lý sheet CP&STAR - chứa 2 brand
        # Sẽ phân biệt bằng vị trí dòng
        ws = wb[sname]

        active_brand = current_brand
        for row in ws.iter_rows(min_row=3, max_row=120, max_col=20):
            # Phát hiện header brand trong sheet (CP, STAR, NASA, NUVO...)
            cell_a = _safe_str(row[0].value) if row[0].value else ''
            cell_b = _safe_str(row[1].value) if len(row) > 1 else ''

            # Chuyển brand khi gặp header
            for key, brand in brand_mapping.items():
                if key in cell_a or key in cell_b:
                    if 'FEED' in cell_a or 'FEED' in cell_b or 'BRAND' in cell_a:
                        active_brand = brand
                        break

            # Cột B (index 1) = mã sản phẩm/mã SAP, Cột C (index 2) = tên bao bì chứa mã sản phẩm
            # Tìm mã sản phẩm (số + chữ, VD: 552, 550S, HT12S)
            cell_c = _safe_str(row[2].value) if len(row) > 2 else ''
            product_raw = cell_b
            if not product_raw:
                continue
                
            # Nếu cột B là mã SAP dài (VD: 1290D203010025), ta chuyển sang cột C để trích xuất mã sản phẩm
            if product_raw.startswith('1') and len(product_raw) > 8:
                product_raw = cell_c
                
            if not product_raw:
                continue

            # Bỏ qua header/tổng
            if any(kw in product_raw for kw in
                   {'STT', 'PRODUCT', 'FEED', 'TOTAL', 'THREAD', 'BALANCE',
                    'REMAIN', 'GRAND', 'SIZE', 'PACK', 'CODE'}):
                continue

            # Trích xuất mã sản phẩm ngắn gọn từ tên bao
            # VD: "BagO25- 552- 54x90-99-B-000-H" → "552"
            product = _normalize_product_code(product_raw)
            if 'BAG' in product_raw and '-' in product_raw:
                # Parse mã từ tên bao bì dạng "BagO25- 552- ..."
                parts = product_raw.split('-')
                for part in parts:
                    part = part.strip()
                    # Tìm phần chứa mã sản phẩm (3-8 ký tự, có số)
                    if (2 <= len(part) <= 10 and any(c.isdigit() for c in part)
                            and not part.startswith('BAG')):
                        product = _normalize_product_code(part)
                        break

            if not product or len(product) < 2:
                continue

            # Cột M (index 12) = TOTAL tồn cuối (bao)
            if len(row) > 12:
                bags = _safe_int(row[12].value)
            else:
                continue

            if bags <= 0:
                continue

            # Lưu kết quả
            if product not in result:
                result[product] = {}

            result[product][active_brand] = (
                result[product].get(active_brand, 0) + bags
            )

    wb.close()

    total_bags = sum(
        sum(brands.values()) for brands in result.values()
    )
    print(f"  ✅ Đọc EMPTY BAG: {len(result)} loại bao, "
          f"tổng {total_bags:,} bao")
    return result


# ============================================================
# 7. LOAD CÔNG SUẤT (Plan.xlsm)
# ============================================================

def load_congsuat(file_path: str) -> dict:
    """
    Đọc Công Suất từ Plan.xlsm → dict {product → ProductSpec}

    Tìm sheet chứa 'CONG' và 'SUAT' (hoặc gần đúng).
    Tìm ton_per_batch (giá trị giữa 5.0 và 10.0).
    Default ton_per_batch = 8.4.

    Args:
        file_path: Đường dẫn file Plan.xlsm

    Returns:
        dict {'552SF': ProductSpec(...), '550S': ProductSpec(...), ...}
    """
    print(f"⚙️ Đọc CÔNG SUẤT: {os.path.basename(file_path)}")
    result = {}

    wb = _open_workbook(file_path)
    if wb is None:
        return result

    # Tìm sheet chứa 'CONG' và 'SUAT' (hoặc 'CÔNG' và 'SUẤT')
    target_ws = None
    target_name = None
    for name in wb.sheetnames:
        name_upper = name.upper()
        # Tìm sheet có chứa cả hai từ (có dấu hoặc không)
        has_cong = 'CONG' in name_upper or 'CÔNG' in name_upper
        has_suat = 'SUAT' in name_upper or 'SUẤT' in name_upper or 'XUẤT' in name_upper
        if has_cong and has_suat:
            target_ws = wb[name]
            target_name = name
            break

    # Nếu không tìm thấy, thử tìm sheet có 'PLAN' hoặc 'CS'
    if target_ws is None:
        for name in wb.sheetnames:
            name_upper = name.upper()
            if 'PLAN' in name_upper or name_upper == 'CS':
                target_ws = wb[name]
                target_name = name
                break

    if target_ws is None:
        print("  ⚠️  Không tìm thấy sheet CÔNG SUẤT")
        wb.close()
        return result

    print(f"  📋 Sheet: {target_name}")

    for row in target_ws.iter_rows(min_row=2, max_row=400, max_col=10):
        val_str = _safe_str(row[0].value)  # Cột A = mã sản phẩm
        if not val_str:
            continue

        if any(kw in val_str for kw in {'TOTAL', 'STT', 'TÊN', 'MÃ',
                                         'SẢN PHẨM', 'PRODUCT'}):
            continue

        product = _normalize_product_code(row[0].value)

        # Tìm ton_per_batch trong các cột sau
        ton_per_batch = 8.4  # Mặc định
        for col_idx in range(1, min(len(row), 10)):
            val = _safe_float(row[col_idx].value)
            if 5.0 <= val <= 10.0:
                ton_per_batch = val
                break

        spec = ProductSpec(
            product_code=product,
            ton_per_batch=ton_per_batch,
        )
        result[product] = spec

    wb.close()
    print(f"  ✅ Đọc CÔNG SUẤT: {len(result)} sản phẩm")
    return result


# ============================================================
# 8. LOAD FEEDCODE
# ============================================================

def load_feedcode(file_path: str) -> dict:
    """
    Đọc FEEDCODE → dict {product → {line_cv, line_pk}}

    Từ file KHSX, sheet FEEDCODE.
    Cột A = mã sản phẩm, Cột B = LINE cám viên, Cột C = LINE đóng bao.

    Args:
        file_path: Đường dẫn file KHSX

    Returns:
        dict {
            '552SF': {'line_cv': '3', 'line_pk': '4'},
            '550S': {'line_cv': '1', 'line_pk': '5'},
            ...
        }
    """
    print(f"🔧 Đọc FEEDCODE: {os.path.basename(file_path)}")
    result = {}

    wb = _open_workbook(file_path)
    if wb is None:
        return result

    # Tìm sheet FEEDCODE
    ws = None
    for name in wb.sheetnames:
        if 'FEEDCODE' in name.upper() or 'FEED CODE' in name.upper():
            ws = wb[name]
            break

    if ws is None:
        print("  ⚠️  Không tìm thấy sheet FEEDCODE")
        wb.close()
        return result

    for row in ws.iter_rows(min_row=2, max_row=500, max_col=5):
        if len(row) < 4:
            continue
        val_str = _safe_str(row[1].value)  # B = FEED NAME (Tên cám)
        if not val_str:
            continue

        if any(kw in val_str for kw in {'TÊN', 'MÃ', 'STT', 'FEED', 'PRODUCT', 'CODE', 'CÁM'}):
            continue

        product = _normalize_product_code(row[1].value)

        line_cv = _safe_str(row[2].value) if len(row) > 2 else ''  # C = MÁY ÉP VIÊN
        line_pk = _safe_str(row[3].value) if len(row) > 3 else ''  # D = MÁY ĐÓNG BAO

        result[product] = {
            'line_cv': line_cv,
            'line_pk': line_pk,
        }

    wb.close()
    print(f"  ✅ Đọc FEEDCODE: {len(result)} mã sản phẩm")
    return result


# ============================================================
# 9. LOAD KHÁNG SINH
# ============================================================

def load_khangsinh(file_path: str) -> dict:
    """
    Đọc mã Kháng Sinh → dict {product → ks_code}

    Từ file KHSX, sheet chứa 'KHANG' hoặc 'KS'.
    Cột A = mã sản phẩm, Cột B = mã kháng sinh (KS/HC).

    Args:
        file_path: Đường dẫn file KHSX

    Returns:
        dict {'552SF': 'KS01', '550S': '', ...}
    """
    print(f"💊 Đọc KHÁNG SINH: {os.path.basename(file_path)}")
    result = {}

    wb = _open_workbook(file_path)
    if wb is None:
        return result

    # Tìm sheet kháng sinh
    ws = None
    for name in wb.sheetnames:
        name_upper = name.upper()
        if ('KHANG' in name_upper or 'KHÁNG' in name_upper
                or name_upper == 'KS' or 'KS' in name_upper.split()):
            ws = wb[name]
            break

    if ws is None:
        print("  ⚠️  Không tìm thấy sheet KHÁNG SINH")
        wb.close()
        return result

    for row in ws.iter_rows(min_row=2, max_row=500, max_col=5):
        if len(row) < 3:
            continue
        val_str = _safe_str(row[1].value)  # B = CÁM (Mã sản phẩm)
        if not val_str:
            continue

        if any(kw in val_str for kw in {'TÊN', 'MÃ', 'STT', 'PRODUCT', 'KHÁNG', 'KHANG', 'CODE', 'CÁM'}):
            continue

        product = _normalize_product_code(row[1].value)

        ks_code = normalize_ks_code(row[2].value) if len(row) > 2 else ''  # C = KHÁNG SINH
        ks_level = _parse_antibiotic_level(ks_code)
        result[product] = {
            'ks_code': ks_code,
            'ks_level': ks_level,
        }

    wb.close()
    print(f"  ✅ Đọc KHÁNG SINH: {len(result)} mã sản phẩm")
    return result


def _parse_antibiotic_level(ks_code: str) -> int:
    """
    Bóc tách mức độ kháng sinh lớn nhất từ chuỗi 'KS/ HC (x)/(y)'.
    Ví dụ: 'KS/ HC (1)/(15)' -> 15.
    Mặc định trả về 1 (cám sạch) nếu không chứa kháng sinh.
    """
    if not ks_code:
        return 1
        
    # Tìm tất cả các số trong ngoặc đơn dạng (x)
    matches = re.findall(r'\((\d+)\)', ks_code)
    if not matches:
        return 1
        
    levels = []
    for m in matches:
        try:
            levels.append(int(m))
        except ValueError:
            continue
            
    if levels:
        return max(levels)
    return 1


def resolve_antibiotic_for_product(product_code: str, base_khangsinh_dict: dict) -> tuple[str, int]:
    """
    Giải quyết kháng sinh động cho mọi mã cám:
    - Sửa các lỗi typo từ forecast (SX -> XS).
    - Ánh xạ mã Silo 'HT' về mã nền trại tương ứng (HT11 -> 551, HT12 -> 552).
    - Ánh xạ mã đại lý 'HG' về mã đại lý gốc (HG16 -> 566, HG17 -> 567, HG16XS34 -> 566XS34).
    - Trả về tuple (ks_code, ks_level).
    """
    p_code = str(product_code).strip().upper().replace(' ', '')
    if not p_code:
        return 'KS/ HC (1)/(2)', 2

    # 1. Kiểm tra trực tiếp trong dict kháng sinh gốc
    if p_code in base_khangsinh_dict:
        data = base_khangsinh_dict[p_code]
        return data.get('ks_code', ''), data.get('ks_level', 1)

    # 2. Xử lý lỗi chính tả thường gặp (ví dụ: SX54 -> XS54)
    p_fixed = p_code.replace('SX54', 'XS54').replace('SX34', 'XS34')
    if p_fixed in base_khangsinh_dict:
        data = base_khangsinh_dict[p_fixed]
        return data.get('ks_code', ''), data.get('ks_level', 1)

    # 3. Xử lý mã Silo (Bắt đầu bằng HT)
    if p_code.startswith('HT'):
        # Thử thay thế HT1 bằng 55
        mapped_code = p_code.replace('HT1', '55', 1)
        if mapped_code in base_khangsinh_dict:
            data = base_khangsinh_dict[mapped_code]
            return data.get('ks_code', ''), data.get('ks_level', 1)
        
        # Thử regex HT1([1-9]) -> 55\1
        m = re.match(r'^HT1([1-9])(.*)', p_code)
        if m:
            mapped_code_3 = f"55{m.group(1)}{m.group(2)}"
            if mapped_code_3 in base_khangsinh_dict:
                data = base_khangsinh_dict[mapped_code_3]
                return data.get('ks_code', ''), data.get('ks_level', 1)

        # Fallback cho cám Silo: Thường là cám sạch nền, cấp độ 2
        return 'KS/ HC (1)/(2)', 2

    # 4. Xử lý mã đại lý (Bắt đầu bằng HG)
    if p_code.startswith('HG'):
        # Thay thế HG1 bằng 56
        mapped_code = p_code.replace('HG1', '56', 1)
        if mapped_code in base_khangsinh_dict:
            data = base_khangsinh_dict[mapped_code]
            return data.get('ks_code', ''), data.get('ks_level', 1)
        
        # Thử regex HG1([6-9]) -> 56\1
        m = re.match(r'^HG1([6-9])(.*)', p_code)
        if m:
            mapped_code_3 = f"56{m.group(1)}{m.group(2)}"
            if mapped_code_3 in base_khangsinh_dict:
                data = base_khangsinh_dict[mapped_code_3]
                return data.get('ks_code', ''), data.get('ks_level', 1)
            
            # Thử sửa cả lỗi SX/XS trong HG nữa (ví dụ: HG16SX34 -> HG16XS34 -> 566XS34)
            mapped_code_4 = mapped_code_3.replace('SX34', 'XS34').replace('SX54', 'XS54')
            if mapped_code_4 in base_khangsinh_dict:
                data = base_khangsinh_dict[mapped_code_4]
                return data.get('ks_code', ''), data.get('ks_level', 1)

        # Fallback cho cám đại lý: Thường là cám sạch nền, cấp độ 2
        return 'KS/ HC (1)/(2)', 2

    # 5. Các dòng đặc biệt khác trong forecast (Ví dụ: 550S(5*5), 551(5*5), 552WDF...)
    base_match = re.match(r'^(\d{3}[A-Z]*).*', p_code)
    if base_match:
        base_code = base_match.group(1)
        if base_code in base_khangsinh_dict:
            data = base_khangsinh_dict[base_code]
            return data.get('ks_code', ''), data.get('ks_level', 1)

    # Fallback cuối cùng: Cám sạch nền tiêu chuẩn, cấp độ 2
    return 'KS/ HC (1)/(2)', 2


def load_stt_khangsinh(file_path: str) -> dict:
    """
    Đọc danh sách 26 cấp độ kháng sinh từ sheet 'STT Khang sinh' của Plan.xlsm.
    Trả về dict {stt: medicine_name}
    """
    print(f"💊 Đọc danh mục cấp độ KHÁNG SINH: {os.path.basename(file_path)}")
    result = {}
    
    wb = _open_workbook(file_path)
    if wb is None:
        return result
        
    # Tìm sheet STT Khang sinh
    ws = None
    for name in wb.sheetnames:
        name_upper = name.upper()
        if 'STT KHANG' in name_upper or 'STT KHÁNG' in name_upper or name_upper == 'STT KHANG SINH':
            ws = wb[name]
            break
            
    if ws is None:
        print("  ⚠️ Không tìm thấy sheet 'STT Khang sinh'")
        wb.close()
        return result
        
    for row in ws.iter_rows(min_row=3, max_row=50, max_col=3):
        stt_val = row[0].value
        med_val = row[1].value
        if stt_val is not None:
            try:
                stt = int(stt_val)
                med_name = str(med_val).strip() if med_val else ''
                result[stt] = med_name
            except (ValueError, TypeError):
                continue
                
    wb.close()
    print(f"  ✅ Đọc được {len(result)} cấp độ kháng sinh từ danh mục")
    return result


def load_fix_code_pellet(file_path: str) -> dict:
    """
    Đọc ma trận ưu tiên gán máy Pellet từ sheet 'Fix code pellet' của Plan.xlsm.
    Trả về dict {product_full_code: {'default_line': str, 'priorities': list[str], 'note': str}}
    """
    print(f"⚙️ Đọc MA TRẬN Pellet: {os.path.basename(file_path)}")
    result = {}
    
    wb = _open_workbook(file_path)
    if wb is None:
        return result
        
    # Tìm sheet Fix code pellet
    ws = None
    for name in wb.sheetnames:
        name_upper = name.upper()
        if 'FIX CODE' in name_upper or 'PELLET' in name_upper:
            ws = wb[name]
            break
            
    if ws is None:
        print("  ⚠️ Không tìm thấy sheet 'Fix code pellet'")
        wb.close()
        return result
        
    for row in ws.iter_rows(min_row=3, max_row=300, max_col=32):
        if len(row) < 31:
            continue
            
        code_val = row[1].value # Cột B = Mã đầy đủ
        if code_val is None:
            continue
            
        code = _normalize_product_code(code_val)
        if not code:
            continue
            
        default_line = _safe_str(row[3].value) # Cột D = Line_CV mặc định
        
        # Đọc 5 mức ưu tiên từ cột Z đến AD (index 25 đến 29)
        priorities = []
        for col_idx in range(25, 30):
            val = row[col_idx].value
            if val is not None:
                p_line = str(val).strip().upper()
                if p_line and p_line != 'NONE' and p_line != 'NULL':
                    # Chuẩn hóa tên máy, ví dụ '4' -> 'PL4', 'PL4' -> 'PL4'
                    if p_line.isdigit():
                        p_line = f"PL{p_line}"
                    priorities.append(p_line)
                    
        note = _safe_str(row[30].value) # Cột AE = Ghi chú
        
        result[code] = {
            'default_line': default_line,
            'priorities': priorities,
            'note': note
        }
        
    wb.close()
    print(f"  ✅ Đọc được {len(result)} cấu hình gán máy Pellet từ ma trận")
    return result


# ============================================================
# 10. LOAD KHSX YESTERDAY
# ============================================================

def load_khsx_yesterday(file_path: str, day: int) -> dict:
    """
    Đọc KHSX ngày hôm qua → dict {product → {planned, actual, pct, shortfall}}

    Từ file KHSX, sheet = str(day) hoặc sheet ngày DD-MM-YYYY.
    Cột B(1) = tên cám, Cột C(2) = số mẻ kế hoạch.
    Cột Z-AC (25-28) = số mẻ thực hiện.

    Args:
        file_path: Đường dẫn file KHSX
        day: Ngày cần đọc (1-31)

    Returns:
        dict {
            '552SF': {'planned': 40, 'actual': 38, 'pct': 95.0, 'shortfall': 2},
            ...
        }
    """
    print(f"📋 Đọc KHSX ngày {day}: {os.path.basename(file_path)}")
    result = {}

    wb = _open_workbook(file_path)
    if wb is None:
        return result

    # Tìm sheet phù hợp: thử theo thứ tự sheet_name = str(day), hoặc sheet chứa '-day-', hoặc sheet đầu tiên (active)
    sheet_name = str(day)
    ws = None
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Thử tìm theo định dạng DD-MM-YYYY
        for name in wb.sheetnames:
            if f"{day:02d}-" in name or f"-{day:02d}-" in name:
                ws = wb[name]
                sheet_name = name
                break
        if ws is None:
            # Lấy sheet đầu tiên (active)
            ws = wb.active
            sheet_name = ws.title

    print(f"  📋 Sử dụng sheet: '{sheet_name}'")

    for row in ws.iter_rows(min_row=2, max_row=80, max_col=32):
        # Cột B (index 1) = TÊN CÁM
        if len(row) < 3:
            continue

        val_str = _safe_str(row[1].value)
        if not val_str:
            continue

        if any(kw in val_str for kw in {'TỔNG', 'TOTAL', 'STT',
                                         'TÊN', 'CÁM', 'KHSX'}):
            continue

        product = _normalize_product_code(row[1].value)

        # Cột C (index 2) = KẾ HOẠCH (MẺ)
        planned = _safe_int(row[2].value)
        if planned <= 0:
            continue

        # Tìm actual trong cột Z-AC (index 25-28)
        actual = 0
        for col_idx in range(25, min(29, len(row))):
            val = _safe_int(row[col_idx].value)
            if val > 0:
                actual = val
                break

        # Tính phần trăm hoàn thành
        pct = (actual / planned * 100.0) if planned > 0 else 0.0
        shortfall = max(0, planned - actual)

        result[product] = {
            'planned': planned,
            'actual': actual,
            'pct': round(pct, 1),
            'shortfall': shortfall,
        }

    wb.close()
    print(f"  ✅ Đọc KHSX ngày {day}: {len(result)} sản phẩm, "
          f"thiếu {sum(v['shortfall'] for v in result.values())} mẻ")
    return result


def create_quick_adjustments_template(file_path: str):
    """
    Tự động khởi tạo file Excel DIEU_CHINH_NHANH.xlsx với 4 sheets mẫu
    nếu chưa tồn tại, kèm theo header và dòng hướng dẫn mẫu.
    """
    if os.path.exists(file_path):
        return

    print(f"✨ Khởi tạo file mẫu Điều chỉnh nhanh: {os.path.basename(file_path)}")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    
    # Định dạng style
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    border_side = Side(border_style="thin", color="D9D9D9")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # ─── Sheet 1: THEM_MOI_HOAC_SUA ───
    ws1 = wb.active
    ws1.title = "THEM_MOI_HOAC_SUA"
    headers1 = ["MÃ CÁM", "TẤN", "QUY CÁCH", "LOẠI ƯU TIÊN", "ÉP SỐ MẺ", "ÉP TẤN/MẺ", "GHI CHÚ (Tùy chọn)"]
    ws1.append(headers1)
    
    # Thêm dòng mẫu
    sample_rows1 = [
        ["550S", 10.0, "SILO", "SILO", 1, 10.0, "Bổ sung xe silo ngày mai"],
        ["551", 9.0, "SILO", "SILO", 1, 9.0, "Bổ sung xe silo ngày mai"],
        ["567SXS34", 6.0, "25", "WALKIN", 1, 6.0, "Khách đặt ép đúng 1 mẻ 6 tấn"],
        ["553S", 12.0, "25", "FORECAST", "", "", "Ép số lượng sản xuất"],
    ]
    for row in sample_rows1:
        ws1.append(row)
        
    # Hướng dẫn bên cạnh
    ws1["I2"] = "HƯỚNG DẪN ĐIỀN SHEET THEM_MOI_HOAC_SUA:"
    ws1["I2"].font = Font(bold=True, size=11, color="1F4E78")
    ws1["I3"] = "- MÃ CÁM: Viết hoa, liền nhau (VD: 550S, 567SXS34)"
    ws1["I4"] = "- TẤN: Số lượng sản xuất bổ sung hoặc muốn đổi"
    ws1["I5"] = "- QUY CÁCH: Điền 'SILO' hoặc size bao ('25', '40', '50')"
    ws1["I6"] = "- LOẠI ƯU TIÊN: SILO, BACANG, WALKIN, hoặc FORECAST"
    ws1["I7"] = "- ÉP SỐ MẺ & ÉP TẤN/MẺ: Để trống nếu muốn hệ thống tự tính. Điền nếu muốn ép máy trộn (VD: 1 mẻ 6.0T)"
    
    # ─── Sheet 2: HUY_KHSX ───
    ws2 = wb.create_sheet("HUY_KHSX")
    headers2 = ["MÃ CÁM", "LOẠI HỦY", "GHI CHÚ"]
    ws2.append(headers2)
    sample_rows2 = [
        ["562", "BAO", "Hủy cám bao Bá Cang ngày thứ 4"],
        ["322", "ALL", "Ngưng sản xuất hoàn toàn mã này"],
        ["550S", "SILO", "Hủy phần xe bồn silo"],
    ]
    for row in sample_rows2:
        ws2.append(row)
        
    ws2["E2"] = "HƯỚNG DẪN ĐIỀN SHEET HUY_KHSX:"
    ws2["E2"].font = Font(bold=True, size=11, color="1F4E78")
    ws2["E3"] = "- LOẠI HỦY: 'SILO' (chỉ hủy xe bồn), 'BAO' (chỉ hủy đóng bao) hoặc 'ALL' (hủy toàn bộ)"
    
    # ─── Sheet 3: THAY_THE_MA_CAM ───
    ws3 = wb.create_sheet("THAY_THE_MA_CAM")
    headers3 = ["MÃ CŨ", "MÃ MỚI", "GHI CHÚ"]
    ws3.append(headers3)
    sample_rows3 = [
        ["553S", "552", "Thay thế mã 553S thành 552 theo yêu cầu phòng bán"],
        ["550XS54PRO", "550XS54", "Sửa lỗi nhập nhầm mã từ đại lý"],
    ]
    for row in sample_rows3:
        ws3.append(row)
        
    ws3["E2"] = "HƯỚNG DẪN ĐIỀN SHEET THAY_THE_MA_CAM:"
    ws3["E2"].font = Font(bold=True, size=11, color="1F4E78")
    ws3["E3"] = "- Thay thế tự động nhu cầu của MÃ CŨ sang MÃ MỚI trước khi tổng hợp kế hoạch."
    
    # ─── Sheet 4: THAY_THE_BAO_BI ───
    ws4 = wb.create_sheet("THAY_THE_BAO_BI")
    headers4 = ["MÃ CÁM", "BAO GỐC", "BAO THAY THẾ", "GHI CHÚ"]
    ws4.append(headers4)
    sample_rows4 = [
        ["521SNPRO", "521SNPRO", "521SPRO", "Sử dụng bao 521SPRO thay cho 521SNPRO (hết bao bì gốc)"],
        ["522S", "522S", "WHITE BAG", "Đóng bao trắng tạm thời do hết vỏ bao"],
    ]
    for row in sample_rows4:
        ws4.append(row)
        
    ws4["F2"] = "HƯỚNG DẪN ĐIỀN SHEET THAY_THE_BAO_BI:"
    ws4["F2"].font = Font(bold=True, size=11, color="1F4E78")
    ws4["F3"] = "- MÃ CÁM: Mã cám cần đổi vỏ bao (VD: 521SNPRO)"
    ws4["F4"] = "- BAO GỐC: Tên vỏ bao nguyên bản (thương hiệu gốc)"
    ws4["F5"] = "- BAO THAY THẾ: HIGRO, CP, STAR, NUVO, NASA, BELL hoặc WHITE BAG (bao trắng)"

    # Style đẹp cho tất cả các sheets
    for ws in [ws1, ws2, ws3, ws4]:
        # Header styles
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
        
        # Border & Align for data rows
        for row in range(2, 7):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and col < 6:
                    cell.alignment = left_align
                    cell.border = border
                    
        # Tự động dãn cột
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = cell.value
                if val:
                    # Bỏ qua các chuỗi hướng dẫn quá dài khi tính toán độ rộng cột
                    if isinstance(val, str) and (len(val) > 40 or val.startswith("-") or val.startswith("HƯỚNG DẪN")):
                        continue
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 16
        if ws == ws1:
            ws.column_dimensions['G'].width = 30
            ws.column_dimensions['I'].width = 40
        elif ws == ws4:
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['F'].width = 45
        else:
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['E'].width = 45
            
    try:
        wb.save(file_path)
        print(f"  ✅ Khởi tạo thành công file: {file_path}")
    except Exception as e:
        print(f"  ❌ Không thể tạo file mẫu Điều chỉnh nhanh: {e}")


def load_quick_adjustments(file_path: str) -> dict:
    """
    Đọc dữ liệu điều chỉnh từ file DIEU_CHINH_NHANH.xlsx
    Trả về dict chứa 4 phần điều chỉnh.
    """
    adjustments = {
        'additions': [],        # list[dict]
        'cancellations': {},    # dict {product_code → 'SILO' | 'BAO' | 'ALL'}
        'substitutions': {},    # dict {old_code → new_code}
        'bag_substitutions': {},# dict {product_code → {original_brand → replacement_brand}}
    }
    
    if not os.path.exists(file_path):
        print(f"  ⚠️  File Điều chỉnh nhanh không tồn tại: {file_path}. Sử dụng cấu hình trống.")
        return adjustments
        
    print(f"🔄 Đọc cấu hình ĐIỀU CHỈNH NHANH: {os.path.basename(file_path)}")
    wb = _open_workbook(file_path, data_only=True, read_only=True)
    if wb is None:
        return adjustments
        
    # 1. THEM_MOI_HOAC_SUA
    if "THEM_MOI_HOAC_SUA" in wb.sheetnames:
        ws = wb["THEM_MOI_HOAC_SUA"]
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if not row or len(row) < 4:
                continue
            product = _normalize_product_code(row[0])
            if not product:
                continue
            
            tons = _safe_float(row[1])
            packing = _safe_str(row[2], '25')
            priority_str = _safe_str(row[3], 'FORECAST')
            
            # Ép mẻ & Ép tấn/mẻ
            force_batches = None
            if len(row) > 4 and row[4] is not None:
                force_batches = _safe_int(row[4])
                if force_batches <= 0:
                    force_batches = None
                    
            force_tpb = None
            if len(row) > 5 and row[5] is not None:
                force_tpb = _safe_float(row[5])
                if force_tpb <= 0:
                    force_tpb = None
            
            adjustments['additions'].append({
                'product_code': product,
                'tons': tons,
                'packing_size': packing,
                'priority': priority_str,
                'force_batches': force_batches,
                'force_tpb': force_tpb
            })
            
    # 2. HUY_KHSX
    if "HUY_KHSX" in wb.sheetnames:
        ws = wb["HUY_KHSX"]
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if not row or len(row) < 2:
                continue
            product = _normalize_product_code(row[0])
            if not product:
                continue
            type_cancel = _safe_str(row[1], 'ALL')
            adjustments['cancellations'][product] = type_cancel
            
    # 3. THAY_THE_MA_CAM
    if "THAY_THE_MA_CAM" in wb.sheetnames:
        ws = wb["THAY_THE_MA_CAM"]
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if not row or len(row) < 2:
                continue
            old_code = _normalize_product_code(row[0])
            new_code = _normalize_product_code(row[1])
            if old_code and new_code:
                adjustments['substitutions'][old_code] = new_code
                
    # 4. THAY_THE_BAO_BI
    if "THAY_THE_BAO_BI" in wb.sheetnames:
        ws = wb["THAY_THE_BAO_BI"]
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if not row or len(row) < 3:
                continue
            product = _normalize_product_code(row[0])
            orig_brand = _safe_str(row[1])
            repl_brand = _safe_str(row[2])
            if product and orig_brand and repl_brand:
                if product not in adjustments['bag_substitutions']:
                    adjustments['bag_substitutions'][product] = {}
                adjustments['bag_substitutions'][product][orig_brand] = repl_brand
                
    wb.close()
    
    # In tóm tắt điều chỉnh nhanh đọc được
    print(f"  ✅ Đã tải cấu hình điều chỉnh nhanh:")
    print(f"     ➕ Bổ sung/Sửa đổi:  {len(adjustments['additions'])} dòng")
    print(f"     ❌ Hủy kế hoạch:    {len(adjustments['cancellations'])} mã cám")
    print(f"     🔄 Thay thế mã:      {len(adjustments['substitutions'])} quy tắc")
    print(f"     🎒 Đổi vỏ bao:       {sum(len(v) for v in adjustments['bag_substitutions'].values())} quy tắc")
    
    return adjustments


# ============================================================
# 11. HELPER DOWNLOAD SHAREPOINT
# ============================================================

def download_sharepoint_file(url, local_path) -> bool:
    """Tải file từ link SharePoint/OneDrive trực tiếp với cơ chế giả lập trình duyệt và giải nén thư mục tự động nếu là ZIP"""
    if not url:
        return False
    url = url.strip()
    if not url.startswith("http"):
        return False
        
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    
    # Thiết lập link tải trực tiếp (?download=1)
    direct_url = url
    if "sharepoint.com" in url and "download=1" not in url:
        if "?" in url:
            direct_url = url + "&download=1"
        else:
            direct_url = url + "?download=1"
            
    print(f"📥 Đang tải từ SharePoint: {url[:60]}... -> {local_path}")
    try:
        import urllib.request
        import zipfile
        import io
        import re
        import datetime
        
        # User-Agent để vượt qua các lớp chặn từ hệ thống bảo mật của SharePoint
        req = urllib.request.Request(
            direct_url,
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'}
        )
        with urllib.request.urlopen(req, timeout=60) as response:
            response_data = response.read()
            
        # Kiểm tra tính hợp lệ của file Excel/ZIP (nếu là HTML redirect thì bỏ qua và báo lỗi)
        is_zip = response_data.startswith(b'PK\x03\x04')
        is_xls = response_data.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1')
        
        if not (is_zip or is_xls):
            # Nếu chứa HTML (ví dụ trang đăng nhập Microsoft)
            sample = response_data[:1000].decode('utf-8', errors='ignore')
            if '<html' in sample.lower() or '<!doctype' in sample.lower() or 'microsoft' in sample.lower() or 'redirect' in sample.lower():
                raise Exception("SharePoint yêu cầu xác thực đăng nhập doanh nghiệp (MFA/ADFS). Vui lòng sử dụng nút 'Mở thư mục SharePoint' tải về máy rồi Kéo thả để cập nhật nhanh.")
            raise Exception("Dữ liệu nhận về không phải định dạng Excel (.xlsx/.xlsm) hoặc ZIP hợp lệ.")
            
        if is_zip:
            # Có thể là file Excel (.xlsx) trực tiếp hoặc file ZIP của cả thư mục
            with zipfile.ZipFile(io.BytesIO(response_data)) as z:
                names = z.namelist()
                # Nếu là file Excel trực tiếp, nó sẽ chứa thư mục xl/
                is_direct_excel = 'xl/workbook.xml' in names or '[Content_Types].xml' in names
                
                if is_direct_excel:
                    print("📄 Nhận diện dữ liệu tải về là tệp Excel trực tiếp (.xlsx/.xlsm).")
                    with open(local_path, 'wb') as f:
                        f.write(response_data)
                    print(f"✅ Tải thành công file SharePoint trực tiếp: {os.path.basename(local_path)}")
                    return True
                else:
                    print("📦 Nhận diện dữ liệu tải về là dạng nén ZIP (SharePoint Shared Folder).")
                    excel_files = []
                    for name in names:
                        if name.endswith('/') or os.path.basename(name).startswith('~$'):
                            continue
                        ext = os.path.splitext(name)[1].lower()
                        if ext in ['.xlsx', '.xls', '.xlsm']:
                            excel_files.append(name)
                    
                    if excel_files:
                        date_pattern = re.compile(r'(\d{1,2})[-_/\s](\d{1,2})[-_/\s](\d{4})')
                        
                        def get_file_sort_key(filename):
                            match = date_pattern.search(os.path.basename(filename))
                            if match:
                                try:
                                    d, m, y = int(match.group(1)), int(match.group(2)), int(match.group(3))
                                    return (2, datetime.date(y, m, d), filename)
                                except Exception:
                                    pass
                            try:
                                zinfo = z.getinfo(filename)
                                dt = datetime.datetime(*zinfo.date_time)
                                return (1, dt, filename)
                            except Exception:
                                pass
                            return (0, filename, filename)
                        
                        excel_files.sort(key=get_file_sort_key, reverse=True)
                        target_zip_file = excel_files[0]
                        print(f"📂 Đã tìm thấy {len(excel_files)} file Excel trong thư mục. File mới nhất: {target_zip_file}")
                        
                        with open(local_path, 'wb') as f:
                            f.write(z.read(target_zip_file))
                        print(f"✅ Tải và giải nén thành công file SharePoint: {os.path.basename(local_path)}")
                        return True
                    else:
                        raise Exception("Không tìm thấy file Excel hợp lệ nào bên trong tệp ZIP tải về.")
        else:
            # File .xls cổ điển trực tiếp
            with open(local_path, 'wb') as f:
                f.write(response_data)
            print(f"✅ Tải thành công file SharePoint trực tiếp (.xls): {os.path.basename(local_path)}")
            return True
            
    except Exception as e:
        print(f"❌ Lỗi tải file SharePoint về {local_path}: {e}")
        return False


def download_sharepoint_files_if_configured(config):
    """Đồng bộ tải toàn bộ file từ SharePoint về local nếu có thiết lập"""
    print("\n🌐 BẮT ĐẦU ĐỒNG BỘ TẢI FILE TỪ CLOUD SHAREPOINT...")
    
    # 1. Forecast
    forecast_url = getattr(config, 'SHAREPOINT_FORECAST_URL', '')
    if forecast_url:
        local_path = os.path.join(config.FORECAST_DIR, 'SharePoint_Forecast.xlsx')
        download_sharepoint_file(forecast_url, local_path)
        
    # 2. Silo
    silo_url = getattr(config, 'SHAREPOINT_SILO_URL', '')
    if silo_url:
        local_path = os.path.join(config.SILO_DIR, 'SharePoint_Silo.xlsx')
        download_sharepoint_file(silo_url, local_path)
        
    # 3. Ba Cang
    bacang_url = getattr(config, 'SHAREPOINT_BACANG_URL', '')
    if bacang_url:
        local_path = os.path.join(config.BACANG_DIR, 'SharePoint_BaCang.xlsx')
        download_sharepoint_file(bacang_url, local_path)
        
    # 4. FFStock
    ffstock_url = getattr(config, 'SHAREPOINT_FFSTOCK_URL', '')
    if ffstock_url:
        local_path = os.path.join(config.FSTOCK_DIR, 'SharePoint_FFStock.xlsx')
        download_sharepoint_file(ffstock_url, local_path)
        
    # 5. Empty Bag
    empty_bag_url = getattr(config, 'SHAREPOINT_EMPTY_BAG_URL', '')
    if empty_bag_url:
        local_path = os.path.join(config.FSTOCK_DIR, 'SharePoint_EmptyBag.xlsx')
        download_sharepoint_file(empty_bag_url, local_path)
        
    # 6. Ton Bon
    tonbon_url = getattr(config, 'SHAREPOINT_TONBON_URL', '')
    if tonbon_url:
        local_path = os.path.join(config.TONBON_DIR, 'SharePoint_TonBon.xlsx')
        download_sharepoint_file(tonbon_url, local_path)
        
    # 7. Plan File (Plan.xlsm)
    plan_url = getattr(config, 'SHAREPOINT_PLAN_URL', '')
    if plan_url:
        download_sharepoint_file(plan_url, config.PLAN_FILE)


# ============================================================
# 12. LOAD ALL DATA
# ============================================================

def load_all_data(config, target_date=None) -> dict:
    if getattr(config, 'USE_POSTGRESQL', False):
        import db_manager
        db_uri = getattr(config, 'DB_URI', db_manager.DB_URI)
        return db_manager.load_all_data_from_db(db_uri, target_date)

    # Tải các tệp SharePoint về thư mục local tương ứng nếu được cấu hình
    download_sharepoint_files_if_configured(config)

    print("=" * 60)
    print("🏭 BẮT ĐẦU ĐỌC DỮ LIỆU KHSX")
    print("=" * 60)

    data = {}

    # ─── 0. ĐIỀU CHỈNH NHANH (QUICK ADJUSTMENTS) ───────────────
    print("\n" + "─" * 40)
    quick_adjust_file = getattr(config, 'QUICK_ADJUST_FILE', None)
    if quick_adjust_file:
        create_quick_adjustments_template(quick_adjust_file)
        data['adjustments'] = load_quick_adjustments(quick_adjust_file)
    else:
        data['adjustments'] = {
            'additions': [],
            'cancellations': {},
            'substitutions': {},
            'bag_substitutions': {},
        }
        print("  ⚠️  Bỏ qua ĐIỀU CHỈNH NHANH")

    # ─── 1. FORECAST ─────────────────────────────────────────
    print("\n" + "─" * 40)
    forecast_file = _find_latest_file(
        config.FORECAST_DIR, '*FORECAST*.xlsx'
    )
    if not forecast_file:
        fallback_dir = getattr(config, 'FORECAST_DIR_FALLBACK', None)
        if fallback_dir and fallback_dir != config.FORECAST_DIR:
            print(f"  ℹ️  Không tìm thấy file Forecast trong OneDrive, thử tìm ở thư mục dự phòng: {fallback_dir}")
            forecast_file = _find_latest_file(fallback_dir, '*FORECAST*.xlsx')

    if forecast_file:
        data['forecast'] = load_forecast(forecast_file)
    else:
        data['forecast'] = []
        print("  ⚠️  Bỏ qua FORECAST")

    # ─── 2. SILO PLAN ────────────────────────────────────────
    print("\n" + "─" * 40)
    silo_file = _find_latest_file(
        config.SILO_DIR, '*SILO*.xlsx'
    )
    if silo_file:
        data['silo_plan'] = load_silo_plan(silo_file)
    else:
        data['silo_plan'] = {d: {} for d in range(1, 7)}
        print("  ⚠️  Bỏ qua SILO PLAN")

    # ─── 3. BA CANG ──────────────────────────────────────────
    print("\n" + "─" * 40)
    bacang_file = _find_latest_file(
        config.BACANG_DIR, '*CANG*.xlsx'
    )
    if bacang_file:
        data['bacang'] = load_bacang(bacang_file)
    else:
        data['bacang'] = {d: {} for d in range(1, 7)}
        print("  ⚠️  Bỏ qua BA CANG")

    # ─── 4. FFSTOCK ──────────────────────────────────────────
    print("\n" + "─" * 40)
    ffstock_dir = getattr(config, 'FSTOCK_DIR_FFSTOCK', config.FSTOCK_DIR)
    ffstock_file = _find_latest_file(
        ffstock_dir, '*FFSTOCK*.xls*'
    )
    if ffstock_file:
        data['ffstock'] = load_ffstock(ffstock_file)
        data['ffstock_details'] = load_ffstock_details(ffstock_file)
    else:
        data['ffstock'] = {}
        data['ffstock_details'] = {}
        print("  ⚠️  Bỏ qua FFSTOCK")

    # ─── 5. TỒN BỒN ─────────────────────────────────────────
    print("\n" + "─" * 40)
    tonbon_file = _find_latest_file(
        config.TONBON_DIR, '*ton bon*.*'
    )
    if not tonbon_file:
        fallback_dir = getattr(config, 'TONBON_DIR_FALLBACK', None)
        if fallback_dir and fallback_dir != config.TONBON_DIR:
            print(f"  ℹ️  Không tìm thấy file Tồn Bồn trong OneDrive, thử tìm ở thư mục dự phòng: {fallback_dir}")
            tonbon_file = _find_latest_file(fallback_dir, '*ton bon*.*')

    if tonbon_file:
        data['tonbon'] = load_tonbon(tonbon_file)
        data['tonbon_detail'] = load_tonbon_detail(tonbon_file)
    else:
        data['tonbon'] = {}
        data['tonbon_detail'] = {}
        print("  ⚠️  Bỏ qua TỒN BỒN")

    # ─── 6. EMPTY BAG ────────────────────────────────────────
    print("\n" + "─" * 40)
    bag_dir = getattr(config, 'FSTOCK_DIR_EMPTYBAG', config.FSTOCK_DIR)
    bag_file = _find_latest_file(
        bag_dir, '*EMPTY BAG*.xls*'
    )
    if not bag_file:
        fallback_dir = getattr(config, 'FSTOCK_DIR_FALLBACK', None)
        if fallback_dir and fallback_dir != bag_dir:
            print(f"  ℹ️  Không tìm thấy file Bao Bì trong OneDrive, thử tìm ở thư mục dự phòng: {fallback_dir}")
            bag_file = _find_latest_file(fallback_dir, '*EMPTY BAG*.xls*')

    if bag_file:
        data['empty_bag'] = load_empty_bag(bag_file)
    else:
        data['empty_bag'] = {}
        print("  ⚠️  Bỏ qua EMPTY BAG")

    # ─── 7. CÔNG SUẤT, KHÁNG SINH, PELLET MATRIX ─────────────
    print("\n" + "─" * 40)
    plan_file = getattr(config, 'PLAN_FILE', None)
    if plan_file and os.path.isfile(plan_file):
        data['congsuat'] = load_congsuat(plan_file)
        data['stt_khangsinh'] = load_stt_khangsinh(plan_file)
        data['fix_code_pellet'] = load_fix_code_pellet(plan_file)
    else:
        data['congsuat'] = {}
        data['stt_khangsinh'] = {}
        data['fix_code_pellet'] = {}
        print("  ⚠️  Bỏ qua CÔNG SUẤT, STT KHÁNG SINH, FIX CODE PELLET")

    # ─── 8. FEEDCODE ─────────────────────────────────────────
    print("\n" + "─" * 40)
    khsx_file = getattr(config, 'KHSX_FILE', None)
    if khsx_file and os.path.isfile(khsx_file):
        data['feedcode'] = load_feedcode(khsx_file)
    else:
        data['feedcode'] = {}
        print("  ⚠️  Bỏ qua FEEDCODE")

    # ─── 9. KHÁNG SINH ───────────────────────────────────────
    print("\n" + "─" * 40)
    if khsx_file and os.path.isfile(khsx_file):
        data['khangsinh'] = load_khangsinh(khsx_file)
    else:
        data['khangsinh'] = {}
        print("  ⚠️  Bỏ qua KHÁNG SINH")

    # ─── 10. KHSX HÔM QUA ────────────────────────────────────
    print("\n" + "─" * 40)
    # Xác định ngày hôm qua (ngày sản xuất trước đó)
    from datetime import date, datetime
    if target_date is None:
        t = date.today()
    elif isinstance(target_date, datetime):
        t = target_date.date()
    else:
        t = target_date
        
    yesterday_day = t.day - 1
    if yesterday_day < 1:
        yesterday_day = 1  # Fallback ngày 1

    if khsx_file and os.path.isfile(khsx_file):
        data['khsx_yesterday'] = load_khsx_yesterday(khsx_file, yesterday_day)
    else:
        data['khsx_yesterday'] = {}
        print("  ⚠️  Bỏ qua KHSX HÔM QUA")

    # ─── TÓM TẮT ─────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("📊 TÓM TẮT DỮ LIỆU ĐÃ ĐỌC:")
    print(f"  📈 Forecast:      {len(data['forecast'])} sản phẩm")
    print(f"  🚛 Silo Plan:     {sum(len(d) for d in data['silo_plan'].values())} "
          f"sản phẩm × ngày")
    print(f"  📦 Ba Cang:       {sum(len(d) for d in data['bacang'].values())} "
          f"sản phẩm × ngày")
    print(f"  🏪 FFSTOCK:       {len(data['ffstock'])} sản phẩm ({len(data.get('ffstock_details', {}))} có DOH)")
    print(f"  🏗️ Tồn bồn:       {len(data['tonbon'])} mã cám ({len(data.get('tonbon_detail', {}))} bồn chi tiết)")
    print(f"  🎒 Empty Bag:     {len(data['empty_bag'])} loại bao")
    print(f"  ⚙️  Công suất:     {len(data['congsuat'])} sản phẩm")
    print(f"  💊 Cấp độ KS:     {len(data.get('stt_khangsinh', {}))} cấp độ")
    print(f"  🔧 Ma trận Pellet: {len(data.get('fix_code_pellet', {}))} cấu hình")
    print(f"  🔩 Feedcode:      {len(data['feedcode'])} mã")
    print(f"  💊 Kháng sinh SP: {len(data['khangsinh'])} mã")
    print(f"  📋 KHSX hôm qua: {len(data['khsx_yesterday'])} sản phẩm")
    print("=" * 60)

    return data


# ============================================================
# MAIN - Test nhanh
# ============================================================

if __name__ == '__main__':
    import config
    data = load_all_data(config)

    # In mẫu forecast
    if data['forecast']:
        print("\n🔍 MẪU FORECAST (5 sản phẩm đầu):")
        for item in data['forecast'][:5]:
            print(f"  {item.product_code:12s} | "
                  f"Packing={item.packing_size:5s} | "
                  f"Die={item.die_size:.1f} | "
                  f"Total={item.total_with_silo:.1f}T | "
                  f"Silo={item.silo_tons:.1f}T")
