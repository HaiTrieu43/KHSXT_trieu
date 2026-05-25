import openpyxl
import sys

# Fix encoding
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

plan_file = r"D:\Kê hoạch sản xuât\plan\Plan.xlsm"
wb = openpyxl.load_workbook(plan_file, read_only=True, data_only=True)
ws = wb['Fix code pellet']

print("Columns in 'Fix code pellet' row 2:")
row2 = [ws.cell(row=2, column=c).value for c in range(1, 35)]
print(row2)

print("\nSample rows from 'Fix code pellet':")
count = 0
for r in range(3, 100):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 35)]
    # Filter rows related to 566, 567S, 550, or general interesting codes
    code = str(row_vals[1]).strip().upper() if row_vals[1] is not None else ""
    if code and any(k in code for k in ['566', '567S', '550S', '522', '552PRO']):
        print(f"Row {r}: Code={code}, Line_CV={row_vals[3]}, cols[25:31]={row_vals[25:31]}")
        count += 1
        if count > 20:
            break

wb.close()
