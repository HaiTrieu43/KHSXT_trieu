import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FILE = r'D:\Kê hoạch sản xuât\BACANG\KẾ HOẠCH CÁM TUẦN VÕ BÁ CANG 2026 (1).xlsx'
wb = openpyxl.load_workbook(FILE, data_only=True)
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")

for idx, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f"\n{'#' * 100}")
    print(f"SHEET {idx+1}: '{name}' (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'#' * 100}")
    
    max_r = min(60, ws.max_row)
    print(f"\n--- Rows 1-{max_r} ---")
    for r in range(1, max_r + 1):
        row_data = []
        for c in range(1, min(ws.max_column + 1, 40)):
            v = ws.cell(r, c).value
            if v is not None:
                row_data.append(f"C{c}={repr(v)}")
        if row_data:
            print(f"  Row {r}: " + " | ".join(row_data))
        else:
            print(f"  Row {r}: [EMPTY]")

    if ws.max_row > 60:
        mid = ws.max_row // 2
        print(f"\n--- Middle rows ({mid-3} to {mid+3}) ---")
        for r in range(mid-3, mid+4):
            if r < 1 or r > ws.max_row: continue
            row_data = []
            for c in range(1, min(ws.max_column + 1, 40)):
                v = ws.cell(r, c).value
                if v is not None:
                    row_data.append(f"C{c}={repr(v)}")
            if row_data:
                print(f"  Row {r}: " + " | ".join(row_data))

    if ws.max_row > 60:
        start = max(61, ws.max_row - 19)
        print(f"\n--- Last rows ({start} to {ws.max_row}) ---")
        for r in range(start, ws.max_row + 1):
            row_data = []
            for c in range(1, min(ws.max_column + 1, 40)):
                v = ws.cell(r, c).value
                if v is not None:
                    row_data.append(f"C{c}={repr(v)}")
            if row_data:
                print(f"  Row {r}: " + " | ".join(row_data))

    if ws.merged_cells.ranges:
        print(f"\n--- Merged cells (first 15) ---")
        for i, mc in enumerate(ws.merged_cells.ranges):
            if i >= 15: break
            print(f"  {mc}")
