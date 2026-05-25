import openpyxl
import os
import sys
import io

# Fix encoding
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

plan_file = r"D:\Kê hoạch sản xuât\plan\Plan.xlsm"
print(f"Checking {plan_file}...")
if not os.path.exists(plan_file):
    print("Plan file does not exist!")
    exit(1)

wb = openpyxl.load_workbook(plan_file, read_only=True, data_only=True)
print("Sheet names:", wb.sheetnames)

# Inspect 'STT Khang sinh' if exists
sheet_name = None
for s in wb.sheetnames:
    if 'STT' in s.upper() or 'KHANG' in s.upper():
        print(f"Found sheet related to antibiotic: {s}")

if 'STT Khang sinh' in wb.sheetnames:
    ws = wb['STT Khang sinh']
    print("\n--- Content of 'STT Khang sinh' (first 40 rows) ---")
    for r in range(1, 40):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 8)]
        if any(row_vals):
            print(f"Row {r}: {row_vals}")

# Inspect 'Fix code pellet' if exists
if 'Fix code pellet' in wb.sheetnames:
    ws = wb['Fix code pellet']
    print("\n--- Content of 'Fix code pellet' (first 25 rows) ---")
    for r in range(1, 26):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        if any(row_vals):
            print(f"Row {r}: {row_vals}")
            
# Inspect 'KHÁNG SINH' if exists
if 'KHÁNG SINH' in wb.sheetnames:
    ws = wb['KHÁNG SINH']
    print("\n--- Content of 'KHÁNG SINH' (first 20 rows) ---")
    for r in range(1, 21):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 6)]
        if any(row_vals):
            print(f"Row {r}: {row_vals}")

wb.close()
