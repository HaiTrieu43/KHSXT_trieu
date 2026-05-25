import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FILE = r'D:\Kê hoạch sản xuât\KHSX THANG 5-20261.xlsm'
wb = openpyxl.load_workbook(FILE, data_only=True)
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")

for idx, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f"\n{'#' * 120}")
    print(f"SHEET {idx+1}: '{name}' (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'#' * 120}")
    
    max_r = min(80, ws.max_row)
    for r in range(1, max_r + 1):
        row_data = []
        for c in range(1, min(ws.max_column + 1, 60)):
            v = ws.cell(r, c).value
            if v is not None:
                row_data.append(f"C{c}={repr(v)}")
        if row_data:
            print(f"  Row {r}: " + " | ".join(row_data))
        else:
            print(f"  Row {r}: [EMPTY]")

    if ws.max_row > 80:
        mid = ws.max_row // 2
        print(f"\n--- Middle rows ({mid-5} to {mid+5}) ---")
        for r in range(mid-5, mid+6):
            if r < 1 or r > ws.max_row: continue
            row_data = []
            for c in range(1, min(ws.max_column + 1, 60)):
                v = ws.cell(r, c).value
                if v is not None:
                    row_data.append(f"C{c}={repr(v)}")
            if row_data:
                print(f"  Row {r}: " + " | ".join(row_data))

    if ws.max_row > 80:
        start = max(81, ws.max_row - 25)
        print(f"\n--- Last rows ({start} to {ws.max_row}) ---")
        for r in range(start, ws.max_row + 1):
            row_data = []
            for c in range(1, min(ws.max_column + 1, 60)):
                v = ws.cell(r, c).value
                if v is not None:
                    row_data.append(f"C{c}={repr(v)}")
            if row_data:
                print(f"  Row {r}: " + " | ".join(row_data))

    if ws.merged_cells.ranges:
        print(f"\n--- Merged cells (first 15) ---")
        for i, mc in enumerate(ws.merged_cells.ranges):
            if i >= 15: break
            print(f"  {mc}")
