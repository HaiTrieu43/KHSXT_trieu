import openpyxl
import sys
import io
import os

# Đảm bảo in tiếng Việt chuẩn
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FILE_PATH = r'D:\Kê hoạch sản xuât\plan\Plan.xlsm'
OUTPUT_FILE = r'D:\Kê hoạch sản xuât\laptrinh vao\scripts\plan_xlsm_structure.txt'

def inspect_workbook():
    if not os.path.exists(FILE_PATH):
        print(f"Lỗi: Không tìm thấy file {FILE_PATH}")
        return

    print(f"Đang phân tích file: {FILE_PATH}...")
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True, read_only=True)
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(f"=== BÁO CÁO CẤU TRÚC FILE PLAN.XLSM ===\n")
        f.write(f"Đường dẫn: {FILE_PATH}\n")
        f.write(f"Số lượng sheet: {len(wb.sheetnames)}\n")
        f.write(f"Danh sách sheet: {wb.sheetnames}\n")
        f.write(f"======================================\n\n")
        
        for sheet_name in wb.sheetnames:
            f.write(f"--- SHEET: {sheet_name} ---\n")
            ws = wb[sheet_name]
            
            # Quét tìm kích thước thực tế
            row_count = 0
            col_count = 0
            
            # Đọc tối đa 1000 dòng để phân tích cấu trúc cột và dữ liệu mẫu
            sample_rows = []
            for r_idx, row in enumerate(ws.iter_rows(max_row=2000, max_col=100, values_only=True), 1):
                row_values = list(row)
                # Kiểm tra dòng trống
                if any(v is not None for v in row_values):
                    row_count = r_idx
                    # Tìm cột cuối cùng có dữ liệu
                    for c_idx, val in enumerate(row_values, 1):
                        if val is not None:
                            col_count = max(col_count, c_idx)
                    
                    # Lưu 50 dòng đầu tiên có dữ liệu làm mẫu phân tích
                    if len(sample_rows) < 100:
                        sample_rows.append((r_idx, row_values))
            
            f.write(f"Kích thước ước tính: {row_count} dòng x {col_count} cột\n")
            f.write(f"Chi tiết nội dung 100 dòng đầu có dữ liệu:\n")
            
            for r_idx, row_values in sample_rows:
                # Chỉ lấy đến cột cuối có dữ liệu
                truncated_row = row_values[:col_count]
                # Format dòng thành chuỗi đẹp
                formatted_cells = []
                for c_idx, val in enumerate(truncated_row, 1):
                    if val is not None:
                        # Rút ngắn các chuỗi quá dài
                        val_str = str(val).strip()
                        if len(val_str) > 40:
                            val_str = val_str[:37] + "..."
                        formatted_cells.append(f"C{c_idx}[{val_str}]")
                
                if formatted_cells:
                    f.write(f"  Dòng {r_idx:4d}: " + " | ".join(formatted_cells) + "\n")
            
            f.write("\n" + "="*50 + "\n\n")
            print(f"  Hoàn thành đọc sheet: {sheet_name} ({row_count} dòng, {col_count} cột)")
            
    wb.close()
    print(f"Đã lưu kết quả phân tích vào: {OUTPUT_FILE}")

if __name__ == '__main__':
    inspect_workbook()
