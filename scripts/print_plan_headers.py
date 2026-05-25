import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FILE = r'D:\Kê hoạch sản xuât\plan\Plan.xlsm'

def print_plan_headers():
    wb = openpyxl.load_workbook(FILE, data_only=False, read_only=True)
    if 'PLAN' not in wb.sheetnames:
        print("Sheet 'PLAN' not found")
        wb.close()
        return
        
    ws = wb['PLAN']
    
    # Read row 5 and row 6
    rows = list(ws.iter_rows(min_row=5, max_row=6, max_col=100, values_only=True))
    row5 = rows[0]
    row6 = rows[1]
    
    print("=== SHEET 'PLAN' COLUMNS ===")
    for col_idx, (r5_val, r6_val) in enumerate(zip(row5, row6), 1):
        if r5_val is not None or r6_val is not None:
            r5_str = str(r5_val).replace('\n', ' ') if r5_val is not None else ""
            r6_str = str(r6_val).replace('\n', ' ') if r6_val is not None else ""
            print(f"Col {col_idx:2d} (Excel {openpyxl.utils.get_column_letter(col_idx)}): {r5_str} | {r6_str}")
            
    wb.close()

if __name__ == '__main__':
    print_plan_headers()
