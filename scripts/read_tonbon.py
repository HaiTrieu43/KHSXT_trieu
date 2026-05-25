import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FILE = r'D:\Kê hoạch sản xuât\BATCHING-TONBON\Bao cao ton bon thanh pham 05.2026.xlsx'
wb = openpyxl.load_workbook(FILE, data_only=True)
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")

for idx, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f"\n{'#' * 100}")
    print(f"SHEET {idx+1}: '{name}' (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'#' * 100}")
    
    max_rows_to_show = min(50, ws.max_row)
    print(f"\n--- Rows 1-{max_rows_to_show} ---")
    for r in range(1, max_rows_to_show + 1):
        row_data = []
        for c in range(1, min(ws.max_column + 1, 60)):
            v = ws.cell(r, c).value
            if v is not None:
                row_data.append(f"C{c}={repr(v)}")
        if row_data:
            print(f"  Row {r}: " + " | ".join(row_data))
        else:
            print(f"  Row {r}: [EMPTY]")
    
    if ws.max_row > 100:
        mid = ws.max_row // 2
        print(f"\n--- Middle rows ({mid-2} to {mid+2}) ---")
        for r in range(mid-2, mid+3):
            row_data = []
            for c in range(1, min(ws.max_column + 1, 60)):
                v = ws.cell(r, c).value
                if v is not None:
                    row_data.append(f"C{c}={repr(v)}")
            if row_data:
                print(f"  Row {r}: " + " | ".join(row_data))

    if ws.max_row > 50:
        start = max(51, ws.max_row - 19)
        print(f"\n--- Last rows ({start} to {ws.max_row}) ---")
        for r in range(start, ws.max_row + 1):
            row_data = []
            for c in range(1, min(ws.max_column + 1, 60)):
                v = ws.cell(r, c).value
                if v is not None:
                    row_data.append(f"C{c}={repr(v)}")
            if row_data:
                print(f"  Row {r}: " + " | ".join(row_data))

    # Key labels
    print(f"\n--- Key text labels ---")
    count = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and isinstance(v, str) and v.strip():
            row_data = []
            for c in range(1, min(ws.max_column + 1, 25)):
                cv = ws.cell(r, c).value
                if cv is not None:
                    row_data.append(f"C{c}={repr(cv)}")
            print(f"  Row {r}: " + " | ".join(row_data))
            count += 1
            if count > 30:
                print("  ... (truncated)")
                break
    
    if ws.merged_cells.ranges:
        print(f"\n--- Merged cells (first 10) ---")
        for i, mc in enumerate(ws.merged_cells.ranges):
            if i >= 10: break
            print(f"  {mc}")
