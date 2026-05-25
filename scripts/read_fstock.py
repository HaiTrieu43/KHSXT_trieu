import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FILE = r'D:\Kê hoạch sản xuât\FSTOCK-BAG\FFSTOCK 17 -05-2026.xlsm'
wb = openpyxl.load_workbook(FILE, data_only=True)

print(f"FILE: {FILE}")
print(f"Total sheets: {len(wb.sheetnames)}")
print()

for idx, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f"\n{'#' * 100}")
    print(f"SHEET {idx+1}: '{name}' (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'#' * 100}")
    
    # Print first 30 rows
    print(f"\n--- Rows 1-30 ---")
    for r in range(1, min(31, ws.max_row + 1)):
        row_data = []
        for c in range(1, min(ws.max_column + 1, 50)):
            v = ws.cell(r, c).value
            if v is not None:
                row_data.append(f"C{c}={repr(v)}")
        if row_data:
            print(f"  Row {r}: " + " | ".join(row_data))
        else:
            print(f"  Row {r}: [EMPTY]")
    
    # Print rows around middle
    if ws.max_row > 60:
        mid = ws.max_row // 2
        print(f"\n--- Middle rows ({mid-2} to {mid+2}) ---")
        for r in range(mid-2, mid+3):
            row_data = []
            for c in range(1, min(ws.max_column + 1, 50)):
                v = ws.cell(r, c).value
                if v is not None:
                    row_data.append(f"C{c}={repr(v)}")
            if row_data:
                print(f"  Row {r}: " + " | ".join(row_data))
    
    # Print last 15 rows
    if ws.max_row > 30:
        print(f"\n--- Last 15 rows ({ws.max_row - 14} to {ws.max_row}) ---")
        for r in range(max(31, ws.max_row - 14), ws.max_row + 1):
            row_data = []
            for c in range(1, min(ws.max_column + 1, 50)):
                v = ws.cell(r, c).value
                if v is not None:
                    row_data.append(f"C{c}={repr(v)}")
            if row_data:
                print(f"  Row {r}: " + " | ".join(row_data))

    # Text labels
    print(f"\n--- Key text labels in Column A ---")
    count = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and isinstance(v, str):
            vs = v.strip()
            if vs and not vs[0].isdigit():
                row_data = []
                for c in range(1, min(ws.max_column + 1, 20)):
                    cv = ws.cell(r, c).value
                    if cv is not None:
                        row_data.append(f"C{c}={repr(cv)}")
                print(f"  Row {r}: " + " | ".join(row_data))
                count += 1
                if count > 30:
                    print("  ... (truncated)")
                    break

    # Merged cells
    if ws.merged_cells.ranges:
        print(f"\n--- Merged cells (first 15) ---")
        for i, mc in enumerate(ws.merged_cells.ranges):
            if i >= 15:
                print(f"  ... and more")
                break
            print(f"  {mc}")
