import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FILE = r'D:\Kê hoạch sản xuât\plan\Plan.xlsm'
wb = openpyxl.load_workbook(FILE, data_only=True, read_only=True)
print(f"Sheet names: {wb.sheetnames}")

# Find CONGSUAT sheet
target = None
for name in wb.sheetnames:
    if 'CONG' in name.upper() or 'SUAT' in name.upper() or 'CS' in name.upper():
        target = name
        print(f"Found: '{name}'")

if target:
    ws = wb[target]
    for r_idx, row in enumerate(ws.iter_rows(max_row=200, max_col=20, values_only=False), 1):
        row_data = []
        for cell in row:
            if cell.value is not None:
                row_data.append(f"C{cell.column}={repr(cell.value)}")
        if row_data:
            print(f"  Row {r_idx}: " + " | ".join(row_data))
else:
    print("CONGSUAT not found, listing all sheets with first rows:")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- Sheet: '{name}' ---")
        for r_idx, row in enumerate(ws.iter_rows(max_row=3, max_col=10, values_only=False), 1):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append(f"C{cell.column}={repr(cell.value)}")
            if row_data:
                print(f"  Row {r_idx}: " + " | ".join(row_data))

wb.close()
