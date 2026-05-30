"""
khsx_auto.py - Chương trình chính: Tự động lập Kế Hoạch Sản Xuất
C.P. Vietnam - Chi nhánh Bình Dương

Cách chạy:
    python khsx_auto.py
    python khsx_auto.py --day 3        # Lập KH cho thứ 3 trong tuần
    python khsx_auto.py --date 20260520  # Lập KH cho ngày cụ thể
"""
import sys
import os
import io
import argparse
from datetime import datetime

# Fix encoding cho Windows console
if sys.platform == 'win32':
    if not isinstance(sys.stdout, io.TextIOWrapper) or sys.stdout.encoding != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    if not isinstance(sys.stderr, io.TextIOWrapper) or sys.stderr.encoding != 'utf-8':
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# Thêm thư mục hiện tại vào path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import config
from models import KHSXResult
from data_loader import load_all_data
from demand_calculator import calculate_daily_demand
from constraint_solver import solve_constraints
from packaging_allocator import allocate_packaging
from output_generator import generate_khsx_excel, print_khsx_console


def main():
    """Chương trình chính"""
    
    # ============================================
    # PARSE ARGUMENTS
    # ============================================
    parser = argparse.ArgumentParser(description='Tự động lập KHSX')
    parser.add_argument('--day', type=int, default=None,
                       help='Ngày trong tuần (1=T2, 6=T7). Mặc định: hôm nay')
    parser.add_argument('--date', type=str, default=None,
                       help='Ngày cụ thể (YYYYMMDD). Mặc định: hôm nay')
    parser.add_argument('--walkin', type=str, default=None,
                       help='File đơn khách vãng lai (CSV: product,tons,packing)')
    parser.add_argument('--no-excel', action='store_true',
                       help='Không xuất file Excel, chỉ in console')
    args = parser.parse_args()
    
    # Xác định ngày
    if args.date:
        today = datetime.strptime(args.date, '%Y%m%d')
    else:
        today = datetime.today()
    
    # Xác định thứ trong tuần (1=Mon, 6=Sat)
    if args.day:
        day_of_week = args.day
    else:
        dow = today.weekday()  # 0=Mon, 6=Sun
        day_of_week = dow + 1  # 1=Mon, 7=Sun
        if day_of_week > 6:
            day_of_week = 6  # Chủ nhật → xem như Thứ 7
    
    date_str = today.strftime('%d-%m-%Y')
    
    print()
    print("╔════════════════════════════════════════════════════════╗")
    print("║   HỆ THỐNG LẬP KẾ HOẠCH SẢN XUẤT TỰ ĐỘNG           ║")
    print("║   C.P. VIETNAM - CHI NHÁNH BÌNH DƯƠNG                ║")
    print("╠════════════════════════════════════════════════════════╣")
    print(f"║   Ngày: {date_str}  |  Thứ: {day_of_week + 1}                      ║")
    print("╚════════════════════════════════════════════════════════╝")
    
    # ============================================
    # BƯỚC 1: ĐỌC DỮ LIỆU
    # ============================================
    try:
        data = load_all_data(config, target_date=today)
    except Exception as e:
        print(f"\n❌ LỖI ĐỌC DỮ LIỆU: {e}")
        return 1
    
    # ============================================
    # BƯỚC 2: ĐỌC ĐƠN KHÁCH VÃNG LAI (nếu có)
    # ============================================
    walkin_orders = []
    if args.walkin and os.path.exists(args.walkin):
        print(f"\n📋 Đọc đơn khách vãng lai: {args.walkin}")
        with open(args.walkin, 'r', encoding='utf-8') as f:
            for line in f:
                parts = line.strip().split(',')
                if len(parts) >= 2:
                    walkin_orders.append({
                        'product': parts[0].strip().upper(),
                        'tons': float(parts[1].strip()),
                        'packing_size': parts[2].strip() if len(parts) > 2 else '25',
                    })
        print(f"  ✅ Đọc được {len(walkin_orders)} đơn hàng vãng lai")
    
    # ============================================
    # BƯỚC 3: TÍNH NHU CẦU SẢN XUẤT
    # ============================================
    # Tính produced_this_week từ KHSX các ngày trước trong tuần
    produced_this_week = {}
    import openpyxl
    from datetime import timedelta
    weekday = today.weekday()  # 0 = Mon, 5 = Sat, 6 = Sun
    if weekday > 5:
        weekday = 5  # Coi chủ nhật như thứ 7
    monday_date = today - timedelta(days=weekday)
    
    def _get_tons_from_batches(product_code, batches_val):
        if batches_val is None:
            return 0.0
        try:
            b = float(batches_val)
        except (ValueError, TypeError):
            return 0.0
        
        prod_upper = str(product_code).strip().upper()
        # Đối với 325F hoặc các mã bắt đầu bằng 550, 551 thì 8.0 tấn/mẻ, các mã khác là 8.4 tấn/mẻ
        if prod_upper == '325F' or prod_upper.startswith('550') or prod_upper.startswith('551'):
            return b * 8.0
        else:
            return b * 8.4

    print("\n🔄 Tính toán sản lượng đã lên kế hoạch từ đầu tuần đến nay:")
    for d_idx in range(weekday):
        prev_date = monday_date + timedelta(days=d_idx)
        prev_date_str = prev_date.strftime('%d-%m-%Y')
        import glob
        pattern = os.path.join(config.OUTPUT_DIR, f"KHSX_{prev_date_str}*.xlsx")
        files = glob.glob(pattern)
        if files:
            # Lấy file có ngày cập nhật mới nhất (phòng trường hợp lưu bản sao _v123)
            files.sort(key=os.path.getmtime, reverse=True)
            prev_file = files[0]
            print(f"  📖 Đọc file: {os.path.basename(prev_file)}")
            try:
                # Đọc nhanh bằng openpyxl (data_only=True)
                prev_wb = openpyxl.load_workbook(prev_file, data_only=True, read_only=True)
                prev_ws = prev_wb.active
                # Đọc từ dòng 7 đến 41 (các cột B = Tên cám, C = Kế hoạch mẻ)
                for r in range(7, 42):
                    prod = prev_ws.cell(row=r, column=2).value  # B
                    batches_val = prev_ws.cell(row=r, column=3).value  # C
                    if prod:
                        prod_code = str(prod).strip().upper().replace(' ', '')
                        if prod_code:
                            t_val = _get_tons_from_batches(prod_code, batches_val)
                            produced_this_week[prod_code] = produced_this_week.get(prod_code, 0.0) + t_val
                prev_wb.close()
            except Exception as ex:
                print(f"  ⚠️ Lỗi đọc file {os.path.basename(prev_file)}: {ex}")
        else:
            print(f"  ℹ️ Không tìm thấy file KHSX cho ngày {prev_date_str}")
            
    if produced_this_week:
        print(f"  ✅ Tổng cộng đã lên kế hoạch {len(produced_this_week)} mã cám từ đầu tuần.")
    else:
        print("  ℹ️ Chưa có sản lượng kế hoạch nào từ đầu tuần.")

    demand_list = calculate_daily_demand(
        today_date=today,
        day_of_week=day_of_week,
        forecast=data['forecast'],
        silo_plan=data['silo_plan'],
        bacang=data['bacang'],
        walkin_orders=walkin_orders,
        ffstock=data['ffstock'],
        tonbon=data['tonbon'],
        khsx_yesterday=data['khsx_yesterday'],
        congsuat=data['congsuat'],
        produced_this_week=produced_this_week,
        ffstock_details=data.get('ffstock_details', {}),
        adjustments=data.get('adjustments'),
    )
    
    # ============================================
    # BƯỚC 4: KIỂM TRA RÀNG BUỘC
    # ============================================
    demand_list, warnings = solve_constraints(
        demand_list=demand_list,
        empty_bag=data['empty_bag'],
        congsuat=data['congsuat'],
        min_tons=config.MIN_DAILY_TONS,
        max_tons=config.MAX_DAILY_TONS,
        target_tons=config.TARGET_DAILY_TONS,
        ffstock_details=data.get('ffstock_details', {}),
    )
    
    # ============================================
    # BƯỚC 5: PHÂN BỔ BAO BÌ
    # ============================================
    allocate_packaging(demand_list, data['forecast'], adjustments=data.get('adjustments'))
    
    # ============================================
    # BƯỚC 6: GÁN LINE MÁY PELLET + KHÁNG SINH
    # ============================================
    feedcode = data['feedcode']
    khangsinh = data['khangsinh']
    fix_code_pellet = data.get('fix_code_pellet', {})
    congsuat = data.get('congsuat', {})
    
    # Định nghĩa hàm đối sánh cấu hình pellet thông minh
    def find_pellet_config(product_code, packing_size, die_size, is_mash, matrix):
        p_code = str(product_code).strip().upper()
        pack = str(packing_size).strip().upper() if packing_size else '25'
        
        # 1. Thử các key kết hợp chính xác
        if is_mash:
            key = f"{p_code}{pack}M"
            if key in matrix:
                return matrix[key]
                
        if die_size:
            die_str = str(die_size)
            if die_str.endswith('.0'):
                die_str_alt = die_str[:-2]
            else:
                die_str_alt = die_str + ".0"
                
            keys_to_try = [
                f"{p_code}{pack}{die_str}",
                f"{p_code}{pack}{die_str_alt}"
            ]
            if pack == 'SILO':
                keys_to_try.append(f"{p_code}SILOSILO")
                
            for k in keys_to_try:
                if k in matrix:
                    return matrix[k]
                    
        # 2. Thử tìm gần đúng: bắt đầu bằng p_code và chứa pack
        prefix = f"{p_code}{pack}"
        matches = []
        for k, cfg in matrix.items():
            if k.startswith(prefix):
                matches.append((k, cfg))
        if matches:
            return matches[0][1]
            
        # 3. Thử tìm rộng hơn: chỉ bắt đầu bằng p_code
        matches_broad = []
        for k, cfg in matrix.items():
            if k.startswith(p_code):
                matches_broad.append((k, cfg))
        if matches_broad:
            return matches_broad[0][1]
            
        return None

    # Bảng ưu tiên máy ép viên
    LINE_ORDER = {
        'PL1': 1, '1': 1,
        'PL2': 2, '2': 2,
        'PL3': 3, '3': 3,
        'PL4': 4, '4': 4,
        'PL5': 5, '5': 5,
        'PL6': 6, '6': 6,
        'PL7': 7, '7': 7,
        'M': 8, 'PLM': 8,
        '': 99
    }
    
    print("\n⚙️ Phân bổ máy ép viên tối ưu và áp dụng ràng buộc kỹ thuật:")
    for item in demand_list:
        # Lấy thông tin từ bảng công suất để biết die_size và line_cv mặc định
        spec = congsuat.get(item.product_code, None)
        die_size = spec.die_size if spec else 0.0
        line_cv_default = spec.line_cv if spec else ''
        is_mash = (line_cv_default == 'M')
        
        # Tìm cấu hình gán máy Pellet từ ma trận
        cfg = find_pellet_config(item.product_code, item.packing_size, die_size, is_mash, fix_code_pellet)
        
        assigned_line = ''
        if cfg and cfg['priorities']:
            # Duyệt từ Ưu tiên 1 -> 5
            for p_line in cfg['priorities']:
                norm_line = str(p_line).strip().upper()
                if norm_line.isdigit():
                    norm_line = f"PL{norm_line}"
                    
                # Ràng buộc loại trừ PL2 cho 566 và 567S
                is_excluded = False
                if item.product_code.startswith('566') or item.product_code.startswith('567S'):
                    if norm_line in {'PL2', '2'}:
                        is_excluded = True
                        
                if not is_excluded:
                    assigned_line = norm_line
                    break
                    
        # Nếu chưa gán được, dùng default_line từ ma trận
        if not assigned_line and cfg and cfg['default_line']:
            default_line = str(cfg['default_line']).strip().upper()
            if default_line.isdigit():
                default_line = f"PL{default_line}"
            # Kiểm tra loại trừ
            is_excluded = False
            if item.product_code.startswith('566') or item.product_code.startswith('567S'):
                if default_line in {'PL2', '2'}:
                    is_excluded = True
            if not is_excluded:
                assigned_line = default_line
                
        # Nếu vẫn chưa gán được, dùng line_cv từ feedcode
        if not assigned_line:
            fc_data = feedcode.get(item.product_code, {})
            assigned_line = str(fc_data.get('line_cv', '')).strip().upper()
            if assigned_line.isdigit():
                assigned_line = f"PL{assigned_line}"
                
        # Ràng buộc loại trừ PL2 cưỡng chế cuối cùng
        if item.product_code.startswith('566') or item.product_code.startswith('567S'):
            if assigned_line in {'PL2', '2', ''}:
                old_line = assigned_line or 'chưa xác định'
                assigned_line = 'PL3'
                msg = f"⚠️ Cám {item.product_code} bị loại trừ khỏi máy {old_line} (thiếu sifter bụi mịn), đã tự động chuyển sang máy {assigned_line}."
                warnings.append(msg)
                print(f"  {msg}")
                
        # Chuẩn hóa assigned_line để lưu trữ sạch
        if assigned_line.isdigit():
            assigned_line = f"PL{assigned_line}"
        item.line_cv = assigned_line
        
        # Gán Line đóng bao (line_pk)
        fc_data = feedcode.get(item.product_code, {})
        item.line_pk = fc_data.get('line_pk', '')
        
        # QUY TẮC: Nếu 100% SILO (không đóng bao) → LINE PK = "SILO"
        silo_tons = item.silo_truck or 0
        bag_tons = item.tons - silo_tons
        if bag_tons <= 0.01 and silo_tons > 0:
            item.line_pk = 'SILO'
            
        # Gán Kháng Sinh & Cấp độ kháng sinh (1-26)
        from data_loader import resolve_antibiotic_for_product
        ks_code, ks_level = resolve_antibiotic_for_product(item.product_code, khangsinh)
        item.ks_code = ks_code
        item.ks_level = ks_level
    
    # ============================================
    # BƯỚC 7: SẮP XẾP MIXER SEQUENCE XEN KẼ (INTERLEAVED LINE ORDER)
    # ============================================
    # Sắp xếp: nhóm theo cấp độ kháng sinh (ks_level) tăng dần (sạch trước, thuốc sau)
    # Trong cùng ks_level: chia giỏ theo line_cv và trộn xen kẽ (round-robin)
    # để phân tán đều mẻ trộn cho các máy, tránh dồn ép tải vào 1 máy
    print("\n🧬 Sắp xếp Mixer sequence theo an toàn sinh học kháng sinh và xen kẽ máy CV...")
    from collections import defaultdict
    ks_groups = defaultdict(list)
    for item in demand_list:
        ks_groups[item.ks_level].append(item)
        
    sorted_interleaved = []
    for ks_lvl in sorted(ks_groups.keys()):
        items_in_ks = ks_groups[ks_lvl]
        
        # Chia thành các giỏ theo line_cv
        line_buckets = defaultdict(list)
        for item in items_in_ks:
            line_buckets[item.line_cv].append(item)
            
        # Sắp xếp các sản phẩm trong mỗi giỏ theo tấn giảm dần (mẻ lớn chạy trước ổn định máy)
        for line in line_buckets:
            line_buckets[line].sort(key=lambda x: -x.tons)
            
        # Sắp xếp các line_cv theo thứ tự để round-robin có hệ thống
        active_lines = sorted(list(line_buckets.keys()), key=lambda l: LINE_ORDER.get(str(l).strip().upper(), 99))
        
        # Round-robin xen kẽ các dòng máy
        ks_interleaved = []
        while any(len(line_buckets[l]) > 0 for l in active_lines):
            for line in active_lines:
                if len(line_buckets[line]) > 0:
                    ks_interleaved.append(line_buckets[line].pop(0))
                    
        sorted_interleaved.extend(ks_interleaved)
        
    demand_list = sorted_interleaved
    
    # ============================================
    # BƯỚC 7.5: LẬP KẾ HOẠCH PELLET (PL) CHI TIẾT THEO CA
    # ============================================
    from pellet_planner import plan_pellet_shifts
    pellet_plan_res = plan_pellet_shifts(
        demand_list=demand_list,
        tonbon_detail=data.get('tonbon_detail', {}),
        congsuat_dict=data.get('congsuat', {}),
        fix_code_pellet_dict=data.get('fix_code_pellet', {}),
        khangsinh_dict=data.get('khangsinh', {}),
        target_date=today,
        feedcode_dict=data.get('feedcode', {}),
        code_mapping=data.get('code_mapping', {})
    )


    # ============================================
    # BƯỚC 8: TẠO KẾ QUẢ
    # ============================================
    result = KHSXResult(
        date=date_str,
        items=demand_list,
        warnings=warnings,
    )
    
    # ============================================
    # BƯỚC 9: XUẤT KẾT QUẢ
    # ============================================
    # In ra console
    print_khsx_console(result)
    
    # Xuất Excel
    if not args.no_excel:
        filepath = generate_khsx_excel(
            result=result,
            output_dir=config.OUTPUT_DIR,
            pellet_plan_result=pellet_plan_res,
            congsuat=data.get('congsuat', {}),
            fix_code_pellet=data.get('fix_code_pellet', {})
        )
        print(f"\n✅ HOÀN TẤT! File KHSX: {filepath}")
    
    print(f"\n{result.summary()}")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
