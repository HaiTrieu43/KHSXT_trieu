"""
output_generator.py - Xuất kế hoạch sản xuất ra file Excel chất lượng cao dựa trên template
"""
import os
import sys
import io
from typing import Optional
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from models import KHSXResult, DemandItem
import config

# Fix encoding cho Windows
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass


def generate_khsx_excel(result, output_dir, filename=None, pellet_plan_result=None, congsuat=None, fix_code_pellet=None):
    """
    Xuất KHSX ra file Excel theo format nhà máy bằng cách sử dụng file mẫu (KHSX_FILE).
    
    Args:
        result: KHSXResult
        output_dir: Thư mục output
        filename: Tên file (None = auto)
    
    Returns:
        str: Đường dẫn file đã tạo
    """
    if filename is None:
        filename = f"KHSX_{result.date}.xlsx"
    
    filepath = os.path.join(output_dir, filename)
    
    # 1. Load workbook template (KHSX_FILE)
    print(f"  📖 Đang đọc file mẫu: {config.KHSX_FILE}...")
    wb = openpyxl.load_workbook(config.KHSX_FILE, data_only=False)
    
    # 2. Sao chép sheet '19' làm template cho sheet ngày mới
    if '19' in wb.sheetnames:
        template_sheet_name = '19'
    else:
        # Fallback nếu không thấy sheet 19, lấy sheet đầu tiên
        template_sheet_name = [s for s in wb.sheetnames if s.isdigit()][0]
        
    print(f"  📋 Sử dụng sheet '{template_sheet_name}' làm layout mẫu...")
    template_sheet = wb[template_sheet_name]
    new_sheet = wb.copy_worksheet(template_sheet)
    new_sheet.title = result.date
    
    # 3. Di chuyển sheet mới lên vị trí đầu tiên (index 0)
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(new_sheet)))
    
    # 4. Định dạng và điền ngày vào cell U3
    # result.date dạng DD-MM-YYYY, đổi sang Ngày:…DD…/……MM….. /…YYYY
    parts = result.date.split('-')
    if len(parts) == 3:
        day = parts[0]
        month = str(int(parts[1]))  # Bỏ số 0 ở đầu nếu có
        year = parts[2]
        date_formatted = f"Ngày:…{day}…/……{month}…../…{year}"
    else:
        date_formatted = f"Ngày: {result.date}"
        
    new_sheet['U3'].value = date_formatted
    new_sheet['U3'].font = Font(name='Times New Roman', size=11, bold=True)
    new_sheet['U3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 5. Phủ màu theo mức độ ưu tiên để giao diện trực quan và chuyên nghiệp
    priority_colors = {
        1: PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid'),  # Đỏ nhạt (Silo)
        2: PatternFill(start_color='FFE8CC', end_color='FFE8CC', fill_type='solid'),  # Cam nhạt (Vãng lai)
        3: PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),  # Vàng nhạt (Bù thiếu)
        4: PatternFill(start_color='FFD2D2', end_color='FFD2D2', fill_type='solid'),  # Hồng cảnh báo (Đứt hàng DOH < 3)
        5: PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid'),  # Xanh lá cây nhạt (Forecast thường)
    }
    no_fill = PatternFill(fill_type=None)
    
    # Fonts tiêu chuẩn của nhà máy
    font_regular = Font(name='Times New Roman', size=12, bold=False)
    font_bold = Font(name='Times New Roman', size=12, bold=True)
    
    # 6. Ghi dữ liệu vào lưới 35 dòng (dòng 7 đến dòng 41)
    data_start = 7
    data_end = 41
    
    for r in range(data_start, data_end + 1):
        idx = r - data_start
        item: Optional[DemandItem] = result.items[idx] if idx < len(result.items) else None
        
        # Thiết lập cột STT (cột A)
        cell_stt = new_sheet[f'A{r}']
        cell_stt.value = r - 6
        cell_stt.font = font_bold
        cell_stt.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thiết lập cột TÊN CÁM (cột B)
        cell_name = new_sheet[f'B{r}']
        cell_name.value = item.product_code if item else None
        cell_name.font = font_bold
        cell_name.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thiết lập cột KẾ HOẠCH MẺ (cột C)
        cell_batches = new_sheet[f'C{r}']
        cell_batches.value = item.batches if item else None
        cell_batches.font = font_bold
        cell_batches.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thiết lập cột TỔNG TẤN (cột D) - dùng giá trị thực thay vì công thức
        # vì sản phẩm gộp (SILO + đóng bao) có tons ≠ batches × tpb cố định
        cell_tons = new_sheet[f'D{r}']
        cell_tons.value = item.tons if item else None
        cell_tons.font = font_bold
        cell_tons.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thiết lập cột bao bì (Cột E đến S) và Silo Truck (Cột T)
        cols_pkg = {
            'E': item.higro_25 if item else 0.0,
            'F': item.higro_40 if item else 0.0,
            'G': item.cp_25 if item else 0.0,
            'H': item.cp_40 if item else 0.0,
            'I': item.star_25 if item else 0.0,
            'J': item.star_40 if item else 0.0,
            'K': item.nuvo_25 if item else 0.0,
            'L': item.nuvo_40 if item else 0.0,
            'M': item.bell_25 if item else 0.0,
            'N': item.bell_40 if item else 0.0,
            'O': item.nasa_25 if item else 0.0,
            'P': item.nasa_40 if item else 0.0,
            'Q': item.white_bag_25 if item else 0.0,
            'R': item.white_bag_40 if item else 0.0,
            'S': item.white_bag_50 if item else 0.0,
            'T': item.silo_truck if item else 0.0,
        }
        
        for col_letter, val in cols_pkg.items():
            cell_pkg = new_sheet[f'{col_letter}{r}']
            cell_pkg.value = val if (val and val > 0) else None
            cell_pkg.font = font_regular
            cell_pkg.alignment = Alignment(horizontal='center', vertical='center')
            
        # Thiết lập giá trị trực tiếp cho các cột (tránh lỗi công thức VLOOKUP lệch kiểu dữ liệu hoặc ghi đè ràng buộc thông minh)
        # Cột U: KHÁNG SINH (sử dụng công thức Excel động liên kết với sheet KHÁNG SINH)
        new_sheet[f'U{r}'].value = f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r}, \'KHÁNG SINH\'!$B$3:$C$2000, 2, 0), "SẠCH (KHÔNG KS)"))'
        new_sheet[f'U{r}'].font = font_regular
        new_sheet[f'U{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột V: LINE CV
        new_sheet[f'V{r}'].value = item.line_cv if (item and item.line_cv) else None
        new_sheet[f'V{r}'].font = font_regular
        new_sheet[f'V{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột W: LINE PK
        new_sheet[f'W{r}'].value = item.line_pk if (item and item.line_pk) else None
        new_sheet[f'W{r}'].font = font_regular
        new_sheet[f'W{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột X: CHÊNH LỆCH
        new_sheet[f'X{r}'].value = f'=IF(D{r}="","",D{r}-SUM(E{r}:T{r}))'
        new_sheet[f'X{r}'].font = font_regular
        new_sheet[f'X{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột Y: Dòng trống phân tách
        new_sheet[f'Y{r}'].value = None
        
        # Cột Z: THỰC HIỆN MẺ
        new_sheet[f'Z{r}'].value = f'=IFERROR(VLOOKUP(B{r},$AC$7:$AD$41,2,0),0)'
        new_sheet[f'Z{r}'].font = font_regular
        new_sheet[f'Z{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột AA: HOÀN THÀNH (%)
        new_sheet[f'AA{r}'].value = f'=IF(C{r}=0,0,Z{r}/C{r}*100)'
        new_sheet[f'AA{r}'].font = font_regular
        new_sheet[f'AA{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cột AB: LÝ DO CHI TIẾT
        new_sheet[f'AB{r}'].value = None
        new_sheet[f'AB{r}'].font = font_regular
        new_sheet[f'AB{r}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Cột AC và AD: Dữ liệu thực tế pha trộn (để trống chờ nhập)
        new_sheet[f'AC{r}'].value = None
        new_sheet[f'AC{r}'].font = font_regular
        new_sheet[f'AC{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        new_sheet[f'AD{r}'].value = None
        new_sheet[f'AD{r}'].font = font_regular
        new_sheet[f'AD{r}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Áp dụng màu nền hàng theo độ ưu tiên
        row_fill = priority_colors.get(item.priority.value) if item else no_fill
        for c in range(1, 31):  # Cột A đến AD (cột 1 đến 30)
            new_sheet.cell(row=r, column=c).fill = row_fill
            
    # Điền công thức VLOOKUP cột U cho các dòng từ 42 đến 450 (bỏ qua các dòng chữ ký và footer từ 42 đến 50)
    for r in range(42, 451):
        if r in {42, 43, 44, 45, 46, 47, 48, 49, 50}:
            continue
        new_sheet[f'U{r}'].value = f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r}, \'KHÁNG SINH\'!$B$3:$C$2000, 2, 0), "SẠCH (KHÔNG KS)"))'
        new_sheet[f'U{r}'].font = font_regular
        new_sheet[f'U{r}'].alignment = Alignment(horizontal='center', vertical='center')
            
    # 7. Cập nhật các dòng tổng kết và ký duyệt ở chân trang (Footer)
    # Dòng 42: TỔNG (TOTAL)
    new_sheet['A42'].value = 'TỔNG (TOTAL)'
    new_sheet['A42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['A42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['C42'].value = '=IF(SUM(C7:C41)=0,"",SUM(C7:C41))'
    new_sheet['C42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['C42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['D42'].value = '=IF(SUM(D7:D41)=0,"",SUM(D7:D41))'
    new_sheet['D42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['D42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['Z42'].value = '=SUM(Z7:Z41)'
    new_sheet['Z42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['Z42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AA42'].value = '=IF(C42=0,0,(Z42/C42)*100)'
    new_sheet['AA42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['AA42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AD42'].value = '=SUM(AD7:AD41)'
    new_sheet['AD42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['AD42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AE42'].value = '=$AD$42-$C$42'
    new_sheet['AE42'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['AE42'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AF42'].value = 'SỐ MẺ'
    new_sheet['AF42'].font = Font(name='Times New Roman', size=12, bold=True)
    new_sheet['AF42'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Dòng 43: Sổ sách chất lượng và tỷ lệ blender
    new_sheet['A43'].value = 'QT-SX-01/BM04\nLần ban hành: 04\nNgày hiệu lực: 01/10//2025'
    new_sheet['A43'].font = Font(name='Times New Roman', size=12, bold=False)
    new_sheet['A43'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    new_sheet['H43'].value = 'NGƯỜI LẬP KẾ HOẠCH '
    new_sheet['H43'].font = Font(name='Times New Roman', size=12, bold=True)
    new_sheet['H43'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['U43'].value = 'NGƯỜI THẨM TRA'
    new_sheet['U43'].font = Font(name='Times New Roman', size=12, bold=True)
    new_sheet['U43'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AE43'].value = '=IF(C42=0,0,($AD$42/$C$42)*100)'
    new_sheet['AE43'].font = Font(name='Times New Roman', size=14, bold=True)
    new_sheet['AE43'].alignment = Alignment(horizontal='center', vertical='center')
    
    new_sheet['AF43'].value = '%'
    new_sheet['AF43'].font = Font(name='Times New Roman', size=12, bold=True)
    new_sheet['AF43'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Dòng 47: Tên người ký lập kế hoạch
    new_sheet['H47'].value = 'Hồ Đăng Xuân Thành'
    new_sheet['H47'].font = Font(name='Times New Roman', size=9, bold=True)
    new_sheet['H47'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 7.5. Chèn động các mã cám mới và kháng sinh vào sheet 'KHÁNG SINH'
    if 'KHÁNG SINH' in wb.sheetnames:
        ks_sheet = wb['KHÁNG SINH']
        existing_codes = set()
        
        # Tìm dòng cuối cùng thực tế có dữ liệu để tránh ghi đè dữ liệu cũ
        max_filled_row = 2
        for r in range(3, 2000):
            cell = ks_sheet.cell(row=r, column=2)
            val = cell.value
            if val is not None and str(val).strip() != '':
                # Đồng bộ kiểu dữ liệu thành String để tránh lỗi VLOOKUP lệch kiểu (str vs int) trong Excel
                cell.value = str(val).strip()
                max_filled_row = r
                existing_codes.add(str(val).strip().upper().replace(' ', ''))
        empty_row = max_filled_row + 1
                    
        for item in result.items:
            if not item or not item.product_code:
                continue
            p_code = str(item.product_code).strip().upper().replace(' ', '')
            if p_code and p_code not in existing_codes:
                ks_sheet.cell(row=empty_row, column=1).value = empty_row - 2
                ks_sheet.cell(row=empty_row, column=2).value = str(item.product_code).strip()
                ks_sheet.cell(row=empty_row, column=3).value = item.ks_code if item.ks_code else 'KS/ HC (1)/(2)'
                existing_codes.add(p_code)
                empty_row += 1

    # Đồng bộ hóa kiểu dữ liệu cho sheet FEEDCODE để tránh lỗi VLOOKUP lệch kiểu tương tự
    if 'FEEDCODE' in wb.sheetnames:
        fd_sheet = wb['FEEDCODE']
        for r in range(2, fd_sheet.max_row + 1):
            cell = fd_sheet.cell(row=r, column=2)  # Cột B = Tên cám
            val = cell.value
            if val is not None and str(val).strip() != '':
                cell.value = str(val).strip()

    # Ghi dữ liệu Kế hoạch Pellet (PL) và cám bột
    if pellet_plan_result:
        print("\n⚙️ Đang ghi Kế hoạch Pellet (PL) và Mash Feed sang sheet 'KẾ HOẠCH PL'...")
        # 1. Mở file mẫu khpl.xlsx
        khpl_template_path = os.path.join(config.DATA_DIR, 'plan', 'khpl.xlsx')
        if not os.path.exists(khpl_template_path):
            khpl_template_path = os.path.join(config.DATA_DIR, 'plan', 'Plan.xlsm')
            
        if os.path.exists(khpl_template_path):
            print(f"  📖 Đọc sheet 'KẾ HOẠCH PL' từ file mẫu: {os.path.basename(khpl_template_path)}...")
            src_wb = openpyxl.load_workbook(khpl_template_path, data_only=True)
            if 'KẾ HOẠCH PL' in src_wb.sheetnames:
                src_ws = src_wb['KẾ HOẠCH PL']
                
                # Tạo sheet mới KẾ HOẠCH PL trong dest_wb
                dest_ws = wb.create_sheet(title='KẾ HOẠCH PL')
                
                # Copy định dạng từ src sang dest
                from copy import copy
                for col, dim in src_ws.column_dimensions.items():
                    dest_ws.column_dimensions[col].width = dim.width
                    
                for row_idx, dim in src_ws.row_dimensions.items():
                    dest_ws.row_dimensions[row_idx].height = dim.height
                    
                for r in range(1, src_ws.max_row + 1):
                    for c in range(1, src_ws.max_column + 1):
                        cell_src = src_ws.cell(row=r, column=c)
                        cell_dest = dest_ws.cell(row=r, column=c)
                        cell_dest.value = cell_src.value
                        
                        if cell_src.has_style:
                            try:
                                cell_dest.font = copy(cell_src.font)
                                cell_dest.fill = copy(cell_src.fill)
                                cell_dest.border = copy(cell_src.border)
                                cell_dest.alignment = copy(cell_src.alignment)
                                cell_dest.number_format = copy(cell_src.number_format)
                            except Exception:
                                pass
                                
                # Thực hiện gộp ô sau khi đã điền xong dữ liệu và style để tránh lỗi MergedCell
                for merged_range in src_ws.merged_cells.ranges:
                    try:
                        dest_ws.merge_cells(str(merged_range))
                    except Exception:
                        pass
                
                # Xóa sạch các dữ liệu động cũ trong các dòng 3-16 cho các máy PL, cột B-AC
                for r in range(3, 17):
                    for c in range(2, 30):
                        dest_ws.cell(row=r, column=c).value = None
                        
                # Xóa dữ liệu cám bột cũ (cột AD-AH dòng 3 đến 45)
                for r in range(3, 46):
                    for c in range(30, 35):
                        dest_ws.cell(row=r, column=c).value = None

                # Cập nhật ngày tháng năm
                parts_dt = result.date.split('-')
                if len(parts_dt) == 3:
                    dest_ws.cell(row=1, column=23).value = f"{parts_dt[2]}-{parts_dt[1]}-{parts_dt[0]}"
                else:
                    dest_ws.cell(row=1, column=23).value = result.date

                # Ghi Tồn Đầu vào dòng 3-4
                ton_dau_data = pellet_plan_result.get('ton_dau_by_machine', {})
                machine_cols = {
                    'PL1': 2, 'PL2': 6, 'PL3': 10, 'PL4': 14, 'PL5': 18, 'PL6': 22, 'PL7': 26
                }
                
                from pellet_planner import get_tph
                c_dict = congsuat or {}
                f_dict = fix_code_pellet or {}
                
                for m, items in ton_dau_data.items():
                    col_idx = machine_cols.get(m)
                    if not col_idx:
                        continue
                    for idx_td, td in enumerate(items[:2]):
                        r_idx = 3 + idx_td
                        dest_ws.cell(row=r_idx, column=col_idx).value = td['product_code']
                        dest_ws.cell(row=r_idx, column=col_idx+2).value = td['tons']
                        
                        tph = get_tph(td['product_code'], m, c_dict, f_dict)
                        run_hrs = td['tons'] / tph if tph > 0 else 1.0
                        dest_ws.cell(row=r_idx, column=col_idx+3).value = run_hrs

                # Ghi Kế Hoạch 3 Ca
                pellet_plan = pellet_plan_result.get('pellet_plan', {})
                for m, runs in pellet_plan.items():
                    col_idx = machine_cols.get(m)
                    if not col_idx:
                        continue
                    
                    non_carryover_runs = [r for r in runs if not r.get('is_carryover', False)]
                    ca1_runs = [r for r in non_carryover_runs if r['shift'] == 'CA 1']
                    ca2_runs = [r for r in non_carryover_runs if r['shift'] == 'CA 2']
                    ca3_runs = [r for r in non_carryover_runs if r['shift'] == 'CA 3']
                    
                    # Ghi Ca 1
                    for idx_run, r_data in enumerate(ca1_runs[:4]):
                        r_idx = 5 + idx_run
                        dest_ws.cell(row=r_idx, column=col_idx).value = r_data['product_code']
                        dest_ws.cell(row=r_idx, column=col_idx+1).value = r_data['batches']
                        dest_ws.cell(row=r_idx, column=col_idx+2).value = r_data['tons']
                        dest_ws.cell(row=r_idx, column=col_idx+3).value = r_data['total_hours']
                        
                    # Ghi Ca 2
                    for idx_run, r_data in enumerate(ca2_runs[:4]):
                        r_idx = 9 + idx_run
                        dest_ws.cell(row=r_idx, column=col_idx).value = r_data['product_code']
                        dest_ws.cell(row=r_idx, column=col_idx+1).value = r_data['batches']
                        dest_ws.cell(row=r_idx, column=col_idx+2).value = r_data['tons']
                        dest_ws.cell(row=r_idx, column=col_idx+3).value = r_data['total_hours']
                        
                    # Ghi Ca 3
                    for idx_run, r_data in enumerate(ca3_runs[:4]):
                        r_idx = 13 + idx_run
                        dest_ws.cell(row=r_idx, column=col_idx).value = r_data['product_code']
                        dest_ws.cell(row=r_idx, column=col_idx+1).value = r_data['batches']
                        dest_ws.cell(row=r_idx, column=col_idx+2).value = r_data['tons']
                        dest_ws.cell(row=r_idx, column=col_idx+3).value = r_data['total_hours']

                # Ghi tổng hao hụt
                loss_by_machine = pellet_plan_result.get('loss_by_machine', {})
                for m, l_val in loss_by_machine.items():
                    col_idx = machine_cols.get(m)
                    if col_idx:
                        dest_ws.cell(row=18, column=col_idx+3).value = l_val

                # Ghi NEXT gối đầu
                for m, runs in pellet_plan.items():
                    col_idx = machine_cols.get(m)
                    if col_idx and runs:
                        last_run = runs[-1]
                        dest_ws.cell(row=19, column=col_idx).value = last_run['product_code']
                        dest_ws.cell(row=19, column=col_idx+2).value = last_run['tons']

                # Ghi Cám Bột Mash Feed
                mash_plan = pellet_plan_result.get('mash_plan', [])
                for idx_m, m_item in enumerate(mash_plan):
                    r_idx = 3 + idx_m
                    if r_idx > 45:
                        break
                    
                    pkg_str = ""
                    if m_item.higro_25: pkg_str = "HG 25"
                    elif m_item.cp_25: pkg_str = "CP 25"
                    elif m_item.star_25: pkg_str = "STAR 25"
                    elif m_item.bell_25: pkg_str = "BELL 25"
                    else: pkg_str = m_item.packing_size or "25"
                    
                    dest_ws.cell(row=r_idx, column=30).value = pkg_str
                    dest_ws.cell(row=r_idx, column=31).value = f"{int(m_item.tons * 1000 / 25)} BAO" if pkg_str != "SILO" else "SILO"
                    dest_ws.cell(row=r_idx, column=32).value = m_item.product_code
                    dest_ws.cell(row=r_idx, column=33).value = m_item.batches
                    dest_ws.cell(row=r_idx, column=34).value = m_item.tons

                # Cập nhật tổng tấn và giờ thực tế
                for m, runs in pellet_plan.items():
                    col_idx = machine_cols.get(m)
                    if col_idx:
                        total_tons_act = sum(r['tons'] for r in runs if not r.get('is_carryover', False))
                        total_hours_act = sum(r['total_hours'] for r in runs if not r.get('is_carryover', False))
                        dest_ws.cell(row=24, column=col_idx).value = total_tons_act
                        dest_ws.cell(row=23, column=col_idx).value = total_hours_act
                
                # Mixer summary
                dest_ws.cell(row=3, column=35).value = f"  1. (3 Tr) {sum(r['batches'] for runs in pellet_plan.values() for r in runs if r['shift'] == 'CA 1') + sum(m.batches for m in mash_plan) // 3} mẻ"
                dest_ws.cell(row=4, column=35).value = f"  2. (4 Tr) {sum(r['batches'] for runs in pellet_plan.values() for r in runs if r['shift'] == 'CA 2') + sum(m.batches for m in mash_plan) // 3} mẻ"
                dest_ws.cell(row=5, column=35).value = f"  3. (9 Tr) {sum(r['batches'] for runs in pellet_plan.values() for r in runs if r['shift'] == 'CA 3') + sum(m.batches for m in mash_plan) // 3} mẻ"

                src_wb.close()
                print("  ✅ Đã điền đầy đủ dữ liệu Kế hoạch PL và cám bột thành công!")
            else:
                print("  ⚠️ Không tìm thấy sheet 'KẾ HOẠCH PL' trong file mẫu.")

    # 8. Xóa toàn bộ các sheet không liên quan để file tinh gọn
    # Chỉ giữ lại sheet kế hoạch ngày mới, FEEDCODE, KHÁNG SINH, LINE PK VÀ CV và KẾ HOẠCH PL (nếu có)
    sheets_to_keep = {result.date, 'FEEDCODE', 'KHÁNG SINH', 'LINE PK VÀ CV', 'KẾ HOẠCH PL'}
    for name in list(wb.sheetnames):
        if name not in sheets_to_keep:
            wb.remove(wb[name])
            
    # Sửa lỗi openpyxl copy/delete sheet làm hỏng danh sách fill toàn cục
    from openpyxl.styles.fills import Fill
    cleaned_fills = []
    for f in list(wb._fills):
        if isinstance(f, Fill):
            cleaned_fills.append(f)
    wb._fills = cleaned_fills
            
    # 9. Lưu workbook mới
    os.makedirs(output_dir, exist_ok=True)
    try:
        wb.save(filepath)
        wb.close()
    except PermissionError:
        # File đang mở ở Excel, lưu bản sao mới
        import time
        suffix = int(time.time()) % 1000
        base, ext = os.path.splitext(filepath)
        filepath = f"{base}_v{suffix}{ext}"
        print(f"\n⚠️ CẢNH BÁO: File gốc đang mở trong Excel hoặc bị khóa! Tự động lưu bản sao...")
        wb.save(filepath)
        wb.close()
            
    print(f"\n  💾 Đã xuất KHSX chất lượng cao: {filepath}")

    # 10. Tự động đồng bộ kết quả lên Neon Tech PostgreSQL Cloud (nếu được bật)
    try:
        if getattr(config, 'USE_POSTGRESQL', False):
            import db_manager
            print("\n☁️ Tự động đồng bộ kết quả kế hoạch sản xuất lên Neon Tech Cloud...")
            
            # Đọc lại khpl_raw_grid từ sheet 'KẾ HOẠCH PL' vừa lưu
            wb_read = openpyxl.load_workbook(filepath, data_only=True)
            khpl_raw_grid = []
            if 'KẾ HOẠCH PL' in wb_read.sheetnames:
                pl_sheet = wb_read['KẾ HOẠCH PL']
                for r in range(1, 27):
                    row_cells = []
                    for c in range(1, 36):
                        val = pl_sheet.cell(row=r, column=c).value
                        row_cells.append(val if val is not None else '')
                    khpl_raw_grid.append(row_cells)
            wb_read.close()
            
            # Chuẩn bị sequence
            sequence_data = []
            for it in result.items:
                sequence_data.append({
                    'product_code': it.product_code,
                    'batches': it.batches,
                    'tons': it.tons,
                    'packing_size': it.packing_size,
                    'silo_tons': it.silo_truck,
                    'bag_25_tons': it.higro_25 + it.cp_25 + it.star_25 + it.nuvo_25 + it.nasa_25 + it.bell_25,
                    'bag_40_tons': it.higro_40 + it.cp_40 + it.star_40 + it.nuvo_40 + it.nasa_40 + it.bell_40,
                    'bag_50_tons': it.white_bag_50,
                    'line_cv': it.line_cv,
                    'line_pk': it.line_pk,
                    'ks_code': it.ks_code,
                    'ks_level': it.ks_level,
                    'higro_25': it.higro_25, 'higro_40': it.higro_40,
                    'cp_25': it.cp_25, 'cp_40': it.cp_40,
                    'star_25': it.star_25, 'star_40': it.star_40,
                    'nuvo_25': it.nuvo_25, 'nuvo_40': it.nuvo_40,
                    'bell_25': it.bell_25, 'bell_40': it.bell_40,
                    'nasa_25': it.nasa_25, 'nasa_40': it.nasa_40,
                    'white_25': it.white_bag_25, 'white_40': it.white_bag_40, 'white_50': it.white_bag_50,
                    'silo_truck': it.silo_truck
                })
                
            # Chuẩn bị packaging
            packaging_data = []
            for it in result.items:
                total_pack = (it.higro_25 + it.cp_25 + it.star_25 + it.nuvo_25 + it.nasa_25 + it.bell_25 +
                              it.higro_40 + it.cp_40 + it.star_40 + it.nuvo_40 + it.nasa_40 + it.bell_40 +
                              it.white_bag_25 + it.white_bag_40 + it.white_bag_50 + it.silo_truck)
                if total_pack > 0:
                    packaging_data.append({
                        'product_code': it.product_code,
                        'tons': it.tons,
                        'packing_size': it.packing_size,
                        'line_pk': it.line_pk,
                        'higro_25': it.higro_25, 'cp_25': it.cp_25, 'star_25': it.star_25, 'nuvo_25': it.nuvo_25, 'nasa_25': it.nasa_25, 'bell_25': it.bell_25,
                        'higro_40': it.higro_40, 'cp_40': it.cp_40, 'star_40': it.star_40, 'nuvo_40': it.nuvo_40, 'nasa_40': it.nasa_40, 'bell_40': it.bell_40,
                        'white_25': it.white_bag_25, 'white_40': it.white_bag_40, 'white_50': it.white_bag_50,
                        'silo_truck': it.silo_truck
                    })
                    
            summary_data = {
                'total_batches': result.total_batches,
                'total_tons': result.total_tons,
                'product_count': result.product_count,
                'warnings': result.warnings
            }
            
            db_manager.save_plan_output_to_db(
                date_str=result.date,
                filename=os.path.basename(filepath),
                khpl_raw_grid=khpl_raw_grid,
                sequence=sequence_data,
                packaging=packaging_data,
                summary=summary_data,
                conn_str=getattr(config, 'DB_URI', db_manager.DB_URI)
            )
    except Exception as db_err:
        import traceback
        traceback.print_exc()
        print(f"⚠️ Lỗi tự động đồng bộ kết quả lên Cloud: {db_err}")

    return filepath


def print_khsx_console(result):
    """In KHSX ra console dạng bảng để kiểm tra nhanh"""
    print(f"\n{'═'*90}")
    print(f"  KẾ HOẠCH SẢN XUẤT - NGÀY {result.date}")
    print(f"{'═'*90}")
    
    print(f"{'STT':>4} {'TÊN CÁM':<12} {'MẺ':>4} {'TẤN':>8} {'BAO BÌ':<30} {'LINE':>6} {'NGUỒN':<10}")
    print(f"{'─'*90}")
    
    for i, item in enumerate(result.items, 1):
        pkg_parts = []
        if item.silo_truck: pkg_parts.append(f"SILO:{item.silo_truck:.1f}")
        if item.white_bag_50: pkg_parts.append(f"WH50:{item.white_bag_50:.1f}")
        if item.higro_25: pkg_parts.append(f"HG:{item.higro_25:.1f}")
        if item.cp_25: pkg_parts.append(f"CP:{item.cp_25:.1f}")
        if item.star_25: pkg_parts.append(f"ST:{item.star_25:.1f}")
        if item.white_bag_25: pkg_parts.append(f"WH25:{item.white_bag_25:.1f}")
        pkg_str = ' '.join(pkg_parts) if pkg_parts else '-'
        
        priority_icon = {1: '🔴', 2: '🟠', 3: '🟡', 4: '🟢'}.get(item.priority.value, '⚪')
        
        print(f"{i:>4} {item.product_code:<12} {item.batches:>4} {item.tons:>8.1f} "
              f"{pkg_str:<30} {item.line_cv:>6} {priority_icon} {item.source:<10}")
    
    print(f"{'─'*90}")
    print(f"     {'TỔNG':<12} {result.total_batches:>4} {result.total_tons:>8.1f}")
    print(f"{'═'*90}")
    
    if result.warnings:
        print(f"\n⚠️ CẢNH BÁO ({len(result.warnings)}):")
        for w in result.warnings:
            print(f"  {w}")
