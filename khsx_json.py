"""
khsx_json.py - JSON output helper for Flask Web App
"""
import sys
import os
import io

# Fix encoding for Windows console immediately before imports
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import json
import math
from datetime import datetime

# Add current dir to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import config
from models import Priority, DemandItem, ForecastItem
from data_loader import load_all_data, resolve_antibiotic_for_product
from demand_calculator import calculate_daily_demand
from constraint_solver import solve_constraints
from packaging_allocator import allocate_packaging

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "message": "Thiếu ngày kế hoạch YYYY-MM-DD"}))
        return
        
    date_str = sys.argv[1] # YYYY-MM-DD
    try:
        today = datetime.strptime(date_str, '%Y-%m-%d')
    except Exception as e:
        print(json.dumps({"success": False, "message": f"Ngày không hợp lệ: {e}"}))
        return
        
    # Determine day_of_week (1 = Mon, 6 = Sat)
    dow = today.weekday() # 0 = Mon, 6 = Sun
    day_of_week = dow + 1 # 1 = Mon, 7 = Sun
    if day_of_week > 6:
        day_of_week = 6  # Sunday -> map to Saturday
        
    # Load all data from Excel files
    try:
        data = load_all_data(config, target_date=today)
    except Exception as e:
        print(json.dumps({"success": False, "message": f"Lỗi đọc dữ liệu Excel: {e}"}))
        return
        
    # Calculate produced_this_week from previous day KHSX output sheets
    produced_this_week = {}
    import openpyxl
    from datetime import timedelta
    weekday = today.weekday()
    if weekday > 5:
        weekday = 5
    monday_date = today - timedelta(days=weekday)
    
    def _get_tons_from_batches(product_code, batches_val):
        if batches_val is None:
            return 0.0
        try:
            b = float(batches_val)
        except:
            return 0.0
        prod_upper = str(product_code).strip().upper()
        if prod_upper == '325F' or prod_upper.startswith('550') or prod_upper.startswith('551'):
            return b * 8.0
        else:
            return b * 8.4

    for d_idx in range(weekday):
        prev_date = monday_date + timedelta(days=d_idx)
        prev_date_str = prev_date.strftime('%d-%m-%Y')
        import glob
        pattern = os.path.join(config.OUTPUT_DIR, f"KHSX_{prev_date_str}*.xlsx")
        files = glob.glob(pattern)
        if files:
            files.sort(key=os.path.getmtime, reverse=True)
            prev_file = files[0]
            try:
                prev_wb = openpyxl.load_workbook(prev_file, data_only=True, read_only=True)
                prev_ws = prev_wb.active
                for r in range(7, 42):
                    prod = prev_ws.cell(row=r, column=2).value
                    batches_val = prev_ws.cell(row=r, column=3).value
                    if prod:
                        prod_code = str(prod).strip().upper().replace(' ', '')
                        if prod_code:
                            t_val = _get_tons_from_batches(prod_code, batches_val)
                            produced_this_week[prod_code] = produced_this_week.get(prod_code, 0.0) + t_val
                prev_wb.close()
            except:
                pass
                
    # Calculate demand using original algorithm
    demand_list = calculate_daily_demand(
        today_date=today,
        day_of_week=day_of_week,
        forecast=data['forecast'],
        silo_plan=data['silo_plan'],
        bacang=data['bacang'],
        walkin_orders=[],
        ffstock=data['ffstock'],
        tonbon=data['tonbon'],
        khsx_yesterday=data['khsx_yesterday'],
        congsuat=data['congsuat'],
        produced_this_week=produced_this_week,
        ffstock_details=data.get('ffstock_details', {}),
        adjustments=data.get('adjustments'),
    )
    
    # Solve constraints
    demand_list, warnings = solve_constraints(
        demand_list=demand_list,
        empty_bag=data['empty_bag'],
        congsuat=data['congsuat'],
        min_tons=config.MIN_DAILY_TONS,
        max_tons=config.MAX_DAILY_TONS,
        target_tons=config.TARGET_DAILY_TONS,
        ffstock_details=data.get('ffstock_details', {}),
    )
    
    # Allocate packaging
    allocate_packaging(demand_list, data['forecast'], adjustments=data.get('adjustments'))
    
    # Pellet assignment and Bio-security sorting
    feedcode = data['feedcode']
    khangsinh = data['khangsinh']
    fix_code_pellet = data.get('fix_code_pellet', {})
    congsuat = data.get('congsuat', {})
    
    def find_pellet_config(product_code, packing_size, die_size, is_mash, matrix):
        p_code = str(product_code).strip().upper()
        pack = str(packing_size).strip().upper() if packing_size else '25'
        if is_mash:
            key = f"{p_code}{pack}M"
            if key in matrix:
                return matrix[key]
        if die_size:
            die_str = str(die_size)
            if die_str.endswith('.0'):
                die_str_alt = die_str[:-2]
            else:
                die_str_alt = die_str + ".0"
            keys_to_try = [f"{p_code}{pack}{die_str}", f"{p_code}{pack}{die_str_alt}"]
            if pack == 'SILO':
                keys_to_try.append(f"{p_code}SILOSILO")
            for k in keys_to_try:
                if k in matrix:
                    return matrix[k]
        prefix = f"{p_code}{pack}"
        matches = []
        for k, cfg in matrix.items():
            if k.startswith(prefix):
                matches.append((k, cfg))
        if matches:
            return matches[0][1]
        matches_broad = []
        for k, cfg in matrix.items():
            if k.startswith(p_code):
                matches_broad.append((k, cfg))
        if matches_broad:
            return matches_broad[0][1]
        return None

    LINE_ORDER = {
        'PL1': 1, '1': 1, 'PL2': 2, '2': 2, 'PL3': 3, '3': 3,
        'PL4': 4, '4': 4, 'PL5': 5, '5': 5, 'PL6': 6, '6': 6,
        'PL7': 7, '7': 7, 'M': 8, 'PLM': 8, '': 99
    }
    
    for item in demand_list:
        spec = congsuat.get(item.product_code, None)
        die_size = spec.die_size if spec else 0.0
        line_cv_default = spec.line_cv if spec else ''
        is_mash = (line_cv_default == 'M')
        cfg = find_pellet_config(item.product_code, item.packing_size, die_size, is_mash, fix_code_pellet)
        assigned_line = ''
        if cfg and cfg['priorities']:
            for p_line in cfg['priorities']:
                norm_line = str(p_line).strip().upper()
                if norm_line.isdigit():
                    norm_line = f"PL{norm_line}"
                is_excluded = False
                if item.product_code.startswith('566') or item.product_code.startswith('567S'):
                    if norm_line in {'PL2', '2'}:
                        is_excluded = True
                if not is_excluded:
                    assigned_line = norm_line
                    break
        if not assigned_line and cfg and cfg['default_line']:
            default_line = str(cfg['default_line']).strip().upper()
            if default_line.isdigit():
                default_line = f"PL{default_line}"
            is_excluded = False
            if item.product_code.startswith('566') or item.product_code.startswith('567S'):
                if default_line in {'PL2', '2'}:
                    is_excluded = True
            if not is_excluded:
                assigned_line = default_line
        if not assigned_line:
            fc_data = feedcode.get(item.product_code, {})
            assigned_line = str(fc_data.get('line_cv', '')).strip().upper()
            if assigned_line.isdigit():
                assigned_line = f"PL{assigned_line}"
        if item.product_code.startswith('566') or item.product_code.startswith('567S'):
            if assigned_line in {'PL2', '2', ''}:
                assigned_line = 'PL3'
        if assigned_line.isdigit():
            assigned_line = f"PL{assigned_line}"
        item.line_cv = assigned_line
        fc_data = feedcode.get(item.product_code, {})
        item.line_pk = fc_data.get('line_pk', '')
        silo_tons = item.silo_truck or 0
        bag_tons = item.tons - silo_tons
        if bag_tons <= 0.01 and silo_tons > 0:
            item.line_pk = 'SILO'
        ks_code, ks_level = resolve_antibiotic_for_product(item.product_code, khangsinh)
        item.ks_code = ks_code
        item.ks_level = ks_level
        
    demand_list.sort(key=lambda x: (
        LINE_ORDER.get(str(x.line_cv).strip().upper(), 99),
        x.ks_level,
        -x.tons
    ))
    
    # Map back items into a simple list of dicts for frontend JSON consumption
    output_list = []
    for item in demand_list:
        stock_tons = data['ffstock'].get(item.product_code, 0.0)
        doh = 999.0
        if item.product_code in data.get('ffstock_details', {}):
            doh = data['ffstock_details'][item.product_code].get('doh', 999.0)
            
        # Format a comprehensive note
        gc = f"{item.batches} mẻ (Máy: {item.line_cv or 'PL?'}) - Line PK: {item.line_pk or ''}"
        if item.ks_code:
            gc += f" - Kháng sinh: {item.ks_code} (lv {item.ks_level})"
        if item.silo_truck > 0:
            gc += f" - Bồn Silo: {item.silo_truck:.1f}T"
            
        output_list.append({
            "code": item.product_code,
            "so_luong": float(item.tons * 1000.0), # convert to Kg
            "stock": float(stock_tons * 1000.0), # convert to Kg
            "doh": float(round(doh, 1)) if doh != 999.0 else 999,
            "loai": item.source,
            "ghi_chu": gc
        })
        
    # We prefix a clean delimiter so we can separate stdout printed messages from actual JSON
    print("---JSON_START---")
    print(json.dumps({
        "success": True,
        "danh_sach": output_list,
        "warnings": warnings
    }, ensure_ascii=False))

if __name__ == '__main__':
    main()
