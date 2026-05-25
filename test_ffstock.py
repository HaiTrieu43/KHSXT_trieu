import os
from openpyxl import load_workbook

def test_ffstock_cols():
    file_path = r"D:\Kê hoạch sản xuât\FSTOCK-BAG\FFSTOCK 17 -05-2026.xlsm"
    wb = load_workbook(file_path, data_only=True, read_only=True)
    target_sheets = ['PRO-NASA', 'BRAN', 'INTGRATE']
    
    for sname in wb.sheetnames:
        sname_upper = sname.upper()
        is_target = any(t in sname_upper for t in ['PRO', 'NASA', 'BRAN', 'INTGRATE', 'INTEGRATE'])
        if not is_target:
            continue
            
        print(f"\n--- Sheet: {sname} ---")
        ws = wb[sname]
        
        # Đọc 15 dòng đầu, 30 cột đầu để phân tích tiêu đề
        headers = []
        for r in range(1, 16):
            row_vals = []
            for c in range(1, 31):
                val = ws.cell(row=r, column=c).value
                row_vals.append((c-1, val))
            headers.append((r, row_vals))
            
        # Tìm cột tiêu đề động
        feed_code_col = None
        balance_kg_col = None
        balance_bag_col = None
        sales_avg_col = None
        doh_col = None
        
        # Duyệt qua các cột
        for col_idx in range(30):
            # Lấy tất cả giá trị của cột này trong 15 dòng đầu
            col_vals = []
            for r_idx in range(15):
                val = headers[r_idx][1][col_idx][1]
                if val is not None:
                    col_vals.append((r_idx + 1, str(val).strip().upper()))
            
            # Kiểm tra xem cột này có phải là cột cần tìm không
            col_str = " | ".join([f"R{r}:{v}" for r, v in col_vals])
            
            # 1. Mã cám
            if any("FEED CODE" in v or "MÃ CÁM" in v or "TEN FEED" in v for r, v in col_vals):
                feed_code_col = col_idx
            
            # 2. Tồn kho Kg
            # Tìm ô chứa BALANCE hoặc TỒN CUỐI
            is_balance = any("BALANCE" in v or "TỒN CUỐI" in v or "TON CUOI" in v for r, v in col_vals)
            if is_balance:
                # Phân biệt bao và kg:
                # Nếu có chữ "KG" hoặc "KÝ" trong cột này
                has_kg = any("KG" in v or "KÝ" in v or "KILO" in v for r, v in col_vals)
                if has_kg:
                    balance_kg_col = col_idx
                else:
                    # Nếu không có chữ KG, xem cột bên cạnh có chữ KG không, hoặc tạm ghi nhận làm bag
                    balance_bag_col = col_idx
            
            # 3. TB bán
            if any("SALES AVERAGE" in v or "TB BÁN" in v or "BÁN TB" in v or "TB BAN" in v or "AVERAGE" in v for r, v in col_vals):
                sales_avg_col = col_idx
                
            # 4. DOH
            if any("DAY ON HAND" in v or "DOH" in v or "SỐ NGÀY TỒN" in v or "SO NGAY TON" in v or "NGÀY TỒN" in v or "ON HAND" in v for r, v in col_vals):
                doh_col = col_idx

        # Nếu balance_kg_col chưa tìm thấy mà có balance_bag_col, hãy kiểm tra cột ngay sau nó
        if balance_kg_col is None and balance_bag_col is not None:
            # Thử xem cột ngay sau nó có phải là Kg không (hoặc mặc định là cột ngay sau)
            next_col = balance_bag_col + 1
            balance_kg_col = next_col
            print(f"  (Fallback balance_kg_col to next col: {balance_kg_col})")
            
        print(f"  FEED CODE col: {feed_code_col}")
        print(f"  BALANCE KG col: {balance_kg_col} (Bag col: {balance_bag_col})")
        print(f"  SALES AVG col: {sales_avg_col}")
        print(f"  DOH col: {doh_col}")
        
        # Đọc thử 5 dòng dữ liệu đầu tiên
        data_start_row = 1
        # Tìm dòng bắt đầu dữ liệu thực tế (dòng có mã cám ở cột feed_code_col và không phải tiêu đề)
        # Thông thường dòng dữ liệu bắt đầu sau các dòng tiêu đề (sau dòng 6-10)
        # Chúng ta quét từ dòng 4 đến 25, nếu thấy dòng nào có giá trị ở feed_code_col mà không chứa các từ khóa tiêu đề
        for r in range(4, 25):
            val = ws.cell(row=r, column=feed_code_col+1).value if feed_code_col is not None else None
            if val is not None:
                val_str = str(val).strip().upper()
                if val_str and not any(kw in val_str for kw in ['TOTAL', 'STT', 'FEED', 'GRAND', 'HI-GRO', 'BRAND', 'FARM', 'CODE', 'TÊN']):
                    data_start_row = r
                    break
        
        print(f"  Data start row: {data_start_row}")
        
        # Đọc thử 5 dòng
        count = 0
        for r in range(data_start_row, data_start_row + 15):
            f_val = ws.cell(row=r, column=feed_code_col+1).value if feed_code_col is not None else None
            b_val = ws.cell(row=r, column=balance_kg_col+1).value if balance_kg_col is not None else None
            s_val = ws.cell(row=r, column=sales_avg_col+1).value if sales_avg_col is not None else None
            d_val = ws.cell(row=r, column=doh_col+1).value if doh_col is not None else None
            
            if f_val is not None:
                print(f"    Row {r}: Code={f_val} | Stock={b_val} | SalesAvg={s_val} | DOH={d_val}")
                count += 1
                if count >= 5:
                    break

    wb.close()

if __name__ == '__main__':
    test_ffstock_cols()
