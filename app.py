"""
app.py - Backend Flask server for KHSX Automator
C.P. Vietnam - Chi nhánh Bình Dương
"""

import sys
import os
import io
import datetime
import subprocess
import tempfile
import json
import traceback
from flask import Flask, render_template, jsonify, request, send_from_directory, Response, session, redirect, url_for

# Fix encoding cho Windows console
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Thêm thư mục hiện tại vào path
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, CURRENT_DIR)

import config
import data_loader
from models import Priority, PackingType

app = Flask(__name__, template_folder=os.path.join(CURRENT_DIR, 'templates'), static_folder=os.path.join(CURRENT_DIR, 'static'))
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY', 'cp_vietnam_khsx_secret_key')

# Khởi tạo thư mục upload tạm thời
TEMP_UPLOAD_DIR = os.path.join(CURRENT_DIR, 'temp_uploads')
os.makedirs(TEMP_UPLOAD_DIR, exist_ok=True)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def get_file_info(directory, pattern, exact_path=None):
    """Lấy thông tin của file mới nhất theo pattern hoặc đường dẫn chính xác"""
    if exact_path:
        path = exact_path
    else:
        path = data_loader._find_latest_file(directory, pattern)
        
    if not path or not os.path.isfile(path):
        return {
            'exists': False,
            'filename': 'Không tìm thấy',
            'last_modified': '-',
            'size': '-',
            'path': ''
        }
        
    stat = os.stat(path)
    mtime = datetime.datetime.fromtimestamp(stat.st_mtime)
    size_kb = round(stat.st_size / 1024, 1)
    
    return {
        'exists': True,
        'filename': os.path.basename(path),
        'last_modified': mtime.strftime('%d-%m-%Y %H:%M:%S'),
        'size': f"{size_kb} KB",
        'path': path
    }


def _safe_float_val(v):
    if v is None or v == '':
        return ''
    try:
        return float(v)
    except:
        return v


def _safe_float_num(v, default=0.0):
    if v is None or v == '':
        return default
    try:
        return float(v)
    except:
        return default


def _safe_int_val(v):
    if v is None or v == '':
        return ''
    try:
        return int(float(v))
    except:
        return v


# ============================================================
# XÁC THỰC NGƯỜI DÙNG (USER AUTHENTICATION)
# ============================================================

@app.before_request
def require_login():
    """Tự động bảo vệ tất cả các endpoints, yêu cầu đăng nhập"""
    # Các endpoint được truy cập tự do không cần đăng nhập
    allowed_endpoints = ['login', 'static', 'healthz', 'health']
    if request.endpoint and request.endpoint not in allowed_endpoints:
        # Bỏ qua kiểm tra favicon
        if request.path == '/favicon.ico':
            return
        if not session.get('logged_in'):
            return redirect(url_for('login'))


@app.route('/healthz', methods=['GET'])
@app.route('/health', methods=['GET'])
def healthz():
    """Endpoint Health Check phục vụ Uptime Robot / Render Keep-Alive"""
    status = {
        "status": "healthy",
        "timestamp": datetime.datetime.now().isoformat(),
        "database": "not_used"
    }
    
    if getattr(config, 'USE_POSTGRESQL', False):
        try:
            import db_manager
            # Thử tạo kết nối nhanh và chạy truy vấn SELECT 1
            conn = db_manager.get_connection(getattr(config, 'DB_URI', db_manager.DB_URI))
            cur = conn.cursor()
            cur.execute("SELECT 1;")
            cur.fetchone()
            cur.close()
            conn.close()
            status["database"] = "connected"
        except Exception as e:
            status["status"] = "degraded"
            status["database"] = "error"
            status["error"] = str(e)
            return jsonify(status), 500
    
    return jsonify(status), 200
            
@app.route('/favicon.ico')
def favicon():
    """Trả về phản hồi trống cho favicon để tránh cảnh báo lỗi 404 trên console trình duyệt"""
    return '', 204


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Giao diện đăng nhập Admin"""
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        if username == getattr(config, 'ADMIN_USERNAME', 'admin') and password == getattr(config, 'ADMIN_PASSWORD', 'cp@123456'):
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('index'))
        else:
            error = 'Tên đăng nhập hoặc mật khẩu không chính xác!'
            
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    """Đăng xuất người dùng, xóa session"""
    session.clear()
    return redirect(url_for('login'))


# ============================================================
# VIEWS & TEMPLATE ROUTES
# ============================================================

@app.route('/')
def index():
    """Trang chủ Single Page App"""
    return render_template('index.html')


# ============================================================
# API: TRẠNG THÁI NGUỒN DỮ LIỆU
# ============================================================

@app.route('/api/data-status', methods=['GET'])
def get_data_status():
    """Lấy thông tin và trạng thái của 10 nguồn dữ liệu đầu vào"""
    try:
        status = {
            'forecast': get_file_info(config.FORECAST_DIR, '*FORECAST*.xlsx'),
            'silo_plan': get_file_info(config.SILO_DIR, '*SILO*.xlsx'),
            'bacang': get_file_info(config.BACANG_DIR, '*CANG*.xlsx'),
            'ffstock': get_file_info(getattr(config, 'FSTOCK_DIR_FFSTOCK', config.FSTOCK_DIR), '*FFSTOCK*.xls*'),
            'tonbon': get_file_info(config.TONBON_DIR, '*ton bon*.*'),
            'empty_bag': get_file_info(getattr(config, 'FSTOCK_DIR_EMPTYBAG', config.FSTOCK_DIR), '*EMPTY BAG*.xls*'),
            'congsuat': get_file_info(None, None, exact_path=config.PLAN_FILE),
            'feedcode': get_file_info(None, None, exact_path=config.KHSX_FILE),
            'khangsinh': get_file_info(None, None, exact_path=config.KHSX_FILE),
            'yesterday_plan': get_file_info(None, None, exact_path=config.KHSX_FILE),
            'adjustments': get_file_info(None, None, exact_path=config.QUICK_ADJUST_FILE)
        }
        return jsonify({'success': True, 'status': status})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/data-freshness', methods=['GET'])
def get_data_freshness():
    """Kiểm tra độ tươi mới của dữ liệu: so sánh ngày báo cáo giữa các nguồn.
    
    Logic:
    - FFStock là báo cáo tham chiếu (anchor) — ngày FFStock mới nhất = ngày chuẩn
    - Các báo cáo hàng ngày (Empty Bag, Tồn Bồn) phải cùng ngày hoặc không quá 1 ngày
    - Các báo cáo hàng tuần (Forecast, SILO, Ba Cảng) phải cùng tuần
    
    Hỗ trợ chế độ chạy Cloud (đọc metadata đồng bộ từ DB) lẫn chế độ Local (quét file).
    """
    import re
    import glob
    
    warnings = []
    source_dates = {}
    
    # 0. Xác định chế độ chạy (File hay Database)
    db_mode = getattr(config, 'USE_POSTGRESQL', False) or os.environ.get('IS_CLOUD') == 'true'
    
    filenames = {}
    bacang_mtime = None
    
    if db_mode:
        try:
            import db_manager
            conn = db_manager.get_connection(getattr(config, 'DB_URI', db_manager.DB_URI))
            cur = conn.cursor()
            # Khởi tạo bảng sync_metadata nếu chưa tồn tại
            cur.execute("""
            CREATE TABLE IF NOT EXISTS sync_metadata (
                category VARCHAR(50) PRIMARY KEY,
                filename VARCHAR(255),
                last_modified VARCHAR(50),
                synced_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """)
            cur.execute("SELECT category, filename, last_modified FROM sync_metadata;")
            rows = cur.fetchall()
            for row in rows:
                cat, fname, mtime_str = row
                filenames[cat] = fname
                if cat == 'bacang' and mtime_str:
                    try:
                        # Parse mtime_str dạng '%Y-%m-%d %H:%M:%S'
                        dt = datetime.datetime.strptime(mtime_str, '%Y-%m-%d %H:%M:%S').date()
                        bacang_mtime = dt
                    except:
                        pass
            cur.close()
            conn.close()
        except Exception as db_err:
            print(f"⚠️ Lỗi đọc sync_metadata từ DB: {db_err}")
            
    # Các hàm trợ giúp trích xuất ngày và tuần từ tên file
    def _extract_date_from_filename(filename, date_regex):
        if not filename: return None
        # Kiểm tra nếu tên file có đính kèm metadata ngày đặc biệt dạng [DATE: dd-mm-yyyy]
        date_meta = re.search(r'\[DATE:\s*(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})\]', filename)
        if date_meta:
            try:
                d, m, y = int(date_meta.group(1)), int(date_meta.group(2)), int(date_meta.group(3))
                if y < 100: y += 2000
                return datetime.date(y, m, d)
            except:
                pass
                
        match = re.search(date_regex, filename)
        if match:
            groups = match.groups()
            try:
                if len(groups) == 3:
                    d, m, y = int(groups[0]), int(groups[1]), int(groups[2])
                    if y < 100: y += 2000
                    return datetime.date(y, m, d)
            except (ValueError, OverflowError):
                pass
        return None

    def _extract_week_from_filename(filename):
        if not filename: return None, None
        wk_match = re.search(r'W(\d+)', filename, re.IGNORECASE)
        best_week = int(wk_match.group(1)) if wk_match else None
        best_range = None
        range_match = re.search(r'(\d{1,2})\s*[-/]\s*(\d{1,2})\s*[-/]\s*(\d{1,2})', filename)
        if range_match:
            best_range = f"{range_match.group(1)}-{range_match.group(2)}/{range_match.group(3)}"
        return best_week, best_range

    def _extract_date_from_files(directory, pattern, date_regex, source_name):
        """Trích xuất ngày mới nhất từ tên file cục bộ."""
        if not directory or not os.path.isdir(directory):
            return None
        
        search_pattern = os.path.join(directory, '**', pattern)
        files = glob.glob(search_pattern, recursive=True)
        files = [f for f in files if not os.path.basename(f).startswith('~$')]
        
        if not files:
            return None
        
        dates = []
        for filepath in files:
            dt = _extract_date_from_filename(os.path.basename(filepath), date_regex)
            if dt:
                dates.append(dt)
        return max(dates) if dates else None

    def _extract_week_from_files(directory, pattern, source_name):
        """Trích xuất tuần mới nhất từ tên file Forecast/SILO cục bộ."""
        if not directory or not os.path.isdir(directory):
            return None, None
        
        search_pattern = os.path.join(directory, '**', pattern)
        files = glob.glob(search_pattern, recursive=True)
        files = [f for f in files if not os.path.basename(f).startswith('~$')]
        
        if not files:
            return None, None
        
        best_week = None
        best_range = None
        for filepath in files:
            w, r = _extract_week_from_filename(os.path.basename(filepath))
            if w:
                if best_week is None or w > best_week:
                    best_week = w
                    best_range = r
        return best_week, best_range

    try:
        if db_mode:
            # === CHẾ ĐỘ CLOUD DATABASE ===
            ffstock_date = _extract_date_from_filename(
                filenames.get('ffstock'),
                r'FFSTOCK\s*(\d{1,2})\s*[-/:.\s]\s*(\d{1,2})\s*[-/:.\s]\s*(\d{2,4})'
            )
            emptybag_date = _extract_date_from_filename(
                filenames.get('empty_bag'),
                r'(\d{1,2})\s*[-/:.\s]\s*(\d{1,2})\s*[-/:.\s]\s*(\d{2,4})'
            )
            tonbon_date = _extract_date_from_filename(
                filenames.get('tonbon'),
                r'(\d{1,2})\s*[-./:]\s*(\d{1,2})\s*[-./:]\s*(\d{2,4})'
            )
            forecast_week, forecast_range = _extract_week_from_filename(filenames.get('forecast'))
            silo_week, silo_range = _extract_week_from_filename(filenames.get('silo_plan'))
        else:
            # === CHẾ ĐỘ LOCAL FILE ===
            # 1. FFStock — Báo cáo tham chiếu (anchor)
            ffstock_dir = getattr(config, 'FSTOCK_DIR_FFSTOCK', config.FSTOCK_DIR)
            ffstock_date = _extract_date_from_files(
                ffstock_dir, '*FFSTOCK*.xls*',
                r'FFSTOCK\s*(\d{1,2})\s*[-/:.\s]\s*(\d{1,2})\s*[-/:.\s]\s*(\d{2,4})',
                'FFStock'
            )
            
            # 2. Empty Bag — Báo cáo hàng ngày
            emptybag_dir = getattr(config, 'FSTOCK_DIR_EMPTYBAG', config.FSTOCK_DIR)
            emptybag_date = _extract_date_from_files(
                emptybag_dir, '*EMPTY BAG*.xls*',
                r'(\d{1,2})\s*[-/:.\s]\s*(\d{1,2})\s*[-/:.\s]\s*(\d{2,4})',
                'Empty Bag'
            )
            
            # 3. Tồn Bồn — Báo cáo hàng ngày
            import data_loader
            tonbon_file = data_loader._find_latest_file(config.TONBON_DIR, '*ton bon*.*')
            if not tonbon_file:
                fallback_dir = getattr(config, 'TONBON_DIR_FALLBACK', None)
                if fallback_dir and fallback_dir != config.TONBON_DIR:
                    tonbon_file = data_loader._find_latest_file(fallback_dir, '*ton bon*.*')
                    
            tonbon_date = None
            if tonbon_file:
                actual_tb_date = data_loader.get_tonbon_actual_date(tonbon_file)
                if actual_tb_date:
                    try:
                        d, m, y = map(int, actual_tb_date.split('-'))
                        tonbon_date = datetime.date(y, m, d)
                    except:
                        pass
            
            # 4. Forecast — Báo cáo tuần
            forecast_week, forecast_range = _extract_week_from_files(
                config.FORECAST_DIR, '*FORECAST*.xlsx', 'Forecast'
            )
            
            # 5. SILO — Báo cáo tuần
            silo_week, silo_range = _extract_week_from_files(
                config.SILO_DIR, '*SILO*.xlsx', 'SILO'
            )
            
            # 6. Ba Cảng — Báo cáo tuần
            bacang_dir = config.BACANG_DIR
            bacang_files = glob.glob(os.path.join(bacang_dir, '*CANG*.xlsx')) if os.path.isdir(bacang_dir) else []
            bacang_files = [f for f in bacang_files if not os.path.basename(f).startswith('~$')]
            if bacang_files:
                latest_bc = max(bacang_files, key=os.path.getmtime)
                bacang_mtime = datetime.date.fromtimestamp(os.path.getmtime(latest_bc))

        # Lưu thông tin ngày hiển thị
        if ffstock_date:
            source_dates['ffstock'] = {'date': ffstock_date, 'label': 'FFStock Tồn Kho'}
        if emptybag_date:
            source_dates['empty_bag'] = {'date': emptybag_date, 'label': 'Bao Bì (Empty Bag)'}
        if tonbon_date:
            source_dates['tonbon'] = {'date': tonbon_date, 'label': 'Tồn Bồn Thành Phẩm'}
            
        # === SO SÁNH VÀ TẠO CẢNH BÁO ===
        today = datetime.date.today()
        anchor_date = ffstock_date  # FFStock là ngày chuẩn
        
        if not anchor_date:
            warnings.append({
                'level': 'critical',
                'source': 'FFStock',
                'message': 'Không tìm thấy dữ liệu FFStock — không thể xác định ngày tham chiếu!',
                'icon': '🚫'
            })
        else:
            # Cảnh báo nếu FFStock cũng đã cũ so với hôm nay
            days_old = (today - anchor_date).days
            if days_old > 2:
                warnings.append({
                    'level': 'warning',
                    'source': 'FFStock',
                    'message': f'FFStock mới nhất ngày {anchor_date.strftime("%d/%m/%Y")} — cũ hơn {days_old} ngày so với hôm nay ({today.strftime("%d/%m/%Y")})',
                    'icon': '📦'
                })
            
            # So sánh Empty Bag vs FFStock
            if emptybag_date:
                delta = (anchor_date - emptybag_date).days
                if delta > 0:
                    warnings.append({
                        'level': 'critical' if delta > 1 else 'warning',
                        'source': 'Empty Bag',
                        'message': f'Dữ liệu Bao Bì mới nhất ngày {emptybag_date.strftime("%d/%m/%Y")} — THIẾU {delta} ngày so với FFStock ({anchor_date.strftime("%d/%m/%Y")})',
                        'icon': '📋',
                        'action': 'Kiểm tra và đồng bộ báo cáo Empty Bag mới nhất'
                    })
            else:
                warnings.append({
                    'level': 'critical',
                    'source': 'Empty Bag',
                    'message': 'Không tìm thấy dữ liệu báo cáo Bao Bì (Empty Bag)!',
                    'icon': '📋'
                })
            
            # So sánh Tồn Bồn vs FFStock  
            if tonbon_date:
                delta = (anchor_date - tonbon_date).days
                if delta > 0:
                    warnings.append({
                        'level': 'critical' if delta > 1 else 'warning',
                        'source': 'Tồn Bồn',
                        'message': f'Dữ liệu Tồn Bồn mới nhất ngày {tonbon_date.strftime("%d/%m/%Y")} — THIẾU {delta} ngày so với FFStock ({anchor_date.strftime("%d/%m/%Y")})',
                        'icon': '🏭',
                        'action': 'Kiểm tra và đồng bộ báo cáo Tồn Bồn mới nhất'
                    })
            else:
                warnings.append({
                    'level': 'warning',
                    'source': 'Tồn Bồn',
                    'message': 'Không tìm thấy dữ liệu báo cáo Tồn Bồn!',
                    'icon': '🏭'
                })
        
        # So sánh báo cáo tuần: Forecast vs SILO
        current_week = today.isocalendar()[1]
        
        if forecast_week:
            if forecast_week < current_week:
                warnings.append({
                    'level': 'warning',
                    'source': 'Forecast',
                    'message': f'Forecast đang ở Tuần {forecast_week} ({forecast_range or "?"}) — hiện tại là Tuần {current_week}',
                    'icon': '📊',
                    'action': 'Cần đồng bộ Forecast tuần mới'
                })
        else:
            warnings.append({
                'level': 'critical',
                'source': 'Forecast',
                'message': 'Không tìm thấy dữ liệu Forecast!',
                'icon': '📊'
            })
        
        if silo_week:
            if silo_week < current_week:
                warnings.append({
                    'level': 'warning',
                    'source': 'SILO',
                    'message': f'Kế hoạch SILO đang ở Tuần {silo_week} ({silo_range or "?"}) — hiện tại là Tuần {current_week}',
                    'icon': '🏗️',
                    'action': 'Cần đồng bộ kế hoạch SILO tuần mới'
                })
        else:
            warnings.append({
                'level': 'warning',
                'source': 'SILO',
                'message': 'Không tìm thấy dữ liệu kế hoạch SILO!',
                'icon': '🏗️'
            })
        
        # Ba Cảng — kiểm tra mtime
        if bacang_mtime:
            bc_days_old = (today - bacang_mtime).days
            if bc_days_old > 7:
                warnings.append({
                    'level': 'warning',
                    'source': 'Ba Cảng',
                    'message': f'Kế hoạch Ba Cảng chưa cập nhật — dữ liệu cũ {bc_days_old} ngày',
                    'icon': '⚓',
                    'action': 'Cần đồng bộ kế hoạch Ba Cảng tuần mới'
                })
        
        # Build response
        response = {
            'success': True,
            'reference_date': anchor_date.strftime('%d/%m/%Y') if anchor_date else None,
            'current_date': today.strftime('%d/%m/%Y'),
            'current_week': current_week,
            'sources': {},
            'warnings': warnings,
            'warning_count': len(warnings),
            'has_critical': any(w['level'] == 'critical' for w in warnings)
        }
        
        # Tóm tắt ngày mỗi nguồn
        for key, info in source_dates.items():
            response['sources'][key] = {
                'date': info['date'].strftime('%d/%m/%Y'),
                'label': info['label'],
                'days_from_ref': (anchor_date - info['date']).days if anchor_date else None
            }
        
        if forecast_week:
            response['sources']['forecast'] = {
                'week': forecast_week, 'range': forecast_range, 'label': 'Forecast Tuần',
                'is_current_week': forecast_week >= current_week
            }
        if silo_week:
            response['sources']['silo'] = {
                'week': silo_week, 'range': silo_range, 'label': 'SILO Tuần',
                'is_current_week': silo_week >= current_week
            }
        
        return jsonify(response)
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e), 'warnings': []})


@app.route('/api/latest-target-date', methods=['GET'])
def get_latest_target_date():
    """Tự động quét tìm ngày dữ liệu FFSTOCK mới nhất và tính toán ngày mục tiêu tiếp theo"""
    try:
        # Nếu đang sử dụng PostgreSQL, thử lấy từ lịch sử kế hoạch plan_outputs trước
        if getattr(config, 'USE_POSTGRESQL', False):
            try:
                import db_manager
                conn = db_manager.get_connection(getattr(config, 'DB_URI', db_manager.DB_URI))
                cur = conn.cursor()
                cur.execute("SELECT date_str FROM plan_outputs ORDER BY created_at DESC LIMIT 1;")
                row = cur.fetchone()
                cur.close()
                conn.close()
                if row:
                    latest_plan_date = row[0]
                    parts = latest_plan_date.split('-')
                    if len(parts) == 3:
                        target_iso = f"{parts[2]}-{parts[1]}-{parts[0]}"
                        return jsonify({
                            'success': True,
                            'data_date': latest_plan_date,
                            'target_date': latest_plan_date,
                            'target_date_iso': target_iso,
                            'has_existing_plan': True,
                            'existing_plan_filename': f"KHSX_{latest_plan_date}.xlsx"
                        })
            except Exception as db_ex:
                print(f"⚠️ Lỗi đọc ngày mục tiêu từ database: {db_ex}")

        import glob
        import re
        
        # 1. Tìm tất cả các file FFSTOCK trong thư mục FSTOCK_DIR
        ffstock_dir = getattr(config, 'FSTOCK_DIR_FFSTOCK', config.FSTOCK_DIR)
        pattern = os.path.join(ffstock_dir, "*FFSTOCK*.xls*")
        files = glob.glob(pattern)
        
        # Lọc bỏ file tạm thời
        files = [f for f in files if not os.path.basename(f).startswith('~$')]
        
        if not files:
            return jsonify({
                'success': False, 
                'message': 'Không tìm thấy bất kỳ file FFSTOCK nào trong thư mục'
            })
            
        # Regex linh hoạt hỗ trợ khoảng trắng
        date_pattern = re.compile(r'FFSTOCK\s*(\d{1,2})\s*[-/:\s]\s*(\d{1,2})\s*[-/:\s]\s*(\d{4})', re.IGNORECASE)
        
        dates = []
        for filepath in files:
            filename = os.path.basename(filepath)
            match = date_pattern.search(filename)
            if match:
                d = int(match.group(1))
                m = int(match.group(2))
                y = int(match.group(3))
                try:
                    dt = datetime.date(y, m, d)
                    dates.append((dt, filepath))
                except Exception:
                    pass
                    
        if not dates:
            return jsonify({
                'success': False,
                'message': 'Không thể trích xuất ngày từ tên các file FFSTOCK'
            })
            
        # Tìm ngày lớn nhất (ngày mới nhất)
        dates.sort(key=lambda x: x[0], reverse=True)
        latest_data_date, latest_file_path = dates[0]
        
        # Tính toán ngày mục tiêu (ngày sau đó 1 ngày)
        target_date = latest_data_date + datetime.timedelta(days=1)
        
        # Format các chuỗi ngày
        data_date_str = latest_data_date.strftime('%d-%m-%Y')
        target_date_str = target_date.strftime('%d-%m-%Y')
        target_date_iso = target_date.strftime('%Y-%m-%d') # cho input date
        
        # 2. Kiểm tra xem file KHSX cho ngày mục tiêu đã được lập chưa
        out_pattern = os.path.join(config.OUTPUT_DIR, f"KHSX_{target_date_str}*.xlsx")
        out_files = glob.glob(out_pattern)
        
        has_existing_plan = False
        existing_plan_filename = ""
        
        if out_files:
            out_files.sort(key=os.path.getmtime, reverse=True)
            has_existing_plan = True
            existing_plan_filename = os.path.basename(out_files[0])
            
        return jsonify({
            'success': True,
            'data_date': data_date_str,
            'target_date': target_date_str,
            'target_date_iso': target_date_iso,
            'has_existing_plan': has_existing_plan,
            'existing_plan_filename': existing_plan_filename
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# API: CHI TIẾT NGUỒN DỮ LIỆU (READ)
# ============================================================

@app.route('/api/data/<category>', methods=['GET'])
def get_detailed_data(category):
    """Đọc dữ liệu chi tiết của từng nguồn đầu vào để hiển thị bảng khoa học"""
    try:
        # Nếu đang sử dụng PostgreSQL, đọc trực tiếp từ Neon Tech Cloud mà không cần Excel
        if getattr(config, 'USE_POSTGRESQL', False):
            import db_manager
            db_data = db_manager.load_all_data_from_db(getattr(config, 'DB_URI', db_manager.DB_URI))
            info = {
                'exists': True,
                'filename': 'Neon Tech Cloud Database (PostgreSQL)',
                'last_modified': datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S'),
                'size_kb': 0.0
            }
            rows = []
            
            if category == 'forecast':
                items = db_data['forecast']
                for it in items:
                    rows.append({
                        'product_code': it.product_code,
                        'packing_size': it.packing_size,
                        'die_size': it.die_size,
                        'dealer_higro': it.dealer_higro,
                        'dealer_cp': it.dealer_cp,
                        'dealer_star': it.dealer_star,
                        'dealer_nuvo': it.dealer_nuvo,
                        'dealer_nasa': it.dealer_nasa,
                        'dealer_total': it.dealer_total,
                        'farm_swine': it.farm_swine,
                        'farm_integrate': it.farm_integrate,
                        'farm_total': it.farm_total,
                        'grand_total_tons': it.grand_total_tons,
                        'silo_tons': it.silo_tons,
                        'total_with_silo': it.total_with_silo
                    })
            elif category == 'silo_plan':
                silo_plan = db_data['silo_plan']
                all_products = sorted(list(set(p for d in silo_plan.values() for p in d)))
                for p in all_products:
                    row = {'product_code': p}
                    for day in range(1, 7):
                        row[f'day_{day}'] = round(silo_plan[day].get(p, 0.0), 1)
                    row['total'] = round(sum(row[f'day_{d}'] for d in range(1, 7)), 1)
                    rows.append(row)
            elif category == 'bacang':
                bacang = db_data['bacang']
                all_products = sorted(list(set(p for d in bacang.values() for p in d)))
                for p in all_products:
                    row = {'product_code': p}
                    for day in range(1, 7):
                        row[f'day_{day}'] = round(bacang[day].get(p, 0.0), 1)
                    row['total'] = round(sum(row[f'day_{d}'] for d in range(1, 7)), 1)
                    rows.append(row)
            elif category == 'ffstock':
                ffstock = db_data['ffstock']
                details = db_data['ffstock_details']
                forecast_map = {}
                for it in db_data['forecast']:
                    forecast_map[it.product_code] = {
                        'plan': it.total_with_silo if it.total_with_silo else it.grand_total_tons,
                        'packing_size': it.packing_size,
                        'die_size': it.die_size
                    }
                product_names = {}
                for p_code, spec in db_data['congsuat'].items():
                    product_names[p_code] = spec.formular_code
                
                all_products = sorted(list(set(list(ffstock.keys()) + list(details.keys()))))
                for p in all_products:
                    det = details.get(p, {})
                    stock_tons = round(ffstock.get(p, 0.0), 1)
                    daily_avg = round(det.get('daily_sales_tons', 0.0), 1)
                    doh_val = det.get('doh', None)
                    doh = round(doh_val, 1) if doh_val is not None else None
                    fc = forecast_map.get(p, {})
                    plan = round(fc.get('plan', 0.0), 1) if fc.get('plan') else 0
                    day5 = round(plan / 6 * 5, 1) if plan > 0 else 0
                    animal = classify_animal_type(p)
                    kq_gc2 = round(stock_tons - plan, 1) if plan > 0 else 0
                    ghi_chu = ''
                    if doh is not None and doh < 1:
                        ghi_chu = f'⚠️ CẦN SX NGAY'
                    elif doh is not None and doh < 3:
                        ghi_chu = f'📋 DOH thấp'
                    rows.append({
                        'product_code': p,
                        'product_name': product_names.get(p, det.get('product_name', '')),
                        'animal_type': animal,
                        'animal_label': VAT_NUOI_LABELS.get(animal, 'HEO'),
                        'animal_color': VAT_NUOI_COLORS.get(animal, '#FF6B6B'),
                        'stock_tons': stock_tons,
                        'safety_stock_tons': round(det.get('safety_stock_tons', 0.0), 1),
                        'daily_sales_tons': daily_avg,
                        'doh': doh,
                        'plan': plan,
                        'day5': day5,
                        'kq_gc2': kq_gc2,
                        'warning': det.get('warning', ''),
                        'ghi_chu': ghi_chu
                    })
            elif category == 'tonbon':
                tonbon = db_data['tonbon']
                for p, t in sorted(tonbon.items()):
                    rows.append({
                        'product_code': p,
                        'tons': round(t, 1)
                    })
            elif category == 'empty_bag':
                empty_bag = db_data['empty_bag']
                for p, brands in sorted(empty_bag.items()):
                    rows.append({
                        'product_code': p,
                        'HIGRO': brands.get('HIGRO', 0),
                        'CP': brands.get('CP', 0),
                        'STAR': brands.get('STAR', 0),
                        'NASA': brands.get('NASA', 0),
                        'NUVO': brands.get('NUVO', 0),
                        'FARM': brands.get('FARM', 0)
                    })
            elif category == 'congsuat':
                congsuat = db_data['congsuat']
                seen_codes = set()
                for p, spec in sorted(congsuat.items()):
                    if spec.product_code not in seen_codes:
                        seen_codes.add(spec.product_code)
                        rows.append({
                            'product_code': spec.product_code,
                            'product_name': spec.product_name,
                            'formular_code': spec.formular_code,
                            'die_size': spec.die_size,
                            'ton_per_batch': spec.ton_per_batch,
                            'line_cv': spec.line_cv,
                            'line_pk': spec.line_pk,
                            'ks_code': spec.ks_code
                        })

            elif category == 'feedcode':
                feedcode = db_data['feedcode']
                for p, lines in sorted(feedcode.items()):
                    rows.append({
                        'product_code': p,
                        'line_cv': lines.get('line_cv', ''),
                        'line_pk': lines.get('line_pk', '')
                    })
            elif category == 'khangsinh':
                khangsinh = db_data['khangsinh']
                for p, ks in sorted(khangsinh.items()):
                    rows.append({
                        'product_code': p,
                        'ks_code': ks.get('ks_code', '') if isinstance(ks, dict) else ks
                    })
            elif category == 'yesterday_plan':
                yesterday = db_data['khsx_yesterday']
                for p, det in sorted(yesterday.items()):
                    rows.append({
                        'product_code': p,
                        'planned_batches': det.get('planned_batches', det.get('planned', 0)),
                        'actual_batches': det.get('actual_batches', det.get('actual', 0)),
                        'planned_tons': round(det.get('planned_tons', det.get('planned', 0) * 8.4), 1),
                        'actual_tons': round(det.get('actual_tons', det.get('actual', 0) * 8.4), 1),
                        'status': det.get('status', 'Thiếu' if det.get('planned', 0) > det.get('actual', 0) else 'Đạt')
                    })
            elif category == 'adjustments':
                adj = db_data['adjustments']
                
                # Flat additions
                additions = []
                for item in adj.get('additions', []):
                    additions.append({
                        'product_code': item.get('product_code', ''),
                        'tons': item.get('tons', 0.0),
                        'packing_size': item.get('packing_size', ''),
                        'priority': item.get('priority', ''),
                        'force_batches': item.get('force_batches', '') or '',
                        'force_tpb': item.get('force_tpb', '') or '',
                        'note': item.get('note', '')
                    })
                    
                # Flat cancellations
                cancellations = []
                for p, cancel_type in adj.get('cancellations', {}).items():
                    cancellations.append({
                        'product_code': p,
                        'cancel_type': cancel_type,
                        'note': ''
                    })
                    
                # Flat substitutions
                substitutions = []
                for old, new in adj.get('substitutions', {}).items():
                    substitutions.append({
                        'old_code': old,
                        'new_code': new,
                        'note': ''
                    })
                    
                # Flat bag substitutions
                bag_substitutions = []
                for p, mapping in adj.get('bag_substitutions', {}).items():
                    for old_bag, new_bag in mapping.items():
                        bag_substitutions.append({
                            'product_code': p,
                            'old_bag': old_bag,
                            'new_bag': new_bag,
                            'note': ''
                        })
                        
                return jsonify({
                    'success': True,
                    'file_info': info,
                    'data': {
                        'additions': additions,
                        'cancellations': cancellations,
                        'substitutions': substitutions,
                        'bag_substitutions': bag_substitutions
                    },
                    'sharepoint_url': getattr(config, 'SHAREPOINT_PLAN_URL', '')
                })

            return jsonify({
                'success': True,
                'file_info': info,
                'data': rows,
                'sharepoint_url': getattr(config, f"SHAREPOINT_{category.upper()}_URL", '')
            })

        info = {}
        rows = []
        
        # Lấy file tương ứng
        if category == 'forecast':
            info = get_file_info(config.FORECAST_DIR, '*FORECAST*.xlsx')
            if info['exists']:
                items = data_loader.load_forecast(info['path'])
                for it in items:
                    rows.append({
                        'product_code': it.product_code,
                        'packing_size': it.packing_size,
                        'die_size': it.die_size,
                        'dealer_higro': it.dealer_higro,
                        'dealer_cp': it.dealer_cp,
                        'dealer_star': it.dealer_star,
                        'dealer_nuvo': it.dealer_nuvo,
                        'dealer_nasa': it.dealer_nasa,
                        'dealer_total': it.dealer_total,
                        'farm_swine': it.farm_swine,
                        'farm_integrate': it.farm_integrate,
                        'farm_total': it.farm_total,
                        'grand_total_tons': it.grand_total_tons,
                        'silo_tons': it.silo_tons,
                        'total_with_silo': it.total_with_silo
                    })
                    
        elif category == 'silo_plan':
            info = get_file_info(config.SILO_DIR, '*SILO*.xlsx')
            if info['exists']:
                silo_plan = data_loader.load_silo_plan(info['path'])
                # silo_plan: dict {day → {product → tons}}
                all_products = sorted(list(set(p for d in silo_plan.values() for p in d)))
                for p in all_products:
                    row = {'product_code': p}
                    for day in range(1, 7):
                        row[f'day_{day}'] = round(silo_plan[day].get(p, 0.0), 1)
                    row['total'] = round(sum(row[f'day_{d}'] for d in range(1, 7)), 1)
                    rows.append(row)
                    
        elif category == 'bacang':
            info = get_file_info(config.BACANG_DIR, '*CANG*.xlsx')
            if info['exists']:
                bacang = data_loader.load_bacang(info['path'])
                # bacang: dict {day → {product → tons}}
                all_products = sorted(list(set(p for d in bacang.values() for p in d)))
                for p in all_products:
                    row = {'product_code': p}
                    for day in range(1, 7):
                        row[f'day_{day}'] = round(bacang[day].get(p, 0.0), 1)
                    row['total'] = round(sum(row[f'day_{d}'] for d in range(1, 7)), 1)
                    rows.append(row)
                    
        elif category == 'ffstock':
            info = get_file_info(getattr(config, 'FSTOCK_DIR_FFSTOCK', config.FSTOCK_DIR), '*FFSTOCK*.xls*')
            if info['exists']:
                ffstock = data_loader.load_ffstock(info['path'])
                details = data_loader.load_ffstock_details(info['path'])
                
                # Load forecast cho Plan column
                forecast_map = {}
                forecast_info = get_file_info(config.FORECAST_DIR, '*FORECAST*.xlsx')
                if forecast_info['exists']:
                    try:
                        fc_items = data_loader.load_forecast(forecast_info['path'])
                        for it in fc_items:
                            forecast_map[it.product_code] = {
                                'plan': it.total_with_silo if it.total_with_silo else it.grand_total_tons,
                                'packing_size': it.packing_size,
                                'die_size': it.die_size
                            }
                    except:
                        pass
                
                # Load tên cám từ FEEDCODE sheet (nếu có)
                product_names = {}
                try:
                    congsuat = data_loader.load_congsuat(config.PLAN_FILE)
                    for spec in congsuat:
                        if spec.product_code and spec.formular_code:
                            product_names[spec.product_code] = spec.formular_code
                except:
                    pass
                
                # Merge all data
                all_products = sorted(list(set(list(ffstock.keys()) + list(details.keys()))))
                for p in all_products:
                    det = details.get(p, {})
                    stock_tons = round(ffstock.get(p, 0.0), 1)
                    daily_avg = round(det.get('daily_sales_tons', 0.0), 1)
                    doh_val = det.get('doh', None)
                    doh = round(doh_val, 1) if doh_val is not None else None
                    
                    # Plan từ forecast
                    fc = forecast_map.get(p, {})
                    plan = round(fc.get('plan', 0.0), 1) if fc.get('plan') else 0
                    
                    # Day5 = forecast / 6 * 5 (5 ngày tới cần bao nhiêu)
                    day5 = round(plan / 6 * 5, 1) if plan > 0 else 0
                    
                    # Phân loại vật nuôi
                    animal = classify_animal_type(p)
                    
                    # KQ GC2: so sánh Stock với nhu cầu
                    kq_gc2 = ''
                    if daily_avg > 0 and stock_tons > 0:
                        kq_gc2 = round(stock_tons - plan, 1) if plan > 0 else 0
                    
                    # Ghi chú tự động
                    ghi_chu = ''
                    if doh is not None and doh < 1:
                        ghi_chu = f'⚠️ CẦN SX NGAY'
                    elif doh is not None and doh < 3:
                        ghi_chu = f'📋 DOH thấp'
                    
                    rows.append({
                        'product_code': p,
                        'product_name': product_names.get(p, det.get('product_name', '')),
                        'animal_type': animal,
                        'animal_label': VAT_NUOI_LABELS.get(animal, 'HEO'),
                        'animal_color': VAT_NUOI_COLORS.get(animal, '#FF6B6B'),
                        'stock_tons': stock_tons,
                        'safety_stock_tons': round(det.get('safety_stock_tons', 0.0), 1),
                        'daily_sales_tons': daily_avg,
                        'doh': doh,
                        'plan': plan,
                        'day5': day5,
                        'kq_gc2': kq_gc2,
                        'warning': det.get('warning', ''),
                        'ghi_chu': ghi_chu
                    })
                    
        elif category == 'tonbon':
            info = get_file_info(config.TONBON_DIR, '*ton bon*.*')
            if info['exists']:
                tonbon = data_loader.load_tonbon(info['path'])
                # tonbon: dict {product → tons}
                for p, t in sorted(tonbon.items()):
                    rows.append({
                        'product_code': p,
                        'tons': round(t, 1)
                    })
                    
        elif category == 'empty_bag':
            info = get_file_info(getattr(config, 'FSTOCK_DIR_EMPTYBAG', config.FSTOCK_DIR), '*EMPTY BAG*.xls*')
            if info['exists']:
                empty_bag = data_loader.load_empty_bag(info['path'])
                # empty_bag: dict {product → {brand → bags}}
                for p, brands in sorted(empty_bag.items()):
                    rows.append({
                        'product_code': p,
                        'HIGRO': brands.get('HIGRO', 0),
                        'CP': brands.get('CP', 0),
                        'STAR': brands.get('STAR', 0),
                        'NASA': brands.get('NASA', 0),
                        'NUVO': brands.get('NUVO', 0),
                        'FARM': brands.get('FARM', 0)
                    })
                    
        elif category == 'congsuat':
            info = get_file_info(None, None, exact_path=config.PLAN_FILE)
            if info['exists']:
                congsuat = data_loader.load_congsuat(info['path'])
                db_data = {'congsuat': congsuat}
                khsx_info = get_file_info(None, None, exact_path=config.KHSX_FILE)
                if khsx_info['exists']:
                    db_data['feedcode'] = data_loader.load_feedcode(khsx_info['path'])
                    db_data['khangsinh'] = data_loader.load_khangsinh(khsx_info['path'])
                db_data['fix_code_pellet'] = data_loader.load_fix_code_pellet(info['path'])
                data_loader.enrich_congsuat(db_data)
                
                seen_codes = set()
                for p, spec in sorted(congsuat.items()):
                    if spec.product_code not in seen_codes:
                        seen_codes.add(spec.product_code)
                        rows.append({
                            'product_code': spec.product_code,
                            'product_name': spec.product_name,
                            'formular_code': spec.formular_code,
                            'die_size': spec.die_size,
                            'ton_per_batch': spec.ton_per_batch,
                            'line_cv': spec.line_cv,
                            'line_pk': spec.line_pk,
                            'ks_code': spec.ks_code
                        })

                    
        elif category == 'feedcode':
            info = get_file_info(None, None, exact_path=config.KHSX_FILE)
            if info['exists']:
                feedcode = data_loader.load_feedcode(info['path'])
                # feedcode: dict {product → {line_cv, line_pk}}
                for p, lines in sorted(feedcode.items()):
                    rows.append({
                        'product_code': p,
                        'line_cv': lines.get('line_cv', ''),
                        'line_pk': lines.get('line_pk', '')
                    })
                    
        elif category == 'khangsinh':
            info = get_file_info(None, None, exact_path=config.KHSX_FILE)
            if info['exists']:
                khangsinh = data_loader.load_khangsinh(info['path'])
                # khangsinh: dict {product → ks_code}
                for p, ks in sorted(khangsinh.items()):
                    rows.append({
                        'product_code': p,
                        'ks_code': ks
                    })
                    
        elif category == 'yesterday_plan':
            info = get_file_info(None, None, exact_path=config.KHSX_FILE)
            if info['exists']:
                # Mặc định lấy ngày hôm qua dựa trên ngày hiện tại
                t = datetime.date.today()
                yesterday_day = t.day - 1
                if yesterday_day < 1: yesterday_day = 1
                
                yesterday = data_loader.load_khsx_yesterday(info['path'], yesterday_day)
                # yesterday: dict {product → dict}
                for p, det in sorted(yesterday.items()):
                    rows.append({
                        'product_code': p,
                        'planned_batches': det.get('planned_batches', 0),
                        'actual_batches': det.get('actual_batches', 0),
                        'planned_tons': round(det.get('planned_tons', 0.0), 1),
                        'actual_tons': round(det.get('actual_tons', 0.0), 1),
                        'status': det.get('status', '')
                    })
                    
        elif category == 'adjustments':
            # Trả về dưới dạng 4 danh sách riêng biệt cho các bảng editor
            info = get_file_info(None, None, exact_path=config.QUICK_ADJUST_FILE)
            if info['exists']:
                adj = data_loader.load_quick_adjustments(info['path'])
                
                # Flat additions
                additions = []
                for item in adj.get('additions', []):
                    additions.append({
                        'product_code': item.get('product_code', ''),
                        'tons': item.get('tons', 0.0),
                        'packing_size': item.get('packing_size', ''),
                        'priority': item.get('priority', ''),
                        'force_batches': item.get('force_batches', ''),
                        'force_tpb': item.get('force_tpb', ''),
                        'note': item.get('note', '')
                    })
                    
                # Flat cancellations
                cancellations = []
                for p, cancel_type in adj.get('cancellations', {}).items():
                    cancellations.append({
                        'product_code': p,
                        'cancel_type': cancel_type,
                        'note': ''
                    })
                    
                # Flat substitutions
                substitutions = []
                for old, new in adj.get('substitutions', {}).items():
                    substitutions.append({
                        'old_code': old,
                        'new_code': new,
                        'note': ''
                    })
                    
                # Flat bag substitutions
                bag_substitutions = []
                for p, mapping in adj.get('bag_substitutions', {}).items():
                    for old_bag, new_bag in mapping.items():
                        bag_substitutions.append({
                            'product_code': p,
                            'old_bag': old_bag,
                            'new_bag': new_bag,
                            'note': ''
                        })
                        
                return jsonify({
                    'success': True,
                    'file_info': info,
                    'data': {
                        'additions': additions,
                        'cancellations': cancellations,
                        'substitutions': substitutions,
                        'bag_substitutions': bag_substitutions
                    },
                    'sharepoint_url': getattr(config, 'SHAREPOINT_PLAN_URL', '')
                })
        
        return jsonify({
            'success': True,
            'file_info': info,
            'data': rows,
            'sharepoint_url': getattr(config, f"SHAREPOINT_{category.upper()}_URL", '')
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# API: UPLOAD FILE ĐẦU VÀO
# ============================================================

@app.route('/api/upload/<category>', methods=['POST'])
def upload_file(category):
    """Admin tải lên báo cáo Excel mới. Tự động lưu vào đúng thư mục cấu hình."""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Không tìm thấy file tải lên'})
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Tên file không hợp lệ'})
            
        # Kiểm tra phần mở rộng file Excel
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ['.xlsx', '.xls', '.xlsm']:
            return jsonify({'success': False, 'message': 'Chỉ chấp nhận file Excel (.xlsx, .xls, .xlsm)'})
            
        # Xác định thư mục đích dựa theo config
        target_dir = None
        prefix = f"UPLOAD_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_"
        
        if category == 'forecast':
            target_dir = config.FORECAST_DIR
        elif category == 'silo_plan':
            target_dir = config.SILO_DIR
        elif category == 'bacang':
            target_dir = config.BACANG_DIR
        elif category == 'ffstock':
            target_dir = getattr(config, 'FSTOCK_DIR_FFSTOCK', config.FSTOCK_DIR)
        elif category == 'tonbon':
            target_dir = config.TONBON_DIR
        elif category == 'empty_bag':
            target_dir = getattr(config, 'FSTOCK_DIR_EMPTYBAG', config.FSTOCK_DIR)
        elif category == 'congsuat':
            # file cố định Plan.xlsm
            target_dir = config.PLAN_DIR
            prefix = ""  # Ghi đè trực tiếp
            filename = "Plan.xlsm"
        elif category == 'feedcode' or category == 'khangsinh' or category == 'yesterday_plan':
            # file cố định KHSX THANG 5-20261.xlsm
            target_dir = config.DATA_DIR
            prefix = ""  # Ghi đè trực tiếp
            filename = "KHSX THANG 5-20261.xlsm"
        elif category == 'adjustments':
            target_dir = config.DATA_DIR
            prefix = ""  # Ghi đè trực tiếp
            filename = "DIEU_CHINH_NHANH.xlsx"
            
        if not target_dir:
            return jsonify({'success': False, 'message': f'Danh mục upload không hợp lệ: {category}'})
            
        # Tạo thư mục nếu chưa có
        os.makedirs(target_dir, exist_ok=True)
        
        if prefix != "":
            filename = prefix + file.filename
            
        target_path = os.path.join(target_dir, filename)
        
        # Nếu ghi đè file Plan/KHSX gốc, sao lưu file cũ trước
        if prefix == "" and os.path.isfile(target_path):
            backup_path = target_path + f".bak_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            try:
                os.rename(target_path, backup_path)
            except Exception as ex:
                print(f"Không thể sao lưu file cũ: {ex}")
                
        file.save(target_path)
        
        # Nếu đang sử dụng PostgreSQL, đồng bộ dữ liệu của file vừa upload lên cloud ngay lập tức
        if getattr(config, 'USE_POSTGRESQL', False):
            try:
                import db_manager
                db_manager.sync_category_to_db(config, category, getattr(config, 'DB_URI', db_manager.DB_URI))
                print(f"⚡ Đã đồng bộ tự động dữ liệu upload {category.upper()} lên PostgreSQL!")
            except Exception as sync_ex:
                print(f"⚠️ Lỗi đồng bộ dữ liệu upload lên database: {sync_ex}")
                # Hủy bỏ file vừa lưu để tránh bất đồng bộ
                try:
                    os.remove(target_path)
                except:
                    pass
                return jsonify({
                    'success': False,
                    'message': f"Upload thành công nhưng đồng bộ lên Database thất bại: {str(sync_ex)}"
                })
        
        # Trả về thông tin file mới
        info = get_file_info(None, None, exact_path=target_path)
        return jsonify({
            'success': True,
            'message': f"Đã upload và đồng bộ thành công dữ liệu {category.upper()} lên hệ thống!",
            'file_info': info
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# API: QUẢN TRỊ ĐIỀU CHỈNH NHANH (DIEU_CHINH_NHANH.xlsx)
# ============================================================

@app.route('/api/adjustments/save', methods=['POST'])
def save_adjustments():
    """Lưu trực tiếp dữ liệu từ bảng editor vào DIEU_CHINH_NHANH.xlsx"""
    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'message': 'Dữ liệu trống'})
            
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        file_path = config.QUICK_ADJUST_FILE
        wb = Workbook()
        
        # Định dạng style mẫu CP
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        # 1. THEM_MOI_HOAC_SUA
        ws1 = wb.active
        ws1.title = "THEM_MOI_HOAC_SUA"
        headers1 = ["MÃ CÁM", "TẤN", "QUY CÁCH", "LOẠI ƯU TIÊN", "ÉP SỐ MẺ", "ÉP TẤN/MẺ", "GHI CHÚ (Tùy chọn)"]
        ws1.append(headers1)
        
        additions = data.get('additions', [])
        for item in additions:
            ws1.append([
                str(item.get('product_code', '')).strip().upper(),
                _safe_float_val(item.get('tons')),
                str(item.get('packing_size', '')).strip().upper(),
                str(item.get('priority', '')).strip().upper(),
                _safe_int_val(item.get('force_batches')),
                _safe_float_val(item.get('force_tpb')),
                item.get('note', '')
            ])
            
        # 2. HUY_KHSX
        ws2 = wb.create_sheet("HUY_KHSX")
        headers2 = ["MÃ CÁM", "LOẠI HỦY", "GHI CHÚ"]
        ws2.append(headers2)
        cancellations = data.get('cancellations', [])
        for item in cancellations:
            ws2.append([
                str(item.get('product_code', '')).strip().upper(),
                str(item.get('cancel_type', '')).strip().upper(),
                item.get('note', '')
            ])
            
        # 3. THAY_THE_MA_CAM
        ws3 = wb.create_sheet("THAY_THE_MA_CAM")
        headers3 = ["MÃ CŨ", "MÃ MỚI", "GHI CHÚ"]
        ws3.append(headers3)
        substitutions = data.get('substitutions', [])
        for item in substitutions:
            ws3.append([
                str(item.get('old_code', '')).strip().upper(),
                str(item.get('new_code', '')).strip().upper(),
                item.get('note', '')
            ])
            
        # 4. THAY_THE_BAO_BI
        ws4 = wb.create_sheet("THAY_THE_BAO_BI")
        headers4 = ["MÃ CÁM", "BAO GỐC", "BAO THAY THẾ", "GHI CHÚ"]
        ws4.append(headers4)
        bag_substitutions = data.get('bag_substitutions', [])
        for item in bag_substitutions:
            ws4.append([
                str(item.get('product_code', '')).strip().upper(),
                str(item.get('old_bag', '')).strip().upper(),
                str(item.get('new_bag', '')).strip().upper(),
                item.get('note', '')
            ])
            
        # Apply header styling cho tất cả sheets
        for sheet in wb.worksheets:
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                
        # Sao lưu file cũ trước khi lưu
        if os.path.isfile(file_path):
            backup_path = file_path + f".bak_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            try: os.rename(file_path, backup_path)
            except: pass
            
        wb.save(file_path)
        
        info = get_file_info(None, None, exact_path=file_path)
        return jsonify({
            'success': True,
            'message': 'Đã ghi nhận thay đổi và cập nhật DIEU_CHINH_NHANH.xlsx thành công!',
            'file_info': info
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# API: CHẠY THUẬT TOÁN KHSX (LIVE SSE STREAM)
# ============================================================

@app.route('/api/generate-plan', methods=['POST'])
def generate_plan():
    """Nhận ngày tính toán, chạy script tối ưu hóa KHSX dưới dạng luồng Stream logs SSE"""
    try:
        req_data = request.json or {}
        target_date_str = req_data.get('date')  # YYYY-MM-DD
        walkin_orders = req_data.get('walkin_orders', [])
        
        if not target_date_str:
            return jsonify({'success': False, 'message': 'Ngày không hợp lệ'})
            
        # Parse ngày sang định dạng YYYYMMDD cho script
        dt = datetime.datetime.strptime(target_date_str, '%Y-%m-%d')
        date_param = dt.strftime('%Y%m%d')
        date_dmy = dt.strftime('%d-%m-%Y')
        
        # 1. Tạo file đơn vãng lai tạm thời nếu có
        walkin_file_path = None
        if walkin_orders:
            temp_fd, walkin_file_path = tempfile.mkstemp(suffix='.csv', text=True)
            with open(walkin_file_path, 'w', encoding='utf-8', newline='') as csvfile:
                writer = csv.writer(csvfile)
                for order in walkin_orders:
                    writer.writerow([
                        str(order.get('product', '')).strip().upper(),
                        float(order.get('tons', 0.0)),
                        str(order.get('packing_size', '25')).strip()
                    ])
            os.close(temp_fd)
            
        # 2. Định nghĩa hàm generator để stream dữ liệu log console
        def sse_generator():
            # Chạy script Python khsx_auto.py dưới dạng tiến trình con
            cmd = ['py', 'khsx_auto.py', '--date', date_param]
            if walkin_file_path:
                cmd.extend(['--walkin', walkin_file_path])
                
            yield "data: " + json.dumps({'type': 'log', 'text': f"🚀 Khởi động chương trình KHSX tự động cho ngày {date_dmy}..."}) + "\n\n"
            
            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding='utf-8',
                cwd=CURRENT_DIR
            )
            
            # Đọc output từng dòng
            for line in process.stdout:
                line_clean = line.rstrip()
                if line_clean:
                    # Gửi log từng dòng về frontend
                    yield "data: " + json.dumps({'type': 'log', 'text': line_clean}) + "\n\n"
                    
            process.wait()
            
            # Xóa file đơn vãng lai tạm thời
            if walkin_file_path and os.path.exists(walkin_file_path):
                try: os.remove(walkin_file_path)
                except: pass
                
            if process.returncode == 0:
                # Tìm file KHSX đầu ra vừa tạo
                import glob
                pattern = os.path.join(config.OUTPUT_DIR, f"KHSX_{date_dmy}*.xlsx")
                files = glob.glob(pattern)
                
                output_filename = ""
                if files:
                    files.sort(key=os.path.getmtime, reverse=True)
                    output_filename = os.path.basename(files[0])
                    
                yield "data: " + json.dumps({
                    'type': 'complete',
                    'success': True,
                    'output_file': output_filename,
                    'message': f"🎉 Đã lập KHSX thành công cho ngày {date_dmy}!"
                }) + "\n\n"
            else:
                yield "data: " + json.dumps({
                    'type': 'complete',
                    'success': False,
                    'message': f"❌ Lỗi thực thi thuật toán (Mã thoát: {process.returncode}). Vui lòng kiểm tra lại dữ liệu Excel đầu vào!"
                }) + "\n\n"
                
        return Response(sse_generator(), mimetype='text/event-stream')
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# HELPER: PHÂN LOẠI VẬT NUÔI TỰ ĐỘNG TỪ MÃ CÁM
# ============================================================

# Bảng mapping vật nuôi
VAT_NUOI_LABELS = {'H': 'HEO', 'G': 'GÀ', 'B': 'BÒ', 'V': 'VỊT', 'C': 'CÚT', 'D': 'DÊ'}
VAT_NUOI_COLORS = {'H': '#FF6B6B', 'G': '#4ECDC4', 'B': '#45B7D1', 'V': '#96CEB4', 'C': '#FFEAA7', 'D': '#DDA0DD'}

def classify_animal_type(product_code: str) -> str:
    """
    Phân loại vật nuôi tự động dựa theo prefix mã cám.
    Quy tắc: 1xx-5xx = Heo, 6xx-7xx = Gà, 8xx = Vịt, 9xx = Bò/Dê/Cút
    Trả về mã 1 ký tự: H/G/B/V/C/D
    """
    code = str(product_code).strip().upper()
    if not code:
        return 'H'
    
    # Lấy ký tự đầu tiên (số)
    first_char = code[0]
    
    if first_char in ('1', '2', '3', '4', '5'):
        return 'H'  # HEO
    elif first_char in ('6', '7'):
        return 'G'  # GÀ
    elif first_char == '8':
        # 8xx có thể là Vịt hoặc đặc biệt
        if code.startswith('85') or code.startswith('86') or code.startswith('87'):
            return 'V'  # VỊT
        return 'V'  # Mặc định 8xx = Vịt
    elif first_char == '9':
        # 9xx: Bò, Cút, Dê - cần kiểm tra chi tiết hơn
        if code.startswith('92') or code.startswith('93'):
            return 'B'  # BÒ
        elif code.startswith('95'):
            return 'C'  # CÚT
        elif code.startswith('96') or code.startswith('97'):
            return 'D'  # DÊ
        return 'B'  # Mặc định 9xx = Bò
    else:
        # Mã không bắt đầu bằng số: kiểm tra keyword
        if 'DUCK' in code or 'VIT' in code:
            return 'V'
        elif 'QUAIL' in code or 'CUT' in code:
            return 'C'
        elif 'GOAT' in code or 'DE' in code:
            return 'D'
        elif 'CATTLE' in code or 'BO' in code or 'DAIRY' in code:
            return 'B'
        elif 'BROILER' in code or 'LAYER' in code or 'GA' in code or 'CHICK' in code:
            return 'G'
        return 'H'  # Mặc định = Heo (chiếm đa số)


def load_doh_data_for_sequence(sequence_items):
    """
    Tải dữ liệu DOH (Days on Hand) cho các sản phẩm trong mixer sequence.
    Nguồn 1: FFSTOCK details (cột DOH có sẵn)
    Nguồn 2: Tính từ forecast = stock / (forecast_week / 6)
    Trả về dict {product_code → {doh, stock, daily_avg}}
    """
    doh_map = {}
    
    try:
        # Nguồn 1: Đọc DOH từ FFSTOCK 
        ffstock_info = get_file_info(getattr(config, 'FSTOCK_DIR_FFSTOCK', config.FSTOCK_DIR), '*FFSTOCK*.xls*')
        if ffstock_info['exists']:
            ffstock = data_loader.load_ffstock(ffstock_info['path'])
            details = data_loader.load_ffstock_details(ffstock_info['path'])
            
            for p in set(list(ffstock.keys()) + list(details.keys())):
                det = details.get(p, {})
                stock_tons = ffstock.get(p, 0.0)
                doh_val = det.get('doh', None)
                daily_avg = det.get('daily_sales_tons', 0.0)
                
                doh_map[p] = {
                    'stock': round(stock_tons, 1),
                    'doh': round(doh_val, 1) if doh_val and doh_val > 0 else None,
                    'daily_avg': round(daily_avg, 1) if daily_avg else 0.0
                }
        
        # Nguồn 2: Bổ sung DOH từ forecast cho sản phẩm chưa có
        forecast_info = get_file_info(config.FORECAST_DIR, '*FORECAST*.xlsx')
        if forecast_info['exists']:
            items = data_loader.load_forecast(forecast_info['path'])
            for it in items:
                pc = it.product_code
                forecast_week = it.total_with_silo if it.total_with_silo else it.grand_total_tons
                
                if pc in doh_map and doh_map[pc]['doh'] is None and forecast_week and forecast_week > 0:
                    stock = doh_map[pc].get('stock', 0.0)
                    daily_forecast = forecast_week / 6  # 6 ngày làm việc
                    if daily_forecast > 0:
                        doh_map[pc]['doh'] = round(stock / daily_forecast, 1)
                        doh_map[pc]['daily_avg'] = round(daily_forecast, 1)
                elif pc not in doh_map and forecast_week and forecast_week > 0:
                    daily_forecast = forecast_week / 6
                    doh_map[pc] = {
                        'stock': 0.0,
                        'doh': 0.0,
                        'daily_avg': round(daily_forecast, 1)
                    }
                    
    except Exception as e:
        print(f"⚠️ Lỗi khi tải DOH data: {e}")
    
    return doh_map


# ============================================================
# API: TẢI CHI TIẾT KẾ HOẠCH ĐÃ LẬP (PL, PACKAGING, SEQUENCE)
# ============================================================

@app.route('/api/plan-details/<date_str>', methods=['GET'])
def get_plan_details(date_str):
    """
    Đọc ngược dữ liệu từ file Excel kết quả KHSX vừa tạo 
    để hiển thị trực tiếp lên UI (Kế hoạch Pellet Line, Kế hoạch Đóng bao, Mixer Sequence)
    """
    try:
        # Nếu đang sử dụng PostgreSQL, đọc kế hoạch trực tiếp từ cơ sở dữ liệu
        if getattr(config, 'USE_POSTGRESQL', False):
            import db_manager
            res_db = db_manager.get_plan_output_from_db(date_str, getattr(config, 'DB_URI', db_manager.DB_URI))
            if res_db.get('exists'):
                print(f"⚡ Đọc kế hoạch ngày {date_str} trực tiếp từ PostgreSQL Cloud...")
                
                sequence = res_db['sequence']
                packaging = res_db['packaging']
                summary = res_db['summary']
                khpl_raw_grid = res_db['khpl_raw_grid']
                filename = res_db['filename']
                
                # Dịch digit code → tên cám dân dã cho khpl_raw_grid từ DB
                db_code_mapping = {}
                try:
                    conn_cm = db_manager.get_connection()
                    cur_cm = conn_cm.cursor()
                    cur_cm.execute("SELECT digit_code, colloquial_name FROM code_mapping;")
                    for dc, cn in cur_cm.fetchall():
                        if dc and cn:
                            db_code_mapping[str(dc).strip().upper()] = str(cn).strip().upper()
                    # Bổ sung từ congsuat
                    cur_cm.execute("SELECT product_code, product_name FROM congsuat;")
                    for pc, pn in cur_cm.fetchall():
                        pc_n = str(pc).strip().upper() if pc else ''
                        pn_n = str(pn).strip().upper() if pn else ''
                        if pc_n and pn_n and pc_n != pn_n and pc_n not in db_code_mapping:
                            db_code_mapping[pc_n] = pn_n
                    cur_cm.close()
                    conn_cm.close()
                except Exception:
                    pass
                
                if db_code_mapping and khpl_raw_grid:
                    product_cols = [1, 5, 9, 13, 17, 21, 25, 29, 31]  # 0-indexed
                    for row_idx in range(2, min(len(khpl_raw_grid), 22)):
                        for col_idx in product_cols:
                            if col_idx < len(khpl_raw_grid[row_idx]):
                                cell_val = khpl_raw_grid[row_idx][col_idx]
                                if cell_val and cell_val != '':
                                    raw_str = str(cell_val).strip()
                                    if raw_str.endswith('.0'):
                                        raw_str = raw_str[:-2]
                                    norm = raw_str.upper().replace(' ', '')
                                    translated = db_code_mapping.get(norm, None)
                                    if translated and translated != norm:
                                        khpl_raw_grid[row_idx][col_idx] = translated
                
                # Tái tạo pl_plans từ sequence
                pl_plans = {f'PL{i}': [] for i in range(1, 8)}
                pl_plans['MASH'] = []
                for item in sequence:
                    line = item.get('line_cv', '').upper().replace(' ', '')
                    if 'PL1' in line: pl_plans['PL1'].append(item)
                    elif 'PL2' in line: pl_plans['PL2'].append(item)
                    elif 'PL3' in line: pl_plans['PL3'].append(item)
                    elif 'PL4' in line: pl_plans['PL4'].append(item)
                    elif 'PL5' in line: pl_plans['PL5'].append(item)
                    elif 'PL6' in line: pl_plans['PL6'].append(item)
                    elif 'PL7' in line: pl_plans['PL7'].append(item)
                    elif 'MASH' in line or item.get('packing_size') == 'M': pl_plans['MASH'].append(item)
                    
                # pl_rich_plans
                pl_rich_plans = {}
                for m in range(1, 8):
                    m_key = f"PL{m}"
                    pl_rich_plans[m_key] = {
                        'ton_dau': [], 'ca_1': [], 'ca_2': [], 'ca_3': [],
                        'loss_hours': 0, 'next_day': None, 'actual_hours': 0, 'actual_tons': 0
                    }
                
                return jsonify({
                    'success': True,
                    'summary': summary,
                    'sequence': sequence,
                    'pl_plans': pl_plans,
                    'pl_rich_plans': pl_rich_plans,
                    'mash_plan_rich': [],
                    'mixer_summary_rich': [],
                    'khpl_raw_grid': khpl_raw_grid,
                    'packaging': packaging,
                    'filename': filename
                })
        # date_str format: dd-mm-yyyy (VD: 21-05-2026)
        import glob
        pattern = os.path.join(config.OUTPUT_DIR, f"KHSX_{date_str}*.xlsx")
        files = glob.glob(pattern)
        
        if not files:
            return jsonify({'success': False, 'message': f'Không tìm thấy file kế hoạch sản xuất cho ngày {date_str}'})
            
        files.sort(key=os.path.getmtime, reverse=True)
        file_path = files[0]
        
        # Đọc dữ liệu từ file Excel KHSX kết quả bằng openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active # Sheet chính
        
        # Xây dựng danh sách mapping line máy mặc định từ sheet FEEDCODE làm fallback
        feedcode_mapping = {}
        if 'FEEDCODE' in wb.sheetnames:
            fc_sheet = wb['FEEDCODE']
            for fcr in range(2, 500):
                f_name = fc_sheet.cell(row=fcr, column=2).value
                f_cv = fc_sheet.cell(row=fcr, column=3).value
                f_pk = fc_sheet.cell(row=fcr, column=4).value
                if f_name:
                    key = str(f_name).replace(" ", "").strip().upper()
                    feedcode_mapping[key] = {
                        'line_cv': str(f_cv or '').strip(),
                        'line_pk': str(f_pk or '').strip()
                    }
        
        # Xây dựng danh sách mapping kháng sinh tự động đề phòng công thức Excel chưa chạy
        ks_mapping = {}
        if 'KHÁNG SINH' in wb.sheetnames:
            ks_sheet = wb['KHÁNG SINH']
            for ksr in range(3, 2000):
                k_prod = ks_sheet.cell(row=ksr, column=2).value
                k_code = ks_sheet.cell(row=ksr, column=3).value
                if k_prod:
                    ks_mapping[str(k_prod).strip().upper()] = str(k_code).strip()
                    
        # 1. Đọc Mixer Sequence (Bảng kế hoạch sản xuất tổng hợp)
        # Bắt đầu từ dòng 7 đến 41
        sequence = []
        summary = {
            'total_batches': 0,
            'total_tons': 0.0,
            'product_count': 0,
            'warnings': []
        }
        
        for r in range(7, 42): # Dòng 7 đến 41
            prod = ws.cell(row=r, column=2).value # B
            if not prod:
                continue
                
            prod_code = str(prod).strip().upper()
            if prod_code in {'TỔNG CỘNG', 'TOTAL', ''}:
                continue
                
            batches = _safe_int_val(ws.cell(row=r, column=3).value)
            if not batches or batches == '':
                batches = 0
                
            # Đọc các cột bao bì để tự tính tons và quy cách an toàn
            higro_25 = round(_safe_float_num(ws.cell(row=r, column=5).value), 1)
            higro_40 = round(_safe_float_num(ws.cell(row=r, column=6).value), 1)
            cp_25 = round(_safe_float_num(ws.cell(row=r, column=7).value), 1)
            cp_40 = round(_safe_float_num(ws.cell(row=r, column=8).value), 1)
            star_25 = round(_safe_float_num(ws.cell(row=r, column=9).value), 1)
            star_40 = round(_safe_float_num(ws.cell(row=r, column=10).value), 1)
            nuvo_25 = round(_safe_float_num(ws.cell(row=r, column=11).value), 1)
            nuvo_40 = round(_safe_float_num(ws.cell(row=r, column=12).value), 1)
            bell_25 = round(_safe_float_num(ws.cell(row=r, column=13).value), 1)
            bell_40 = round(_safe_float_num(ws.cell(row=r, column=14).value), 1)
            nasa_25 = round(_safe_float_num(ws.cell(row=r, column=15).value), 1)
            nasa_40 = round(_safe_float_num(ws.cell(row=r, column=16).value), 1)
            white_25 = round(_safe_float_num(ws.cell(row=r, column=17).value), 1)
            white_40 = round(_safe_float_num(ws.cell(row=r, column=18).value), 1)
            white_50 = round(_safe_float_num(ws.cell(row=r, column=19).value), 1)
            silo_truck = round(_safe_float_num(ws.cell(row=r, column=20).value), 1)
            
            # Tính toán tons từ công thức nếu openpyxl trả về None/rỗng do chưa mở file
            raw_tons = ws.cell(row=r, column=4).value
            if raw_tons is None or raw_tons == '' or (isinstance(raw_tons, str) and raw_tons.startswith('=')):
                # Tự tính dựa trên batches và mã cám giống công thức Excel
                if batches > 0:
                    if prod_code.startswith('550') or prod_code.startswith('551') or prod_code == '325F':
                        tons = float(batches) * 8.0
                    else:
                        tons = float(batches) * 8.4
                else:
                    tons = 0.0
            else:
                tons = _safe_float_num(raw_tons)
            
            tons = round(tons, 1)

            # Xác định quy cách đóng gói (packing_size) - Hỗ trợ quy cách ghép Silo + Bao
            has_silo = (silo_truck > 0)
            has_bag_25 = (higro_25 > 0 or cp_25 > 0 or star_25 > 0 or nuvo_25 > 0 or bell_25 > 0 or nasa_25 > 0 or white_25 > 0)
            has_bag_40 = (higro_40 > 0 or cp_40 > 0 or star_40 > 0 or nuvo_40 > 0 or bell_40 > 0 or nasa_40 > 0 or white_40 > 0)
            has_bag_50 = (white_50 > 0)
            
            if has_silo and (has_bag_25 or has_bag_40 or has_bag_50):
                bag_types = []
                if has_bag_25: bag_types.append("Bao 25")
                if has_bag_40: bag_types.append("Bao 40")
                if has_bag_50: bag_types.append("Bao 50")
                packing_size = "Silo + " + " + ".join(bag_types)
            elif has_silo:
                packing_size = 'M'
            elif has_bag_50:
                packing_size = '50'
            elif has_bag_40:
                packing_size = '40'
            else:
                packing_size = '25'
                
            # Phân rã tấn theo quy cách để truyền lên UI
            silo_tons = round(silo_truck, 1)
            bag_25_tons = round(higro_25 + cp_25 + star_25 + nuvo_25 + bell_25 + nasa_25 + white_25, 1)
            bag_40_tons = round(higro_40 + cp_40 + star_40 + nuvo_40 + bell_40 + nasa_40 + white_40, 1)
            bag_50_tons = round(white_50, 1)
                
            line_cv = str(ws.cell(row=r, column=22).value or '').strip()
            line_pk = str(ws.cell(row=r, column=23).value or '').strip()
            
            # Khôi phục line_cv từ feedcode_mapping nếu trống
            if not line_cv or line_cv == 'None':
                line_cv = feedcode_mapping.get(prod_code, {}).get('line_cv', '')
            if line_cv and line_cv != 'None':
                if line_cv.isdigit():
                    line_cv = f"PL{line_cv}"
                elif not line_cv.upper().startswith('PL') and not line_cv.upper() == 'MASH':
                    line_cv = f"PL{line_cv}"
                    
            # Khôi phục line_pk từ feedcode_mapping nếu trống
            if not line_pk or line_pk == 'None':
                line_pk = feedcode_mapping.get(prod_code, {}).get('line_pk', '')
            if has_silo and (not line_pk or line_pk == 'None'):
                line_pk = 'SILO'
            
            # Đọc kháng sinh (cột 21 - U), nếu là công thức chưa chạy thì tự VLOOKUP bằng Python
            raw_ks = ws.cell(row=r, column=21).value
            if raw_ks is None or raw_ks == '' or (isinstance(raw_ks, str) and raw_ks.startswith('=')):
                ks_code = ks_mapping.get(prod_code, 'SẠCH (KHÔNG KS)')
            else:
                ks_code = str(raw_ks).strip()
            
            # Cấp độ kháng sinh để phối màu
            ks_level = 1
            if ks_code and ks_code != 'SẠCH (KHÔNG KS)' and ks_code != 'Sạch không KS':
                ks_level = 5 # Đặt mức tượng trưng
                
            sequence.append({
                'product_code': prod_code,
                'batches': batches,
                'tons': tons,
                'packing_size': packing_size,
                'silo_tons': silo_tons,
                'bag_25_tons': bag_25_tons,
                'bag_40_tons': bag_40_tons,
                'bag_50_tons': bag_50_tons,
                'higro_25': higro_25,
                'higro_40': higro_40,
                'cp_25': cp_25,
                'cp_40': cp_40,
                'star_25': star_25,
                'star_40': star_40,
                'nuvo_25': nuvo_25,
                'nuvo_40': nuvo_40,
                'bell_25': bell_25,
                'bell_40': bell_40,
                'nasa_25': nasa_25,
                'nasa_40': nasa_40,
                'white_25': white_25,
                'white_40': white_40,
                'white_50': white_50,
                'silo_truck': silo_truck,
                'line_cv': line_cv,
                'line_pk': line_pk,
                'ks_code': ks_code,
                'ks_level': ks_level
            })
            
            summary['total_batches'] += batches
            summary['total_tons'] += tons
            summary['product_count'] += 1
            
        summary['total_tons'] = round(summary['total_tons'], 1)
        
        # Enrichment: Gắn DOH + phân loại vật nuôi cho mỗi sản phẩm
        doh_map = load_doh_data_for_sequence(sequence)
        
        for item in sequence:
            pc = item['product_code']
            
            # Gắn DOH
            doh_info = doh_map.get(pc, {})
            item['doh'] = doh_info.get('doh', None)
            item['stock'] = doh_info.get('stock', 0.0)
            item['daily_avg'] = doh_info.get('daily_avg', 0.0)
            

        
        # Đọc cảnh báo từ dòng 46 trở đi

        for r in range(46, 60):
            cell_val = ws.cell(row=r, column=2).value
            if cell_val and str(cell_val).startswith('⚠️'):
                summary['warnings'].append(str(cell_val).strip())
                
        # 2. Xây dựng Kế hoạch Pellet Line (Phân chia theo PL1 -> PL7 và Mash)
        # Lọc sequence theo line_cv
        pl_plans = {f'PL{i}': [] for i in range(1, 8)}
        pl_plans['MASH'] = []
        
        for item in sequence:
            line = item['line_cv'].upper().replace(' ', '')
            if 'PL1' in line: pl_plans['PL1'].append(item)
            elif 'PL2' in line: pl_plans['PL2'].append(item)
            elif 'PL3' in line: pl_plans['PL3'].append(item)
            elif 'PL4' in line: pl_plans['PL4'].append(item)
            elif 'PL5' in line: pl_plans['PL5'].append(item)
            elif 'PL6' in line: pl_plans['PL6'].append(item)
            elif 'PL7' in line: pl_plans['PL7'].append(item)
            elif 'MASH' in line or item['packing_size'] == 'M': pl_plans['MASH'].append(item)
            
        # 3. Kế hoạch Đóng bao (Packaging Matrix)
        packaging_list = []
        for r in range(7, 42): # Dòng 7 đến 41
            prod = ws.cell(row=r, column=2).value # B
            if not prod: continue
            prod_code = str(prod).strip().upper()
            if prod_code in {'TỔNG CỘNG', 'TOTAL', ''}: continue
            
            # Đọc các cột bao bì và silo truck bằng _safe_float_num
            higro_25 = _safe_float_num(ws.cell(row=r, column=5).value)
            higro_40 = _safe_float_num(ws.cell(row=r, column=6).value)
            cp_25 = _safe_float_num(ws.cell(row=r, column=7).value)
            cp_40 = _safe_float_num(ws.cell(row=r, column=8).value)
            star_25 = _safe_float_num(ws.cell(row=r, column=9).value)
            star_40 = _safe_float_num(ws.cell(row=r, column=10).value)
            nuvo_25 = _safe_float_num(ws.cell(row=r, column=11).value)
            nuvo_40 = _safe_float_num(ws.cell(row=r, column=12).value)
            bell_25 = _safe_float_num(ws.cell(row=r, column=13).value)
            bell_40 = _safe_float_num(ws.cell(row=r, column=14).value)
            nasa_25 = _safe_float_num(ws.cell(row=r, column=15).value)
            nasa_40 = _safe_float_num(ws.cell(row=r, column=16).value)
            white_25 = _safe_float_num(ws.cell(row=r, column=17).value)
            white_40 = _safe_float_num(ws.cell(row=r, column=18).value)
            white_50 = _safe_float_num(ws.cell(row=r, column=19).value)
            silo_truck = _safe_float_num(ws.cell(row=r, column=20).value)
            
            # Tính toán tons từ công thức nếu openpyxl trả về None/rỗng do chưa mở file
            batches = _safe_int_val(ws.cell(row=r, column=3).value)
            if not batches or batches == '':
                batches = 0
                
            raw_tons = ws.cell(row=r, column=4).value
            if raw_tons is None or raw_tons == '' or (isinstance(raw_tons, str) and raw_tons.startswith('=')):
                if batches > 0:
                    if prod_code.startswith('550') or prod_code.startswith('551') or prod_code == '325F':
                        tons = float(batches) * 8.0
                    else:
                        tons = float(batches) * 8.4
                else:
                    tons = 0.0
            else:
                tons = _safe_float_num(raw_tons)

            # Xác định quy cách đóng gói (packing_size) - Hỗ trợ quy cách ghép Silo + Bao
            has_silo = (silo_truck > 0)
            has_bag_25 = (higro_25 > 0 or cp_25 > 0 or star_25 > 0 or nuvo_25 > 0 or bell_25 > 0 or nasa_25 > 0 or white_25 > 0)
            has_bag_40 = (higro_40 > 0 or cp_40 > 0 or star_40 > 0 or nuvo_40 > 0 or bell_40 > 0 or nasa_40 > 0 or white_40 > 0)
            has_bag_50 = (white_50 > 0)
            
            if has_silo and (has_bag_25 or has_bag_40 or has_bag_50):
                bag_types = []
                if has_bag_25: bag_types.append("Bao 25")
                if has_bag_40: bag_types.append("Bao 40")
                if has_bag_50: bag_types.append("Bao 50")
                packing_size = "Silo + " + " + ".join(bag_types)
            elif has_silo:
                packing_size = 'M'
            elif has_bag_50:
                packing_size = '50'
            elif has_bag_40:
                packing_size = '40'
            else:
                packing_size = '25'

            line_pk = str(ws.cell(row=r, column=23).value or '').strip()
            # Khôi phục line_pk từ feedcode_mapping nếu trống
            if not line_pk or line_pk == 'None':
                line_pk = feedcode_mapping.get(prod_code, {}).get('line_pk', '')
            if has_silo and (not line_pk or line_pk == 'None'):
                line_pk = 'SILO'
            
            row_data = {
                'product_code': prod_code,
                'tons': tons,
                'packing_size': packing_size,
                'line_pk': line_pk,
                'higro_25': higro_25,
                'cp_25': cp_25,
                'star_25': star_25,
                'nuvo_25': nuvo_25,
                'nasa_25': nasa_25,
                'bell_25': bell_25,
                'higro_40': higro_40,
                'cp_40': cp_40,
                'star_40': star_40,
                'nuvo_40': nuvo_40,
                'nasa_40': nasa_40,
                'bell_40': bell_40,
                'white_25': white_25,
                'white_40': white_40,
                'white_50': white_50,
                'silo_truck': silo_truck,
            }
            
            # Chỉ lấy các dòng có phân bổ đóng bao hoặc silo (tổng các cột > 0)
            total_pack = (higro_25 + cp_25 + star_25 + nuvo_25 + nasa_25 + bell_25 +
                          higro_40 + cp_40 + star_40 + nuvo_40 + nasa_40 + bell_40 +
                          white_25 + white_40 + white_50 + silo_truck)
            if total_pack > 0:
                packaging_list.append(row_data)
                
        # Parse KẾ HOẠCH PL sheet if present
        pl_rich_plans = {}
        mash_plan_rich = []
        mixer_summary_rich = []
        
        if 'KẾ HOẠCH PL' in wb.sheetnames:
            try:
                pl_sheet = wb['KẾ HOẠCH PL']
                machine_cols = {
                    'PL1': 2, 'PL2': 6, 'PL3': 10, 'PL4': 14, 'PL5': 18, 'PL6': 22, 'PL7': 26
                }
                
                # Tồn Đầu (rows 3 and 4)
                code_mapping = {}
                if os.path.exists(config.PLAN_FILE):
                    try:
                        code_mapping = data_loader.load_code_mapping(config.PLAN_FILE)
                        # Bổ sung từ CONG SUAT — override Code sheet vì congsuat là nguồn chính xác nhất
                        congsuat_data = data_loader.load_congsuat(config.PLAN_FILE)
                        for _k, spec in congsuat_data.items():
                            pc = str(spec.product_code).strip().upper()
                            pn = str(spec.product_name).strip().upper()
                            if pc and pn and pc != pn:
                                code_mapping[pc] = pn
                    except Exception as e:
                        print(f"⚠️ Lỗi load code_mapping: {e}")
                        
                ton_dau = {f"PL{i}": [] for i in range(1, 8)}
                for r in (3, 4):
                    for m, col in machine_cols.items():
                        code = pl_sheet.cell(row=r, column=col).value
                        tons = pl_sheet.cell(row=r, column=col+2).value
                        hours = pl_sheet.cell(row=r, column=col+3).value
                        if code:
                            raw_str = str(code).strip().upper()
                            if raw_str.endswith('.0'):
                                raw_str = raw_str[:-2]
                            norm_code = data_loader._normalize_product_code(raw_str)
                            prod_code = code_mapping.get(norm_code, norm_code)
                            
                            ks_code = ks_mapping.get(prod_code, 'SẠCH (KHÔNG KS)')
                            ton_dau[m].append({
                                'product_code': prod_code,
                                'tons': _safe_float_num(tons),
                                'hours': _safe_float_num(hours),
                                'ks_code': ks_code
                            })

                            
                # Ca 1 (rows 5 to 8)
                ca1 = {f"PL{i}": [] for i in range(1, 8)}
                for r in range(5, 9):
                    for m, col in machine_cols.items():
                        code = pl_sheet.cell(row=r, column=col).value
                        batches = pl_sheet.cell(row=r, column=col+1).value
                        tons = pl_sheet.cell(row=r, column=col+2).value
                        hours = pl_sheet.cell(row=r, column=col+3).value
                        if code:
                            prod_code = str(code).strip().upper()
                            ks_code = ks_mapping.get(prod_code, 'SẠCH (KHÔNG KS)')
                            ca1[m].append({
                                'product_code': prod_code,
                                'batches': _safe_int_val(batches),
                                'tons': _safe_float_num(tons),
                                'hours': _safe_float_num(hours),
                                'ks_code': ks_code
                            })
                            
                # Ca 2 (rows 9 to 12)
                ca2 = {f"PL{i}": [] for i in range(1, 8)}
                for r in range(9, 13):
                    for m, col in machine_cols.items():
                        code = pl_sheet.cell(row=r, column=col).value
                        batches = pl_sheet.cell(row=r, column=col+1).value
                        tons = pl_sheet.cell(row=r, column=col+2).value
                        hours = pl_sheet.cell(row=r, column=col+3).value
                        if code:
                            prod_code = str(code).strip().upper()
                            ks_code = ks_mapping.get(prod_code, 'SẠCH (KHÔNG KS)')
                            ca2[m].append({
                                'product_code': prod_code,
                                'batches': _safe_int_val(batches),
                                'tons': _safe_float_num(tons),
                                'hours': _safe_float_num(hours),
                                'ks_code': ks_code
                            })
                            
                # Ca 3 (rows 13 to 16)
                ca3 = {f"PL{i}": [] for i in range(1, 8)}
                for r in range(13, 17):
                    for m, col in machine_cols.items():
                        code = pl_sheet.cell(row=r, column=col).value
                        batches = pl_sheet.cell(row=r, column=col+1).value
                        tons = pl_sheet.cell(row=r, column=col+2).value
                        hours = pl_sheet.cell(row=r, column=col+3).value
                        if code:
                            prod_code = str(code).strip().upper()
                            ks_code = ks_mapping.get(prod_code, 'SẠCH (KHÔNG KS)')
                            ca3[m].append({
                                'product_code': prod_code,
                                'batches': _safe_int_val(batches),
                                'tons': _safe_float_num(tons),
                                'hours': _safe_float_num(hours),
                                'ks_code': ks_code
                            })
                            
                # Loss (row 18)
                loss = {}
                for m, col in machine_cols.items():
                    val = pl_sheet.cell(row=18, column=col+3).value
                    loss[m] = _safe_float_num(val)
                    
                # Next Day (row 19)
                next_day = {}
                for m, col in machine_cols.items():
                    code = pl_sheet.cell(row=19, column=col).value
                    tons = pl_sheet.cell(row=19, column=col+2).value
                    if code:
                        next_day[m] = {
                            'product_code': str(code).strip().upper(),
                            'tons': _safe_float_num(tons)
                        }
                    else:
                        next_day[m] = None
                        
                # Actual Hours (row 23) & Tons (row 24)
                actual_hours = {}
                actual_tons = {}
                for m, col in machine_cols.items():
                    val_h = pl_sheet.cell(row=23, column=col).value
                    val_t = pl_sheet.cell(row=24, column=col).value
                    actual_hours[m] = _safe_float_num(val_h)
                    actual_tons[m] = _safe_float_num(val_t)
                    
                # Format into rich PL plans
                for m in range(1, 8):
                    m_key = f"PL{m}"
                    pl_rich_plans[m_key] = {
                        'ton_dau': ton_dau[m_key],
                        'ca_1': ca1[m_key],
                        'ca_2': ca2[m_key],
                        'ca_3': ca3[m_key],
                        'loss_hours': loss[m_key],
                        'next_day': next_day[m_key],
                        'actual_hours': actual_hours[m_key],
                        'actual_tons': actual_tons[m_key]
                    }
                    
                # Mash Feed (AD to AH, rows 3 to 45)
                for r in range(3, 46):
                    pkg = pl_sheet.cell(row=r, column=30).value
                    bao = pl_sheet.cell(row=r, column=31).value
                    code = pl_sheet.cell(row=r, column=32).value
                    batches = pl_sheet.cell(row=r, column=33).value
                    tons = pl_sheet.cell(row=r, column=34).value
                    if code:
                        mash_plan_rich.append({
                            'packing_size': str(pkg).strip() if pkg else '',
                            'bag_count': str(bao).strip() if bao else '',
                            'product_code': str(code).strip().upper(),
                            'batches': _safe_int_val(batches),
                            'tons': _safe_float_num(tons)
                        })
                        
                # Mixer Summary (AI, rows 3 to 5)
                for r in range(3, 6):
                    val = pl_sheet.cell(row=r, column=35).value
                    if val:
                        mixer_summary_rich.append(str(val).strip())
                        
                # Raw Excel grid for exact table rendering (Rows 1 to 26, Columns 1 to 35)
                khpl_raw_grid = []
                for r in range(1, 27):
                    row_cells = []
                    for c in range(1, 36):
                        cell_val = pl_sheet.cell(row=r, column=c).value
                        row_cells.append(cell_val if cell_val is not None else '')
                    khpl_raw_grid.append(row_cells)
                
                # Dịch digit code → tên cám dân dã trong các ô sản phẩm
                # Cột chứa mã cám: 2(PL1), 6(PL2), 10(PL3), 14(PL4), 18(PL5), 22(PL6), 26(PL7), 30(MASH), 32(MASH cột phụ)
                product_cols = [1, 5, 9, 13, 17, 21, 25, 29, 31]  # 0-indexed
                for row_idx in range(2, min(len(khpl_raw_grid), 22)):  # Row 3 đến 22 (0-indexed: 2..21)
                    for col_idx in product_cols:
                        if col_idx < len(khpl_raw_grid[row_idx]):
                            cell_val = khpl_raw_grid[row_idx][col_idx]
                            if cell_val and cell_val != '':
                                raw_str = str(cell_val).strip()
                                # Loại bỏ đuôi .0 nếu Excel đọc float
                                if raw_str.endswith('.0'):
                                    raw_str = raw_str[:-2]
                                norm = raw_str.upper().replace(' ', '')
                                translated = code_mapping.get(norm, None)
                                if translated and translated != norm:
                                    khpl_raw_grid[row_idx][col_idx] = translated
            except Exception as pl_ex:
                print(f"⚠️ Lỗi khi parse KẾ HOẠCH PL sheet: {pl_ex}")
                traceback.print_exc()
                
        wb.close()
        
        return jsonify({
            'success': True,
            'summary': summary,
            'sequence': sequence,
            'pl_plans': pl_plans,
            'pl_rich_plans': pl_rich_plans,
            'mash_plan_rich': mash_plan_rich,
            'mixer_summary_rich': mixer_summary_rich,
            'khpl_raw_grid': khpl_raw_grid if 'khpl_raw_grid' in locals() else [],
            'packaging': packaging_list,
            'filename': os.path.basename(file_path)
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# API: TẢI EXCEL KẾT QUẢ VỀ MÁY
# ============================================================

@app.route('/api/download-plan/<filename>', methods=['GET'])
def download_plan(filename):
    """Tải trực tiếp file Excel KHSX động (.xlsx) vừa được tạo ra"""
    try:
        # Làm sạch tên file
        clean_name = os.path.basename(filename)
        return send_from_directory(config.OUTPUT_DIR, clean_name, as_attachment=True)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ============================================================
# KHỞI CHẠY SERVER
# ============================================================

# ============================================================
# API: ĐỒNG BỘ THỦ CÔNG EXCEL LÊN DATABASE CLOUD (NEON TECH)
# ============================================================

@app.route('/api/sync-to-db', methods=['POST'])
def sync_to_db():
    """Đồng bộ thủ công toàn bộ file Excel cục bộ lên Neon Tech Cloud Database"""
    try:
        import db_manager
        print("\n☁️ Bắt đầu đồng bộ thủ công dữ liệu Excel lên Cloud Neon Tech...")
        db_manager.sync_local_to_db(config, getattr(config, 'DB_URI', db_manager.DB_URI))
        return jsonify({'success': True, 'message': 'Đồng bộ toàn bộ dữ liệu Excel lên Neon Tech thành công!'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Lỗi khi đồng bộ lên Neon Tech: {str(e)}'})


@app.route('/api/sync-to-db/<category>', methods=['POST'])
def sync_category(category):
    """Đồng bộ nhanh duy nhất 1 danh mục dữ liệu từ SharePoint lên Neon Tech Cloud Database"""
    try:
        import db_manager
        print(f"\n☁️ Bắt đầu đồng bộ nhanh danh mục: {category.upper()} lên Neon Tech...")
        res = db_manager.sync_category_to_db(config, category, getattr(config, 'DB_URI', db_manager.DB_URI))
        return jsonify({
            'success': True,
            'message': f"Đồng bộ thành công danh mục {category.upper()} lên Cloud (File mới nhất: {res['filename']})!",
            'filename': res['filename'],
            'last_modified': res['last_modified']
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f"Lỗi đồng bộ danh mục {category.upper()}: {str(e)}"})


if __name__ == '__main__':
    # Đảm bảo thư mục lưu trữ tồn tại
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    os.makedirs(TEMP_UPLOAD_DIR, exist_ok=True)
    
    # Khởi tạo database PostgreSQL (Neon Tech) khi khởi chạy server
    if getattr(config, 'USE_POSTGRESQL', False):
        try:
            import db_manager
            db_manager.init_db(getattr(config, 'DB_URI', db_manager.DB_URI))
        except Exception as db_init_err:
            print(f"⚠️ Không thể kết nối khởi tạo Neon Tech DB: {db_init_err}")
            
    # Chạy Flask Server
    app.run(host='0.0.0.0', port=5000, debug=True)
